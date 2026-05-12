"""
oos_functions.py — модуль «Отчёт без OOS» для EFKO FlowManager.
Поддерживает три вида майонезного отчёта:
  'sloboda'    — Слобода  200/400/800
  'provansale' — Провансаль 200/400/800 + конкуренты
  'olive'      — Оливковый 200/400/800
"""

import re
import pandas as pd
import polars as pl
from pathlib import Path
from datetime import date

# ─── Парсинг русских месяцев ──────────────────────────────────────────────
_RU_FULL = {
    "январь": 1,
    "февраль": 2,
    "март": 3,
    "апрель": 4,
    "май": 5,
    "июнь": 6,
    "июль": 7,
    "август": 8,
    "сентябрь": 9,
    "октябрь": 10,
    "ноябрь": 11,
    "декабрь": 12,
}
_RU_SHORT = {
    "янв": 1,
    "фев": 2,
    "мар": 3,
    "апр": 4,
    "май": 5,
    "июн": 6,
    "июл": 7,
    "авг": 8,
    "сен": 9,
    "окт": 10,
    "ноя": 11,
    "дек": 12,
}


def _parse_ru_month(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    s = str(v).strip().lower()
    m = re.match(r"^([а-яё]+)\s+(\d{4})$", s)
    if m:
        mon = _RU_FULL.get(m.group(1))
        if mon:
            return date(int(m.group(2)), mon, 1)
    m = re.match(r"^([а-яё]+)\.(\d{2,4})$", s)
    if m:
        abbr, yr = m.group(1), int(m.group(2))
        mon = _RU_SHORT.get(abbr) or _RU_FULL.get(abbr)
        if mon:
            return date(yr + (2000 if yr < 100 else 0), mon, 1)
    return None


def _parse_month_series(series: pl.Series) -> pl.Series:
    return pl.Series(
        series.name, [_parse_ru_month(v) for v in series.to_list()], dtype=pl.Date
    )


# ─── Конвертация pandas ↔ polars без pyarrow ──────────────────────────────
def _pd2pl(df: pd.DataFrame) -> pl.DataFrame:
    result = {}
    for col in df.columns:
        name = str(col)
        s = df[col]
        dtype_str = str(s.dtype)
        try:
            if hasattr(s.dtype, "numpy_dtype"):
                result[name] = pl.Series(
                    name, s.to_numpy(dtype=s.dtype.numpy_dtype, na_value=None)
                )
            elif dtype_str == "object" or dtype_str.startswith("datetime"):
                result[name] = pl.Series(name, s.tolist())
            elif dtype_str in ("int64", "int32", "int16", "int8") and s.isna().any():
                result[name] = pl.Series(name, s.astype(float).tolist())
            else:
                result[name] = pl.Series(name, s.to_numpy(na_value=None))
        except Exception:
            try:
                result[name] = pl.Series(name, s.tolist(), strict=False)
            except Exception:
                result[name] = pl.Series(
                    name, [None if (v != v) else v for v in s.tolist()]
                )
    return pl.DataFrame(result)


def _pl2pd(df: pl.DataFrame) -> pd.DataFrame:
    result = {}
    for col in df.columns:
        s = df[col]
        if s.dtype == pl.Date:
            result[col] = [
                v.isoformat() if v is not None else None for v in s.to_list()
            ]
        else:
            result[col] = s.to_list()
    return pd.DataFrame(result)


# ─── Чтение листов ────────────────────────────────────────────────────────
def _read_sheet(
    filepath: Path, sheet: str, engine: str, header_row: int
) -> pl.DataFrame:
    raw = pd.read_excel(filepath, sheet_name=sheet, header=None, engine=engine)
    if header_row >= len(raw):
        raise ValueError(f"header_row={header_row} за границей листа '{sheet}'")
    headers = raw.iloc[header_row].tolist()
    df = raw.iloc[header_row + 1 :].copy()
    df.columns = headers
    valid = [
        c
        for c in df.columns
        if c is not None and not (isinstance(c, float) and pd.isna(c))
    ]
    if not valid:
        raise ValueError(f"Лист '{sheet}': нет заголовков в строке {header_row}")
    return _pd2pl(df[valid].dropna(how="all").reset_index(drop=True))


def _read_kub(kub: Path, sheet: str, header_row: int) -> pl.DataFrame:
    return _read_sheet(kub, sheet, "openpyxl", header_row)


def _read_elt(elt: Path, sheet: str, header_row: int = 7) -> pl.DataFrame:
    return _read_sheet(elt, sheet, "openpyxl", header_row)


# ─── Общие запросы ────────────────────────────────────────────────────────
def _query_asst(kub: Path) -> pl.DataFrame:
    df = _read_kub(kub, "асс-т", 8)
    df = df.with_columns(_parse_month_series(df["Месяц"]))
    return df.rename({"Продажи кг; л": "Емкость, кг.,л."})


def _query_emkost(kub: Path) -> pl.DataFrame:
    df = _read_kub(kub, "емкость сети", 9)
    df = df.with_columns(
        [
            pl.col("Сеть UNI").cast(pl.Utf8),
            pl.col("Формат").cast(pl.Utf8),
            pl.col("Ф_ТТ присутствия").cast(pl.Int64),
            pl.col("Продажи кг; л").cast(pl.Float64),
        ]
    )
    df = df.with_columns(_parse_month_series(df["Месяц"]))
    df = df.with_columns(
        (pl.col("Сеть UNI") + pl.col("Формат")).alias("сеть&формат"),
        (pl.col("Продажи кг; л") / 1000).alias("Емкость сети, тн."),
    )
    return df.select(
        [
            "Сеть UNI",
            "Формат",
            "сеть&формат",
            "Год",
            "Месяц",
            "Ф_ТТ присутствия",
            "Емкость сети, тн.",
        ]
    )


def _query_elt_status(elt: Path, sheet: str) -> pl.DataFrame:
    """Выгрузка ELT → (Адрес, Месяц, Статус ТТ(1))"""
    df = _read_elt(elt, sheet, 7)
    df = df.with_columns(_parse_month_series(df["Месяц"]))
    df = df.with_columns(
        [
            pl.col("Адрес").cast(pl.Utf8),
            pl.when(pl.col("Итог").cast(pl.Float64) == 1.0)
            .then(pl.lit("не берем"))
            .otherwise(pl.lit("берем"))
            .alias("Статус"),
        ]
    )
    g = df.group_by(["Адрес", "Месяц", "Статус"]).agg(pl.len().alias("Количество"))
    piv = g.pivot(
        values="Количество",
        index=["Адрес", "Месяц"],
        on="Статус",
        aggregate_function="sum",
    )
    for c in ("берем", "не берем"):
        if c not in piv.columns:
            piv = piv.with_columns(pl.lit(None).cast(pl.Int64).alias(c))
    return piv.with_columns(
        pl.when(
            (pl.col("берем") > 3)
            & (pl.col("не берем").is_null() | (pl.col("не берем") == 0))
        )
        .then(pl.lit("берем"))
        .otherwise(pl.lit("не берем"))
        .alias("Статус ТТ(1)")
    ).select(["Адрес", "Месяц", "Статус ТТ(1)"])


def _query_schet_tt(elt: Path, sheet: str) -> pl.DataFrame:
    """Счётчик ТТ ELT → (Сеть&формат, Месяц, Количество)"""
    df = _read_elt(elt, sheet, 7)
    df = df.with_columns(_parse_month_series(df["Месяц"]))
    df = df.with_columns(
        (pl.col("Торговая сеть").cast(pl.Utf8) + pl.col("Формат").cast(pl.Utf8)).alias(
            "Сеть&формат"
        )
    )
    df = df.select(["Адрес", "Сеть&формат", "Месяц"]).unique()
    df = df.group_by(["Сеть&формат", "Месяц"]).agg(pl.len().alias("Количество"))
    df = df.with_columns(pl.col("Сеть&формат").str.replace("_МТТ", "", literal=True))
    return df.group_by(["Сеть&формат", "Месяц"]).agg(pl.col("Количество").sum())


def _query_kub(
    kub: Path, sheet: str, header_row: int, elt_status: pl.DataFrame, asst: pl.DataFrame
) -> pl.DataFrame:
    df = _read_kub(kub, sheet, header_row)
    df = df.with_columns(_parse_month_series(df["Месяц"]))
    df = df.with_columns(
        [
            pl.col("Сеть UNI").cast(pl.Utf8),
            pl.col("Формат").cast(pl.Utf8),
            pl.col("Адрес").cast(pl.Utf8),
            (pl.col("Сеть UNI").cast(pl.Utf8) + pl.col("Формат").cast(pl.Utf8)).alias(
                "Сеть&формат"
            ),
        ]
    )
    for c in ("Ф_Ин-ть", "Ф Цена продажи с НДС, руб/шт"):
        if c in df.columns:
            df = df.with_columns(pl.col(c).cast(pl.Float64, strict=False))
    if "Ф_SKUср факт" in df.columns:
        df = df.with_columns(
            pl.col("Ф_SKUср факт")
            .cast(pl.Float64, strict=False)
            .round(0)
            .cast(pl.Int64, strict=False)
        )
    df = df.join(
        elt_status.select(["Адрес", "Месяц", "Статус ТТ(1)"]),
        on=["Адрес", "Месяц"],
        how="left",
    )
    df = df.filter(pl.col("Статус ТТ(1)") == "берем")
    asst_cols = ["Адрес", "Месяц"] + [
        c for c in ("Ф_ SKUcрТТприс", "Емкость, кг.,л.") if c in asst.columns
    ]
    return df.join(asst.select(asst_cols), on=["Адрес", "Месяц"], how="inner")


def _query_itogo(
    kub: Path,
    sheet: str,
    emkost: pl.DataFrame,
    schet_tt: pl.DataFrame,
    filter_type: str,
    formats: set,
) -> pl.DataFrame:
    """filter_type: 'include' | 'exclude'"""
    df = _read_kub(kub, sheet, 9)
    df = df.with_columns(_parse_month_series(df["Месяц"]))
    df = df.with_columns(
        [
            pl.col("Сеть UNI").cast(pl.Utf8),
            pl.col("Формат").cast(pl.Utf8),
            pl.col("Ф_Ин-ть").cast(pl.Float64),
            pl.col("Ф_ТТ присутствия").cast(pl.Int64),
            pl.col("Продажи кг; л").cast(pl.Float64),
            (pl.col("Сеть UNI").cast(pl.Utf8) + pl.col("Формат").cast(pl.Utf8)).alias(
                "Сеть&формат"
            ),
        ]
    )
    df = df.select(
        [
            "Сеть UNI",
            "Формат",
            "Сеть&формат",
            "Год",
            "Месяц",
            "Ф_Ин-ть",
            "Ф_ТТ присутствия",
            "Продажи кг; л",
        ]
    )
    df = df.join(
        emkost.rename({"сеть&формат": "Сеть&формат"}).select(
            ["Сеть&формат", "Месяц", "Емкость сети, тн."]
        ),
        on=["Сеть&формат", "Месяц"],
        how="left",
    )
    df = df.join(
        schet_tt.rename({"Количество": "Количество посещаемых ТТ"}),
        on=["Сеть&формат", "Месяц"],
        how="left",
    )
    if filter_type == "include":
        df = df.filter(pl.col("Формат").is_in(formats))
    else:
        df = df.filter(~pl.col("Формат").is_in(formats))
    return df


def _query_competitor(
    kub: Path, sheet: str, header_row: int, elt_status: pl.DataFrame
) -> pl.DataFrame:
    df = _read_kub(kub, sheet, header_row)
    df = df.with_columns(_parse_month_series(df["Месяц"]))
    df = df.with_columns(
        [
            pl.col("Сеть UNI").cast(pl.Utf8),
            pl.col("Формат").cast(pl.Utf8),
            pl.col("Адрес").cast(pl.Utf8),
        ]
    )
    if "Ф Цена продажи с НДС, руб/шт" in df.columns:
        df = df.with_columns(
            pl.col("Ф Цена продажи с НДС, руб/шт").cast(pl.Float64, strict=False)
        )
    df = df.join(
        elt_status.select(["Адрес", "Месяц", "Статус ТТ(1)"]),
        on=["Адрес", "Месяц"],
        how="left",
    )
    df = df.filter(pl.col("Статус ТТ(1)") == "берем")
    return df.with_columns((pl.col("Сеть UNI") + pl.col("Формат")).alias("Сеть&формат"))


# ─── Сохранение в Excel ───────────────────────────────────────────────────
_FORMATS_4 = {"Гипермаркет", "Дискаунтер", "Минимаркет/Дискаунтер", "Супермаркет"}
_FORMATS_EX_P200 = {"Дакстор", "Интернет-магазин", "Эконом"}


def _save_to_excel(report_file: Path, sheets_dict: dict, log):
    from openpyxl import load_workbook, Workbook

    if report_file.exists():
        try:
            wb = load_workbook(report_file)
            log(f"  Открыт шаблон: {report_file.name}")
        except Exception as e:
            log(f"  [WARN] Шаблон не читается ({e}), создаём новый")
            wb = Workbook()
            wb.remove(wb.active)
    else:
        log("  Создаём новый файл")
        wb = Workbook()
        wb.remove(wb.active)

    for sheet_name, pl_df in sheets_dict.items():
        pdf = _pl2pd(pl_df)
        pos = (
            wb.sheetnames.index(sheet_name)
            if sheet_name in wb.sheetnames
            else len(wb.sheetnames)
        )
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name, pos)
        ws.append(list(pdf.columns))
        for row in pdf.itertuples(index=False, name=None):
            ws.append([None if (isinstance(v, float) and v != v) else v for v in row])

    visible = [ws for ws in wb.worksheets if ws.sheet_state == "visible"]
    if not visible:
        wb.worksheets[0].sheet_state = "visible"
    wb.save(report_file)


# ─── Главная функция ──────────────────────────────────────────────────────


def run_oos_report(
    report_type: str, kub_file: str, elt_file: str, report_file: str, log, stop_event
):
    """
    report_type: 'sloboda' | 'provansale' | 'olive'
    kub_file, elt_file, report_file — строки с путями
    log — callable(str)
    stop_event — threading.Event
    """
    kub = Path(kub_file)
    elt = Path(elt_file)
    rep = Path(report_file)

    def chk():
        if stop_event.is_set():
            raise InterruptedError("Остановлено пользователем")

    log("📖 Читаем асс-т + емкость сети...")
    asst = _query_asst(kub)
    chk()
    emkost = _query_emkost(kub)
    chk()
    log(f"  асс-т: {len(asst)} строк, емкость: {len(emkost)} строк")

    # ── Слобода ──
    if report_type == "sloboda":
        log("📋 Слобода: ELT выгрузки...")
        e200 = _query_elt_status(elt, "200")
        chk()
        e400 = _query_elt_status(elt, "400")
        chk()
        e800 = _query_elt_status(elt, "800")
        chk()
        s200 = _query_schet_tt(elt, "200")
        chk()
        s400 = _query_schet_tt(elt, "400")
        chk()
        s800 = _query_schet_tt(elt, "800")
        chk()
        log("📋 Слобода: КУБ + итого сети...")
        k200 = _query_kub(kub, "200", 9, e200, asst)
        chk()
        k400 = _query_kub(kub, "400", 9, e400, asst)
        chk()
        k800 = _query_kub(kub, "800", 9, e800, asst)
        chk()
        i200 = _query_itogo(kub, "итого сеть 200", emkost, s200, "include", _FORMATS_4)
        chk()
        i400 = _query_itogo(kub, "итого сеть 400", emkost, s400, "include", _FORMATS_4)
        chk()
        i800 = _query_itogo(kub, "итого сеть 800", emkost, s800, "include", _FORMATS_4)
        chk()
        sheets = {
            "200": k200,
            "400": k400,
            "800": k800,
            "итого сеть 200": i200,
            "итого сеть 400": i400,
            "итого сеть 800": i800,
        }

    # ── Провансаль ──
    elif report_type == "provansale":
        log("📋 Провансаль: ELT выгрузки...")
        e200 = _query_elt_status(elt, "провансаль 200")
        chk()
        e400 = _query_elt_status(elt, "провансаль 400")
        chk()
        e800 = _query_elt_status(elt, "провансаль 800")
        chk()
        s200 = _query_schet_tt(elt, "провансаль 200")
        chk()
        s400 = _query_schet_tt(elt, "провансаль 400")
        chk()
        s800 = _query_schet_tt(elt, "провансаль 800")
        chk()
        log("📋 Провансаль: КУБ + итого сети...")
        k200 = _query_kub(kub, "провансаль 200", 8, e200, asst)
        chk()
        k400 = _query_kub(kub, "провансаль 400", 8, e400, asst)
        chk()
        k800 = _query_kub(kub, "провансаль 800", 8, e800, asst)
        chk()
        i200 = _query_itogo(
            kub, "итого сеть П200", emkost, s200, "exclude", _FORMATS_EX_P200
        )
        chk()
        i400 = _query_itogo(kub, "итого сеть П400", emkost, s400, "include", _FORMATS_4)
        chk()
        i800 = _query_itogo(kub, "итого сеть П800", emkost, s800, "include", _FORMATS_4)
        chk()
        log("📋 Конкуренты...")
        maheev = _query_competitor(kub, "Махеев", 8, e400)
        chk()
        mr_ricco = _query_competitor(kub, "Mr Ricco", 9, e400)
        chk()
        ryaba = _query_competitor(kub, "Ряба", 9, e400)
        chk()
        yalg = _query_competitor(kub, "ЯЛГ", 9, e400)
        chk()
        sheets = {
            "200": k200,
            "400": k400,
            "800": k800,
            "итого сеть 200": i200,
            "итого сеть 400": i400,
            "итого сеть 800": i800,
            "Махеев": maheev,
            "Mr.Ricco": mr_ricco,
            "Ряба": ryaba,
            "ЯЛГ": yalg,
        }

    # ── Оливковый ──
    elif report_type == "olive":
        log("📋 Оливковый: ELT выгрузки...")
        e200 = _query_elt_status(elt, "оливковый 200")
        chk()
        e400 = _query_elt_status(elt, "оливковый 400")
        chk()
        e800 = _query_elt_status(elt, "оливковый 800")
        chk()
        s200 = _query_schet_tt(elt, "оливковый 200")
        chk()
        s400 = _query_schet_tt(elt, "оливковый 400")
        chk()
        s800 = _query_schet_tt(elt, "оливковый 800")
        chk()
        log("📋 Оливковый: КУБ + итого сети...")
        k200 = _query_kub(kub, "оливковый 200", 8, e200, asst)
        chk()
        k400 = _query_kub(kub, "оливковый 400", 8, e400, asst)
        chk()
        k800 = _query_kub(kub, "оливковый 800", 8, e800, asst)
        chk()
        i200 = _query_itogo(kub, "итого сеть О200", emkost, s200, "include", _FORMATS_4)
        chk()
        i400 = _query_itogo(kub, "итого сеть О400", emkost, s400, "include", _FORMATS_4)
        chk()
        i800 = _query_itogo(kub, "итого сеть О800", emkost, s800, "include", _FORMATS_4)
        chk()
        sheets = {
            "200": k200,
            "400": k400,
            "800": k800,
            "итого сеть 200": i200,
            "итого сеть 400": i400,
            "итого сеть 800": i800,
        }
    else:
        raise ValueError(f"Неизвестный тип отчёта: {report_type}")

    log(f"💾 Сохраняем в {rep.name}...")
    _save_to_excel(rep, sheets, log)

    log("✅ Готово!")
    for name, df in sheets.items():
        log(f"  {name:<22} {len(df):>6} строк")
