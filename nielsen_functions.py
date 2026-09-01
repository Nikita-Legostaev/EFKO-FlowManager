# nielsen_functions.py
"""
Обработчик выгрузок Nielsen.
Полностью открыт для расширения: никаких хардкодов имён листов,
id-колонок или маппингов. Маппинги загружаются из папки Справочники
автоматически и применяются ко всем категориям.
"""

import os
import re
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

# Паттерн даты в заголовках Nielsen ("MAR 21", "APR 22", "Mar 2023", "Apr 2024")
_DATE_COL_RE = re.compile(r"^[A-Z]{3}\s+\d{2,4}$", re.IGNORECASE)

# ── Путь к папке со справочниками ─────────────────────────────────────────────
# Папка рядом с nielsen_functions.py или рядом с exe после сборки.
# Можно переопределить из конфига.
СПРАВОЧНИКИ_FOLDER: str = str(Path(__file__).parent / "Справочники")

# ── Допустимые значения FACT ───────────────────────────────────────────────────
valid_FACT = [
    "Units (in 1000 PACKS)",
    "Volume (in 1000 LTR)",
    "Volume (in 1000 KG)",
    "Value (in 1000 RUR)",
    "Price (Unit)",
    "Price (Volume)",
    "Weighted Distribution (C)",
    "Weighted Distribution (w)",
    "Numeric Distribution (C)",
    "Numeric Distribution (w)",
    "Volume (in 1000 )",
    "Volume (in 1000 LTR/KG)",
    "Value (in 1000)",
]

# ── Описание JOIN-таблиц из справочников ──────────────────────────────────────
# key_col    — колонка в данных Nielsen (левый ключ JOIN)
# sheet_key  — заголовок первой колонки в xlsx-листе (для распознавания)
# aliases    — вариации заголовка первой колонки (разные файлы могут называть по-разному)
# added_cols — какие колонки добавляем (None = все остальные колонки листа)
_JOIN_DEFS = [
    {"key_col": "ATTRIBUTE",        "sheet_keys": ["Период"],                 "added_cols": ["Период скорр", "Год", "Квартал"]},
    {"key_col": "MARKET",           "sheet_keys": ["MARKET"],                 "added_cols": None},
    {"key_col": "FACT",             "sheet_keys": ["FACT ", "FACT", "Fact"],  "added_cols": None},
    {"key_col": "IF REFINED",       "sheet_keys": ["IF REFINED"],             "added_cols": None},
    {"key_col": "PRODUCT BASE",     "sheet_keys": ["PRODUCT BASE"],           "added_cols": None},
    {"key_col": "PACKAGE TYPE",     "sheet_keys": ["PACKAGE TYPE", "упаковка"], "added_cols": None},
    {"key_col": "MANUFACTURER",     "sheet_keys": ["MANUFACTURER"],           "added_cols": None},
    {"key_col": "BRAND",            "sheet_keys": ["BRAND", "Brand"],         "added_cols": None},
    {"key_col": "Long Description", "sheet_keys": ["Long Description", "СКЮ", "справочник", "справочник до 1 пол25", "вкус", "Лист1"], "added_cols": None},
    {"key_col": "PRODUCT TYPE",     "sheet_keys": ["Producte type"],          "added_cols": None},
]

# Суффиксы листов → категория Nielsen
_SUFFIX_TO_CATEGORY: dict[str, str] = {
    "_масло":   "Масло",
    "_майонез": "Майонез",
    "_кетчуп":  "Кетчуп",
    "_соус":    "Соусы",
    "_молоко":  "Растительное молоко",
    "_рам":     "РАМ",
    "_рам2":    "РАМ",
}

# Маппинг первой колонки листа → колонка в данных Nielsen
_SHEET_COL_TO_DATA_COL: dict[str, str] = {
    "BRAND":            "BRAND",
    "MANUFACTURER":     "MANUFACTURER",
    "FACT":             "FACT",
    "FACT ":            "FACT",
    "IF REFINED":       "IF REFINED",
    "PRODUCT BASE":     "PRODUCT BASE",
    "PACKAGE TYPE":     "PACKAGE TYPE",
    "MARKET":           "MARKET",
    "PERIOD":           "ATTRIBUTE",
    "ПЕРИОД":           "ATTRIBUTE",
    "Long Description": "Long Description",
    "СКЮ":              "Long Description",
    "SKU":              "Long Description",
    "Producte type":    "PRODUCT TYPE",
    "PRODUCTE TYPE":    "PRODUCT TYPE",
}


def _sheet_category(sheet_name: str) -> str | None:
    """None = общий лист (все категории), иначе — конкретная категория."""
    lower = sheet_name.lower()
    for suffix, cat in _SUFFIX_TO_CATEGORY.items():
        if lower.endswith(suffix):
            return cat
    return None


def _col_to_data(col_src: str) -> str | None:
    key = col_src.strip()
    return _SHEET_COL_TO_DATA_COL.get(key) or _SHEET_COL_TO_DATA_COL.get(key.upper())


def load_join_tables(
    sprav_path: str | None = None,
    category: str = "",
    region_type: str = "ST",
) -> dict[str, object]:
    """
    Читает единый файл справочника и строит JOIN-таблицы для указанной категории.
    Листы без суффикса — общие (все категории).
    Листы с суффиксом _масло/_кетчуп/... — только для своей категории.
    Регионы_ST / Регионы_NU — выбирается по region_type.
    """
    import polars as pl
    import openpyxl

    if not sprav_path or not os.path.isfile(sprav_path):
        return {}

    try:
        wb = openpyxl.load_workbook(sprav_path, read_only=True, data_only=True)
    except Exception:
        return {}

    frames: dict[str, list] = {}

    for sh in wb.sheetnames:
        sh_lower = sh.lower()

        # Фильтр по региону
        if sh_lower.startswith("регионы_"):
            if sh_lower != f"регионы_{region_type.lower()}":
                continue

        # Фильтр по категории
        sheet_cat = _sheet_category(sh)
        if sheet_cat is not None and sheet_cat != category:
            continue

        try:
            ws = wb[sh]
            rows = list(ws.iter_rows(values_only=True))
        except Exception:
            continue
        if len(rows) < 2 or not rows[0] or not rows[0][0]:
            continue

        col_src = str(rows[0][0]).strip()
        data_col = _col_to_data(col_src)
        if not data_col:
            continue

        # Для упаковки — берём только первые 2 непустые колонки
        header = rows[0]
        if "упаковка" in sh_lower:
            non_empty = [(i, str(h).strip()) for i, h in enumerate(header) if h and str(h).strip()]
            keep_idx = [non_empty[0][0], non_empty[1][0]] if len(non_empty) >= 2 else []
        else:
            keep_idx = [i for i, h in enumerate(header) if h and str(h).strip()]

        if len(keep_idx) < 2:
            continue

        col_names = [str(header[i]).strip() for i in keep_idx]
        col_names[0] = data_col

        str_rows = []
        for row in rows[1:]:
            if not row or row[0] is None:
                continue
            vals = []
            for i in keep_idx:
                v = row[i] if i < len(row) else None
                if v is None:
                    vals.append(None)
                elif hasattr(v, "strftime"):
                    vals.append(v.strftime("%d.%m.%Y"))
                else:
                    s = str(v).strip()
                    vals.append(None if s in ("", "None", "nan") else s)
            if vals[0]:
                str_rows.append(vals)

        if not str_rows:
            continue

        try:
            df_sheet = pl.DataFrame(
                {col_names[j]: [r[j] for r in str_rows] for j in range(len(col_names))}
            ).with_columns(pl.col(col_names[0]).cast(pl.Utf8).str.strip_chars())
            frames.setdefault(data_col, []).append(df_sheet)
        except Exception:
            pass

    wb.close()

    result: dict[str, object] = {}
    for data_col, dfs in frames.items():
        try:
            combined = pl.concat(dfs, how="diagonal").unique(subset=[data_col], keep="last")
            combined = combined.with_columns(pl.col(data_col).cast(pl.Utf8).str.strip_chars())
            result[data_col] = combined
        except Exception:
            pass

    return result


def join_sprav(df, join_tables: dict, log) -> object:
    """
    LEFT JOIN данных Nielsen с таблицами из справочников.
    Для каждого join_table:
      - ключ нормализуется (strip + upper) в обоих датафреймах
      - добавляемые колонки не перезаписывают уже существующие в df
    """
    import polars as pl

    for key_col, ref_df in join_tables.items():
        if key_col not in df.columns:
            continue
        # Не добавляем колонки которые уже есть
        new_cols = [c for c in ref_df.columns if c != key_col and c not in df.columns]
        if not new_cols:
            continue

        ref_slim = ref_df.select([key_col] + new_cols)

        # Нормализуем ключ в основном df для JOIN
        tmp_key = f"__join_{key_col}__"
        df = df.with_columns(
            pl.col(key_col).cast(pl.Utf8).str.strip_chars().str.to_uppercase().alias(tmp_key)
        )
        ref_slim = ref_slim.with_columns(
            pl.col(key_col).cast(pl.Utf8).str.strip_chars().str.to_uppercase().alias(tmp_key)
        ).drop(key_col)

        try:
            df = df.join(ref_slim, on=tmp_key, how="left")
            log(f"  JOIN {key_col} → +{new_cols}")
        except Exception as e:
            log(f"  ⚠ JOIN {key_col}: {e}")

        df = df.drop(tmp_key)

    return df

def _max_workers() -> int:
    return min(4, os.cpu_count() or 2)


def _is_date_col(name: str) -> bool:
    """True если колонка — период Nielsen («MAR 21», «APR 22»)."""
    return bool(_DATE_COL_RE.match(str(name).strip()))


def _detect_sheets(wb_sheet_names: list[str]) -> dict[str, list[str]]:
    """
    Автоматически раскладывает листы книги по ролям:
      sku      — листы с данными SKU (все листы, начинающиеся с «SKU»)
      brand    — листы с данными BRAND
      man      — листы с данными MANUFACTURER / MAN
    Регистронезависимо.
    """
    result: dict[str, list[str]] = {"sku": [], "brand": [], "man": []}
    for name in wb_sheet_names:
        nl = name.strip().upper()
        if nl.startswith("SKU"):
            result["sku"].append(name)
        elif nl.startswith("BRAND"):
            result["brand"].append(name)
        elif nl.startswith("MAN"):
            result["man"].append(name)
    return result


def _cache_dir_for(input_file: str, base_cache_dir: Path) -> Path:
    """
    Возвращает уникальную подпапку кэша для конкретного файла.
    Ключ = хэш от абсолютного пути + времени изменения файла.
    При изменении файла или смене категории кэш не смешивается.
    """
    import hashlib
    mtime = os.path.getmtime(input_file)
    key = f"{os.path.abspath(input_file)}|{mtime}"
    digest = hashlib.md5(key.encode()).hexdigest()[:12]
    cache_path = base_cache_dir / digest
    cache_path.mkdir(parents=True, exist_ok=True)
    return cache_path


def _cleanup_old_caches(base_cache_dir: Path, keep: str, log):
    """
    Удаляет все кэш-папки кроме текущей (keep).
    Вызывается после успешного запуска.
    """
    if not base_cache_dir.exists():
        return
    for entry in base_cache_dir.iterdir():
        if entry.is_dir() and entry.name != keep:
            import shutil
            try:
                shutil.rmtree(entry)
                log(f"🗑 Старый кэш удалён: {entry.name}")
            except Exception as e:
                log(f"⚠ Не удалось удалить кэш {entry.name}: {e}")


def make_unique_columns(columns):
    seen = {}
    new_cols = []
    for col in columns:
        col = col or "Unnamed"
        if col in seen:
            seen[col] += 1
            new_cols.append(f"{col}_{seen[col]}")
        else:
            seen[col] = 0
            new_cols.append(col)
    return new_cols


def load_sheet(sheet_name: str, input_file: str, cache_dir: Path, log):
    import polars as pl

    cache_file = cache_dir / f"{sheet_name}.parquet"
    if cache_file.exists():
        log(f"Из кэша → {sheet_name}")
        return sheet_name, pl.read_parquet(cache_file)

    log(f"Читаем Excel → {sheet_name}")
    df = pl.read_excel(
        source=input_file,
        sheet_name=sheet_name,
        engine="calamine",
        infer_schema_length=100_000,
    )
    df = df.rename(dict(zip(df.columns, make_unique_columns(df.columns))))
    df = df.drop([c for c in df.columns if str(c).startswith("Unnamed")])
    df = df.filter(~pl.all_horizontal(pl.all().is_null()))
    df.write_parquet(cache_file, compression="zstd")
    return sheet_name, df


def _split_id_date(df) -> tuple[list[str], list[str]]:
    """
    Делит колонки датафрейма на id-колонки (измерения) и date-колонки (периоды).
    Date-колонки определяются по паттерну «MMM YY».
    """
    id_cols  = [c for c in df.columns if not _is_date_col(c)]
    date_cols = [c for c in df.columns if _is_date_col(c)]
    return id_cols, date_cols


def melt_polars(df, id_cols: list[str]):
    import polars as pl

    date_cols = [c for c in df.columns if c not in id_cols]
    date_mapping = {}
    for col in date_cols:
        cleaned = col.strip().upper()
        parts = cleaned.split()
        if len(parts) == 2 and len(parts[0]) == 3 and parts[1].isdigit():
            year_part = parts[1]
            # Поддержка и двузначного (21 → 2021) и четырёхзначного (2023) года
            if len(year_part) == 2:
                fmt = "%b %y %d"
            else:
                fmt = "%b %Y %d"
            date_str = f"{parts[0]} {year_part} 1"
            parsed = pl.Series([date_str]).str.to_date(fmt, strict=False)[0]
            date_mapping[col] = (
                parsed.strftime("%d.%m.%Y") if parsed is not None else None
            )
        else:
            date_mapping[col] = None

    return (
        df.unpivot(
            index=id_cols, on=date_cols, variable_name="ATTRIBUTE", value_name="VALUE"
        )
        .with_columns(
            pl.col("ATTRIBUTE")
            .replace_strict(date_mapping, default=None)
            .alias("ATTRIBUTE")
        )
        .filter(pl.col("VALUE").is_not_null())
    )


def optimize_for_size(df):
    """
    Оптимизирует датафрейм для сжатия: строковые колонки → Categorical,
    VALUE → Float32. Работает для любого набора колонок.
    """
    import polars as pl

    for col in df.columns:
        if df[col].dtype in (pl.Utf8, pl.String):
            df = df.with_columns(pl.col(col).cast(pl.Categorical("lexical")))

    if "VALUE" in df.columns:
        # VALUE может прийти как String (если polars не угадал тип при melt)
        # → сначала приводим к Float64, потом округляем и сжимаем до Float32
        if df["VALUE"].dtype not in (pl.Float32, pl.Float64, pl.Int32, pl.Int64):
            df = df.with_columns(
                pl.col("VALUE").cast(pl.Categorical).cast(pl.Utf8)
                .str.replace(",", ".", literal=True)
                .cast(pl.Float64, strict=False)
                .alias("VALUE")
            )
        df = df.with_columns(
            pl.col("VALUE").round(4).cast(pl.Float32)
        )

    return df


def pivot_by_fact(df, log) -> object:
    """
    Разворачивает колонку FACT в отдельные столбцы.

    Было (длинный формат):
      BRAND | FACT                   | ATTRIBUTE | VALUE
      X     | Units (in 1000 PACKS)  | 01.03.2021 | 100
      X     | Value (in 1000 RUR)    | 01.03.2021 | 500

    Стало (широкий формат):
      BRAND | ATTRIBUTE  | Units (in 1000 PACKS) | Value (in 1000 RUR)
      X     | 01.03.2021 | 100                   | 500
    """
    import polars as pl

    if "FACT" not in df.columns or "VALUE" not in df.columns:
        return df

    # VALUE → Float64 (из любого типа: Categorical, Utf8, Float32 и т.д.)
    try:
        value_dtype = df["VALUE"].dtype
        if value_dtype in (pl.Categorical, pl.Utf8, pl.String):
            df = df.with_columns(
                pl.col("VALUE").cast(pl.Utf8)
                .str.replace_all(r"\s", "", literal=False)  # убираем пробелы/неразрывные
                .str.replace(",", ".", literal=True)
                .cast(pl.Float64, strict=False)
                .alias("VALUE")
            )
        elif value_dtype == pl.Float32:
            df = df.with_columns(pl.col("VALUE").cast(pl.Float64).alias("VALUE"))
    except Exception as e:
        log(f"  ⚠ pivot_by_fact: не удалось привести VALUE к числу: {e}")
        return df

    # Убираем только строки где FACT пустой (без VALUE — оставляем, null останется в пивоте)
    df = df.filter(pl.col("FACT").is_not_null())

    if df.is_empty():
        log("  ⚠ pivot_by_fact: нет данных")
        return df

    # Индексные колонки = всё кроме FACT и VALUE
    index_cols = [c for c in df.columns if c not in ("FACT", "VALUE")]

    # Убираем колонки которые зависят 1-к-1 от FACT (добавлены JOIN-ом справочника FACT)
    # Например "Показатель" — русское название FACT. Они делают каждую строку уникальной
    # и ломают пивот, создавая отдельную строку вместо отдельного столбца.
    fact_derived = []
    for col in index_cols:
        try:
            n_per_fact = (
                df.select(["FACT", col])
                .unique()
                .group_by("FACT")
                .agg(pl.col(col).n_unique().alias("n"))
                .filter(pl.col("n") > 1)
            )
            if len(n_per_fact) == 0:  # каждому FACT соответствует ровно 1 значение
                fact_derived.append(col)
        except Exception:
            pass

    if fact_derived:
        log(f"  Удалены FACT-зависимые столбцы: {fact_derived}")
        df = df.drop(fact_derived)
        index_cols = [c for c in index_cols if c not in fact_derived]

    fact_vals = df["FACT"].unique().drop_nulls().to_list()
    log(f"  FACT pivot → {len(fact_vals)} столбцов: {fact_vals}")

    try:
        pivoted = df.pivot(
            values="VALUE",
            index=index_cols,
            on="FACT",
            aggregate_function="mean",
        )
        return pivoted
    except Exception as e:
        log(f"  ⚠ pivot_by_fact ошибка: {e} — возвращаем без пивота")
        return df


def save_optimized(df, name: str, output_dir: Path, fmt: str, log, stop_event):
    import polars as pl

    if stop_event.is_set():
        return
    log(f"Сохранение {name}...")
    df = df.with_columns([pl.col(pl.Utf8).str.strip_chars()])

    if fmt == "csv":
        path = output_dir / f"{name}.csv"
        if "VALUE" in df.columns:
            df = df.with_columns(
                pl.col("VALUE")
                .cast(pl.Utf8)
                .str.replace(".", ",", literal=True)
                .alias("VALUE")
            )
        else:
            # После пивота числовые колонки — бывшие FACT-значения
            numeric_cols = [c for c in df.columns if df[c].dtype in (pl.Float32, pl.Float64, pl.Int32, pl.Int64, pl.Int16, pl.Int8)]
            if numeric_cols:
                df = df.with_columns([
                    pl.col(c).cast(pl.Utf8).str.replace(".", ",", literal=True)
                    for c in numeric_cols
                ])
        df.write_csv(
            path,
            separator=";",
            include_bom=True,
            quote_style="necessary",
            null_value="",
        )
        size_mb = path.stat().st_size / (1024**2)
        log(f"✓ {name}.csv — {size_mb:.1f} МБ")

    elif fmt == "excel":
        max_rows = 1_048_575
        num_parts = (len(df) + max_rows - 1) // max_rows
        if num_parts > 1:
            for i in range(num_parts):
                if stop_event.is_set():
                    return
                chunk = df.slice(i * max_rows, max_rows)
                chunk.write_excel(
                    output_dir / f"{name}_part{i + 1}.xlsx", worksheet="Data"
                )
            log(f"✓ {name} разбит на {num_parts} файлов .xlsx ({len(df)} строк)")
        else:
            df.write_excel(output_dir / f"{name}.xlsx", worksheet="Data")
            log(f"✓ {name}.xlsx сохранён ({len(df)} строк)")


def _apply_mappings(df, category, log, global_maps=None):
    """Оставлен для обратной совместимости — теперь делегирует в join_sprav."""
    if global_maps:
        return join_sprav(df, global_maps, log)
    return df


# ── Основная функция ───────────────────────────────────────────────────────────

def _process_one_source(
    input_file: str,
    region_type: str,     # "ST" или "NU"
    category: str,
    sprav_path: str | None,
    base_cache_dir: Path,
    log,
    stop_event,
) -> dict | None:
    """
    Обрабатывает один исходник Nielsen.
    Возвращает {"SKU_long_unique": df, "MAN_long": df, "BRAND_long": df} или None при ошибке.
    """
    import polars as pl
    import openpyxl

    log(f"── Исходник ({region_type}): {os.path.basename(input_file)}")

    cache_dir = _cache_dir_for(input_file, base_cache_dir)
    log(f"  Кэш: {cache_dir.name}")

    wb_meta = openpyxl.load_workbook(input_file, read_only=True, data_only=True)
    sheet_roles = _detect_sheets(wb_meta.sheetnames)
    wb_meta.close()

    if not sheet_roles["sku"]:
        log(f"  ❌ Листов SKU не найдено в {os.path.basename(input_file)}")
        return None

    log(f"  Листы SKU: {sheet_roles['sku']}")
    log(f"  Листы BRAND: {sheet_roles['brand']}")
    log(f"  Листы MAN: {sheet_roles['man']}")

    # Справочник: регион определяется по исходнику, не пользователем
    join_tables = load_join_tables(sprav_path, category, region_type)
    if join_tables:
        log(f"  Справочники ({region_type}): {len(join_tables)} таблиц")

    data = {}
    all_sheets = sheet_roles["sku"] + sheet_roles["brand"] + sheet_roles["man"]
    log(f"  Загрузка {len(all_sheets)} листов параллельно...")
    with ThreadPoolExecutor(max_workers=_max_workers()) as executor:
        futures = {
            executor.submit(load_sheet, sheet, input_file, cache_dir, log): sheet
            for sheet in all_sheets
        }
        for future in as_completed(futures):
            if stop_event.is_set():
                return None
            name, df = future.result()
            data[name] = df

    if stop_event.is_set():
        return None

    # Собираем melt-задачи
    sku_frames = [data[s] for s in sheet_roles["sku"] if s in data]
    df_sku = pl.concat(sku_frames, how="diagonal") if len(sku_frames) > 1 else sku_frames[0]

    melt_tasks: list[tuple[str, object, str]] = []
    for role, label in [("man", "MAN"), ("brand", "BRAND")]:
        sheets = sheet_roles[role]
        if not sheets:
            continue
        frames = [data[s] for s in sheets if s in data]
        df_role = pl.concat(frames, how="diagonal") if len(frames) > 1 else frames[0]
        melt_tasks.append((f"{label}_long", df_role, label))

    id_cols_sku, _ = _split_id_date(df_sku)
    melt_tasks.append(("SKU_long", df_sku, "SKU"))

    def _melt_task(key, df, label):
        id_cols, _ = _split_id_date(df)
        melted = melt_polars(df, id_cols)
        log(f"  melt {label}: {len(df)} строк → {len(melted)} строк")
        return key, melted, id_cols

    melted: dict = {}
    with ThreadPoolExecutor(max_workers=_max_workers()) as executor:
        futures = {
            executor.submit(_melt_task, k, df, lbl): k
            for k, df, lbl in melt_tasks
        }
        for future in as_completed(futures):
            if stop_event.is_set():
                return None
            try:
                k, m, ic = future.result()
                melted[k] = (m, ic)
            except Exception as e:
                log(f"  ⚠ Ошибка melt {futures[future]}: {e}")

    if stop_event.is_set():
        return None

    # Дедупликация SKU
    df_sku_long, id_cols_sku = melted.pop("SKU_long")
    df_sku_unique = df_sku_long.unique(subset=id_cols_sku + ["ATTRIBUTE"])
    log(f"  SKU после дедупликации: {len(df_sku_unique)} строк")

    log(f"  Справочник: ключи JOIN = {list(join_tables.keys())}")

    # JOIN справочников для всех трёх
    result = {}
    result["SKU_long_unique"] = join_sprav(df_sku_unique, join_tables, log)

    for role_key, (df_role, _) in melted.items():
        label = role_key.replace("_long", "")
        result[f"{label}_long"] = join_sprav(df_role, join_tables, log)

    return result, cache_dir.name   # возвращаем имя кэша чтобы не удалять


def process_nielsen(
    input_file: str,
    output_dir_str: str,
    fmt: str,
    log,
    messagebox,
    stop_event,
    category: str,
    sprav_path: str | None = None,
    input_file2: str | None = None,
    output_dir2: str | None = None,
    pq_file: str | None = None,
    pq_file_nu: str | None = None,
    arch_input: str | None = None,
    arch_input2: str | None = None,
    arch_enabled: bool = False,
):

    if not input_file and not input_file2:
        log("Ни один исходник Nielsen не выбран!")
        messagebox.showwarning("Ошибка", "Выберите хотя бы один исходник Nielsen")
        return

    # Определяем папки сохранения для каждого исходника
    dir1 = Path(output_dir_str) if output_dir_str else None
    dir2 = Path(output_dir2) if output_dir2 else dir1   # fallback на папку 1

    if not dir1 and not dir2:
        log("Папка сохранения не выбрана!")
        messagebox.showwarning("Ошибка", "Выберите папку сохранения")
        return

    # Общая кэш-папка — в первой доступной директории
    base_cache_root = (dir1 or dir2)
    base_cache_root.mkdir(parents=True, exist_ok=True)
    base_cache_dir = base_cache_root / "cache"
    base_cache_dir.mkdir(exist_ok=True)

    log(f"Категория: {category}")

    # ── Исходник 1 → регион ST ────────────────────────────────────────────────
    r1 = None
    used_caches = set()
    if input_file and os.path.isfile(input_file):
        res = _process_one_source(
            input_file, "ST", category, sprav_path, base_cache_dir, log, stop_event
        )
        if stop_event.is_set():
            return
        if res:
            r1, cache_name = res
            used_caches.add(cache_name)

    # ── Исходник 2 → регион NU ────────────────────────────────────────────────
    r2 = None
    if input_file2 and os.path.isfile(input_file2):
        res = _process_one_source(
            input_file2, "NU", category, sprav_path, base_cache_dir, log, stop_event
        )
        if stop_event.is_set():
            return
        if res:
            r2, cache_name = res
            used_caches.add(cache_name)

    # ── Сохранение — каждый исходник в свою папку ────────────────────────────
    def _save_result(result: dict, save_dir: Path, label: str):
        save_dir.mkdir(parents=True, exist_ok=True)
        for name, df in result.items():
            if stop_event.is_set():
                log("⛔ Сохранение остановлено")
                return
            df = pivot_by_fact(df, log)
            save_optimized(optimize_for_size(df), name, save_dir, fmt, log, stop_event)
        log(f"  Сохранено в {save_dir}")

    if r1 is not None and dir1:
        log("── Сохранение исходника 1 (ST)...")
        _save_result(r1, dir1, "ST")

    if stop_event.is_set():
        return

    if r2 is not None and dir2:
        log("── Сохранение исходника 2 (NU)...")
        _save_result(r2, dir2, "NU")

    if stop_event.is_set():
        return

    log(f"Готово! Формат: {fmt.upper()}")
    for cache_name in used_caches:
        _cleanup_old_caches(base_cache_dir, keep=cache_name, log=log)

    # ── Архивные источники (однократный прогон, суффикс _арх) ────────────────
    if arch_enabled:
        def _save_result_arch(result: dict, save_dir: Path, label: str):
            save_dir.mkdir(parents=True, exist_ok=True)
            for name, df in result.items():
                if stop_event.is_set():
                    return
                df = pivot_by_fact(df, log)
                save_optimized(optimize_for_size(df), f"{name}_арх", save_dir, fmt, log, stop_event)
            log(f"  Архив сохранён в {save_dir}")

        if arch_input and os.path.isfile(arch_input) and dir1:
            log("── Архивный исходник ST...")
            res = _process_one_source(arch_input, "ST", category, sprav_path, base_cache_dir, log, stop_event)
            if res and not stop_event.is_set():
                r_arch, cache_name = res
                used_caches.add(cache_name)
                _save_result_arch(r_arch, dir1, "ST_arch")

        if arch_input2 and os.path.isfile(arch_input2) and not stop_event.is_set():
            save_arch2 = dir2 or dir1
            log("── Архивный исходник NU...")
            res = _process_one_source(arch_input2, "NU", category, sprav_path, base_cache_dir, log, stop_event)
            if res and not stop_event.is_set():
                r_arch2, cache_name = res
                used_caches.add(cache_name)
                _save_result_arch(r_arch2, save_arch2, "NU_arch")
    else:
        log("ℹ️ Архивные источники пропущены (галочка не установлена)")

    # ── Обновление Power Query ST (если указан файл) ──────────────────────────
    if pq_file and os.path.isfile(pq_file):
        log(f"Обновление Power Query ST: {os.path.basename(pq_file)}...")
        from promodate_functions import refresh_file
        ok = refresh_file(pq_file, log, stop_event)
        if ok:
            log("✅ Power Query ST обновлён")
        else:
            log("⚠ Не удалось обновить Power Query ST")

    # ── Обновление Power Query NU (если указан файл) ──────────────────────────
    if pq_file_nu and os.path.isfile(pq_file_nu):
        log(f"Обновление Power Query NU: {os.path.basename(pq_file_nu)}...")
        from promodate_functions import refresh_file
        ok = refresh_file(pq_file_nu, log, stop_event)
        if ok:
            log("✅ Power Query NU обновлён")
        else:
            log("⚠ Не удалось обновить Power Query NU")

    messagebox.showinfo("Готово", "Обработка Nielsen завершена!")