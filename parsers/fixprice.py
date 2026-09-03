"""
Парсер магазинов Fix Price — fix-price.com/stores
Итерирует: Страна → каждый Регион → каждый Город, перехватывает /store API.

КЭШИРОВАНИЕ:
    По мере того как для города находятся данные (перехвачен ответ /store),
    они сразу же сохраняются на диск в fixprice_cache.json.
    При повторном запуске уже найденные города берутся из кэша и НЕ
    запрашиваются у сайта заново — прогресс не теряется при падении/остановке.
    Чтобы собрать всё заново — удалите fixprice_cache.json (или см. флаг --fresh).

Установка:
    pip install playwright openpyxl
    playwright install chromium

Запуск:
    python fixprice_scraper.py
    python fixprice_scraper.py --fresh   # игнорировать кэш и собрать заново
"""

import asyncio
import json
import logging
import re
import sys
from collections import Counter
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from playwright.async_api import async_playwright

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)

SITE_URL          = "https://fix-price.com/stores"
OUTPUT            = "fixprice_shops.xlsx"
CACHE_FILE        = Path("fixprice_cache.json")         # {"Регион::Город": [row, ...]}
CITIES_CACHE_FILE = Path("fixprice_cities_cache.json")  # {"Регион": ["Город1","Город2",...]}


# ═══════════════════════════════════════════════════════════════════════════════
# Кэш на диске
# ═══════════════════════════════════════════════════════════════════════════════

def _load_json(path: Path, label: str) -> dict:
    if not path.exists():
        return {}
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        log.warning("%s повреждён (%s), начинаем заново", label, e)
        return {}


def _save_json(path: Path, data: dict) -> None:
    """Атомарная запись: сначала во временный файл, затем replace."""
    tmp = path.with_suffix(".tmp")
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)
    tmp.replace(path)


def load_cache() -> dict:
    data = _load_json(CACHE_FILE, "Кэш магазинов")
    if data:
        total = sum(len(v) for v in data.values())
        log.info("Кэш магазинов: %s — %d городов, %d ТТ", CACHE_FILE, len(data), total)
    return data


def save_cache(cache: dict) -> None:
    _save_json(CACHE_FILE, cache)


def load_city_cache() -> dict:
    data = _load_json(CITIES_CACHE_FILE, "Кэш списков городов")
    if data:
        log.info("Кэш списков городов: %s — %d регионов", CITIES_CACHE_FILE, len(data))
    return data


def save_city_cache(city_cache: dict) -> None:
    _save_json(CITIES_CACHE_FILE, city_cache)


def cache_key(region: str, city: str) -> str:
    return f"{region}::{city}"


def region_fully_cached(region: str, city_cache: dict, store_cache: dict) -> bool:
    """True, если список городов региона известен и КАЖДЫЙ город уже есть в кэше магазинов."""
    cities = city_cache.get(region)
    if not cities:
        return False
    return all(cache_key(region, c) in store_cache for c in cities)


# ═══════════════════════════════════════════════════════════════════════════════
# Вспомогательные функции UI
# ═══════════════════════════════════════════════════════════════════════════════

async def safe_click(locator, timeout=8_000, attempts=3, retry_delay=1.5) -> bool:
    """
    Клик с предварительным скроллом к элементу (список городов/регионов
    виртуализирован — нужный <li> может быть вне зоны рендера), укороченным
    таймаутом и несколькими попытками — сайт иногда не готов принять клик
    с первого раза (анимация/оверлей ещё доигрывает). Ничего не бросает
    наружу — возвращает False при неудаче после всех попыток.
    """
    last_err = None
    for attempt in range(1, attempts + 1):
        try:
            await locator.scroll_into_view_if_needed(timeout=timeout)
            await locator.click(timeout=timeout)
            return True
        except Exception as e:
            last_err = e
            if attempt < attempts:
                await asyncio.sleep(retry_delay)
    log.warning("    Клик не удался (%d попыт.): %s",
                attempts, str(last_err).splitlines()[0])
    return False


async def dismiss_overlays(page):
    """
    Best-effort закрытие баннера cookies/согласия и подобных оверлеев,
    которые могут перекрывать элементы и мешать кликам. Ничего не бросает —
    если баннера нет, просто ничего не делает.
    """
    texts = ["Принять", "Согласен", "Согласна", "Хорошо", "ОК", "Accept", "Понятно"]
    for t in texts:
        try:
            btn = page.locator(f"button:has-text('{t}')").first
            if await btn.is_visible(timeout=500):
                await btn.click(timeout=1_000)
                await asyncio.sleep(0.3)
                return
        except Exception:
            pass


async def open_dropdown(dd_locator, timeout=8_000) -> bool:
    """Открывает дропдаун. Возвращает False вместо исключения при неудаче
    (сайт может ненадолго подвиснуть/анимация не успела завершиться)."""
    ok = await safe_click(dd_locator.locator("button.icon-button"), timeout=timeout)
    await asyncio.sleep(0.6)
    return ok


async def get_list_items(dd_locator, timeout=8_000):
    items = dd_locator.locator('[data-test="dropdown-list"] li')
    try:
        await items.first.wait_for(timeout=timeout)
    except Exception:
        return None
    return items


async def wait_enabled(input_locator, retries=30, delay=0.4):
    for _ in range(retries):
        if await input_locator.get_attribute("disabled") is None:
            return True
        await asyncio.sleep(delay)
    return False


# ═══════════════════════════════════════════════════════════════════════════════
# 2. НОРМАЛИЗАЦИЯ (объявлена раньше сбора, т.к. используется сразу по ходу сбора)
# ═══════════════════════════════════════════════════════════════════════════════

def normalize(raw: dict) -> dict:
    region  = raw.get("_region", "").strip()
    city    = raw.get("_city", "").strip()
    address = raw.get("address", "").strip()
    hours   = raw.get("scheduleWeekdays") or raw.get("workTime") or ""
    return {"region": region, "city": city, "address": address, "hours": hours, "phone": ""}


# ═══════════════════════════════════════════════════════════════════════════════
# 1. СБОР
# ═══════════════════════════════════════════════════════════════════════════════

DEBUG_PORT = 9222


async def collect_stores(playwright, cache: dict, city_cache: dict) -> list[dict]:
    # Подключаемся к уже запущенному системному Chrome/Edge через CDP —
    # так же, как это делает bristol.py. Раньше здесь был
    # playwright.chromium.launch() — отдельно скачиваемый Playwright-
    # Chromium, но в сетях с закрытым доступом к серверам загрузки
    # Microsoft/Google (частая ситуация в корпоративных сетях) его не
    # докачать даже вручную. Приложение само поднимает Chrome/Edge с
    # --remote-debugging-port перед запуском скрипта (см.
    # services/parsing_runner.py::ensure_chrome_cdp), отдельно скачивать
    # браузер для этого больше не нужно.
    log.info(f"Подключаюсь к браузеру на порту {DEBUG_PORT}…")
    browser = await playwright.chromium.connect_over_cdp(
        f"http://127.0.0.1:{DEBUG_PORT}", slow_mo=120
    )
    log.info("Подключено, открываю страницу…")
    context = browser.contexts[0] if browser.contexts else await browser.new_context(locale="ru-RU")
    page    = await context.new_page()

    all_stores: list[dict] = []
    store_event = asyncio.Event()
    captured: list[dict] = []

    # Заполняем all_stores тем, что уже есть в кэше — это учитывается сразу,
    # даже до захода на сайт по конкретному городу.
    cached_cities = 0
    for rows in cache.values():
        all_stores.extend(rows)
        cached_cities += 1
    if cached_cities:
        log.info("Из кэша предзагружено: %d городов, %d ТТ", cached_cities, len(all_stores))

    async def on_response(response):
        if "/buyer/v1/store" not in response.url:
            return
        if response.status != 200:
            return
        try:
            data  = await response.json()
            items = data if isinstance(data, list) else data.get("data", [])
            captured.extend(items)
            store_event.set()
        except Exception:
            pass

    page.on("response", on_response)

    log.info("Открываем %s ...", SITE_URL)
    await page.goto(SITE_URL, wait_until="load", timeout=60_000)
    await asyncio.sleep(2)
    await dismiss_overlays(page)

    dds = page.locator('[data-test="dropdown"]')
    country_dd = dds.nth(0)
    region_dd  = dds.nth(1)
    city_dd    = dds.nth(2)

    # ── Выбираем Россию ────────────────────────────────────────────────────────
    log.info("Открываем 'Страна'...")
    await open_dropdown(country_dd)
    c_items = await get_list_items(country_dd)
    if c_items is None:
        log.error("Список стран не появился!")
        await browser.close()
        return all_stores

    countries = await c_items.all_inner_texts()
    russia_idx = next(
        (i for i, c in enumerate(countries) if "россия" in c.lower()), 0
    )
    log.info("Выбираем: %s", countries[russia_idx])
    if not await safe_click(c_items.nth(russia_idx)):
        log.error("Не удалось выбрать страну!")
        await browser.close()
        return all_stores
    await asyncio.sleep(1.5)

    # ── Открываем список регионов ──────────────────────────────────────────────
    log.info("Ждём активации 'Регион'...")
    if not await wait_enabled(region_dd.locator('[data-test="input"]')):
        log.error("Регион не активировался!")
        await browser.close()
        return all_stores

    await open_dropdown(region_dd)
    r_items = await get_list_items(region_dd)
    if r_items is None:
        log.error("Список регионов пуст!")
        await browser.close()
        return all_stores

    regions = [r.strip() for r in await r_items.all_inner_texts()]
    log.info("Регионов: %d", len(regions))

    # ── Перебираем Регион → Город ──────────────────────────────────────────────
    for reg_idx, region_name in enumerate(regions):
        visible = await r_items.first.is_visible()
        if not visible:
            await open_dropdown(region_dd)
            r_items = await get_list_items(region_dd)
            if r_items is None:
                continue
            regions_now = [r.strip() for r in await r_items.all_inner_texts()]
            try:
                reg_idx_now = regions_now.index(region_name)
            except ValueError:
                reg_idx_now = reg_idx
        else:
            reg_idx_now = reg_idx

        log.info("[Регион %d/%d] %s", reg_idx + 1, len(regions), region_name)
        if not await safe_click(r_items.nth(reg_idx_now)):
            log.warning("  Не удалось выбрать регион, пропускаем")
            continue
        await asyncio.sleep(1.0)

        # ── Весь регион уже полностью в кэше — не открываем дропдаун городов ──
        if region_fully_cached(region_name, city_cache, cache):
            n_cities = len(city_cache[region_name])
            n_stores = sum(
                len(v) for k, v in cache.items() if k.startswith(f"{region_name}::")
            )
            log.info(
                "  → регион полностью из кэша (%d городов, %d ТТ), пропускаем",
                n_cities, n_stores,
            )
            print(f"  ✓ {region_name}: {n_stores} магазинов (кэш)")
            continue

        city_ok = await wait_enabled(
            city_dd.locator('[data-test="input"]'), retries=15, delay=0.3
        )
        if not city_ok:
            log.info("  Города не активировались, пропускаем")
            continue

        await open_dropdown(city_dd)
        ci_items = await get_list_items(city_dd, timeout=5_000)
        if ci_items is None:
            log.info("  Список городов пуст")
            continue

        cities = [c.strip() for c in await ci_items.all_inner_texts()]
        log.info("  Городов: %d", len(cities))

        # Запоминаем список городов региона — понадобится в следующий запуск,
        # чтобы можно было пропустить регион целиком, не открывая дропдаун.
        city_cache[region_name] = cities
        save_city_cache(city_cache)

        for city_idx, city_name in enumerate(cities):
            key = cache_key(region_name, city_name)

            cv = await ci_items.first.is_visible()
            if not cv:
                await open_dropdown(city_dd)
                ci_items = await get_list_items(city_dd, timeout=5_000)
                if ci_items is None:
                    break
                cities_now = [c.strip() for c in await ci_items.all_inner_texts()]
                try:
                    city_idx_now = cities_now.index(city_name)
                except ValueError:
                    city_idx_now = city_idx
            else:
                city_idx_now = city_idx

            # ── Уже есть в кэше — просто кликаем (чтобы не сбить навигацию
            #    дропдауна) и переходим дальше без ожидания сети ────────────
            if key in cache:
                log.info(
                    "  [Город %d/%d] %s → из кэша (%d ТТ)",
                    city_idx + 1, len(cities), city_name, len(cache[key]),
                )
                await safe_click(ci_items.nth(city_idx_now))
                await asyncio.sleep(0.3)
                continue

            captured.clear()
            store_event.clear()

            log.info("  [Город %d/%d] %s", city_idx + 1, len(cities), city_name)
            if not await safe_click(ci_items.nth(city_idx_now)):
                log.warning("    Не удалось кликнуть по городу, пропускаем "
                            "(при следующем запуске попробуется снова)")
                continue

            try:
                await asyncio.wait_for(store_event.wait(), timeout=6)
                city_rows = [
                    normalize({**s, "_region": region_name, "_city": city_name})
                    for s in captured
                ]
                log.info("    → %d ТТ", len(city_rows))
            except asyncio.TimeoutError:
                city_rows = []
                log.info("    → таймаут (0 ТТ)")

            # ── Сохраняем найденное сразу, независимо от того что дальше ───
            all_stores.extend(city_rows)
            cache[key] = city_rows
            save_cache(cache)

            await asyncio.sleep(0.3)

        # Итог по региону в консоль
        region_total = sum(
            len(v) for k, v in cache.items() if k.startswith(f"{region_name}::")
        )
        print(f"  ✓ {region_name}: {region_total} магазинов")

    grand_total = len(all_stores)
    print(f"\n{'─'*45}")
    print(f"  ИТОГО ПО ВСЕМ РЕГИОНАМ: {grand_total} магазинов")
    print(f"{'─'*45}\n")

    await browser.close()
    log.info("Итого собрано: %d записей", len(all_stores))
    return all_stores


# ═══════════════════════════════════════════════════════════════════════════════
# 3. ФИЛЬТРАЦИЯ / ДЕДУПЛИКАЦИЯ
# ═══════════════════════════════════════════════════════════════════════════════

def is_valid_address(addr: str) -> bool:
    addr = addr.strip()
    if not addr:
        return False
    if re.fullmatch(r'\d+\s*[а-яёa-z]?', addr, re.IGNORECASE):
        return False
    if not re.findall(r'[а-яёa-zA-Z]{3,}', addr):
        return False
    return True


def deduplicate(rows: list[dict]) -> list[dict]:
    seen, result = set(), []
    for r in rows:
        key = (r["region"], r["address"])
        if key not in seen:
            seen.add(key)
            result.append(r)
    return result


# ═══════════════════════════════════════════════════════════════════════════════
# 4. EXCEL
# ═══════════════════════════════════════════════════════════════════════════════

def save_to_excel(rows: list[dict], path: str) -> None:
    valid_rows    = deduplicate([r for r in rows if is_valid_address(r.get("address", ""))])
    region_counts = Counter(r["region"] for r in valid_rows)
    log.info("Уникальных ТТ: %d", len(valid_rows))

    wb = openpyxl.Workbook()

    # Стили
    thin      = Side(style="thin", color="CCCCCC")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill  = PatternFill("solid", fgColor="1F4E79")
    alt_fill  = PatternFill("solid", fgColor="EBF3FB")
    tot_fill  = PatternFill("solid", fgColor="D6E4F0")   # итог по области
    grand_fill= PatternFill("solid", fgColor="BDD7EE")   # общий итог
    hdr_font  = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    dat_font  = Font(name="Arial", size=10)
    tot_font  = Font(bold=True, name="Arial", size=10, italic=True)
    grand_font= Font(bold=True, name="Arial", size=11)

    # ── Лист 1: Итог по регионам ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Итог по регионам"
    for ci, (title, width) in enumerate([("Область / Регион", 35), ("Кол-во ТТ", 14)], 1):
        c = ws1.cell(row=1, column=ci, value=title)
        c.font = hdr_font; c.fill = hdr_fill; c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws1.column_dimensions[c.column_letter].width = width
    ws1.row_dimensions[1].height = 22

    grand_total = 0
    for ri, (region, count) in enumerate(
        sorted(region_counts.items(), key=lambda x: -x[1]), 2
    ):
        fill = alt_fill if ri % 2 == 0 else None
        for ci, val in enumerate([region, count], 1):
            c = ws1.cell(row=ri, column=ci, value=val)
            c.font = dat_font; c.border = border
            c.alignment = Alignment(
                horizontal="right" if ci == 2 else "left", vertical="center")
            if fill: c.fill = fill
        grand_total += count

    ri_grand = len(region_counts) + 2
    for ci, val in enumerate(["ИТОГО", grand_total], 1):
        c = ws1.cell(row=ri_grand, column=ci, value=val)
        c.font = grand_font; c.fill = grand_fill; c.border = border
        c.alignment = Alignment(
            horizontal="right" if ci == 2 else "left", vertical="center")

    ws1.freeze_panes    = "A2"
    ws1.auto_filter.ref = f"A1:B{ri_grand - 1}"

    # ── Лист 2: Все адреса — с итоговой строкой после каждой области ──────────
    ws2 = wb.create_sheet("Все адреса")
    COLS2 = [("Регион", 30), ("Город", 22), ("Адрес", 52),
             ("Часы работы", 24), ("Телефон", 18)]
    for ci, (title, width) in enumerate(COLS2, 1):
        c = ws2.cell(row=1, column=ci, value=title)
        c.font = hdr_font; c.fill = hdr_fill; c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws2.column_dimensions[c.column_letter].width = width
    ws2.row_dimensions[1].height = 22

    # Группируем по регионам, сохраняя порядок
    from itertools import groupby
    sorted_rows = sorted(valid_rows, key=lambda r: r["region"])

    row_num = 2
    sheet_total = 0

    for region_name, group in groupby(sorted_rows, key=lambda r: r["region"]):
        group_rows = list(group)
        region_count = len(group_rows)

        # Строки магазинов
        for r in group_rows:
            fill = alt_fill if row_num % 2 == 0 else None
            for ci, key in enumerate(["region", "city", "address", "hours", "phone"], 1):
                c = ws2.cell(row=row_num, column=ci, value=r.get(key, ""))
                c.font = dat_font; c.border = border
                c.alignment = Alignment(vertical="center", wrap_text=(ci == 3))
                if fill: c.fill = fill
            row_num += 1

        # Итоговая строка области
        for ci in range(1, 6):
            c = ws2.cell(row=row_num, column=ci)
            c.font = tot_font; c.fill = tot_fill; c.border = border
            c.alignment = Alignment(vertical="center")
        ws2.cell(row=row_num, column=1).value = (
            f"Итого по «{region_name}» — {region_count} магазинов"
        )
        ws2.cell(row=row_num, column=1).alignment = Alignment(
            vertical="center", horizontal="left")
        ws2.cell(row=row_num, column=2).value = region_count
        ws2.cell(row=row_num, column=2).alignment = Alignment(
            vertical="center", horizontal="center")
        row_num += 1
        sheet_total += region_count

    # Общий итог в конце
    for ci in range(1, 6):
        c = ws2.cell(row=row_num, column=ci)
        c.font = grand_font; c.fill = grand_fill; c.border = border
        c.alignment = Alignment(vertical="center")
    ws2.cell(row=row_num, column=1).value = (
        f"ИТОГО ПО ВСЕМ РЕГИОНАМ — {sheet_total} магазинов"
    )
    ws2.cell(row=row_num, column=1).alignment = Alignment(
        vertical="center", horizontal="left")
    ws2.cell(row=row_num, column=2).value = sheet_total
    ws2.cell(row=row_num, column=2).alignment = Alignment(
        vertical="center", horizontal="center")

    ws2.freeze_panes = "A2"

    wb.save(path)
    log.info("Сохранено → %s (%d регионов, %d ТТ)", path, len(region_counts), sheet_total)


# ═══════════════════════════════════════════════════════════════════════════════
# 5. MAIN
# ═══════════════════════════════════════════════════════════════════════════════

async def main_async(fresh: bool):
    import time
    t0 = time.time()

    cache      = {} if fresh else load_cache()
    city_cache = {} if fresh else load_city_cache()
    if fresh and (CACHE_FILE.exists() or CITIES_CACHE_FILE.exists()):
        log.info("--fresh: игнорируем существующий кэш (файлы не удаляются, будут перезаписаны)")

    log.info("Запускаю драйвер Playwright…")
    async with async_playwright() as playwright:
        log.info("Драйвер Playwright готов")
        rows = await collect_stores(playwright, cache, city_cache)

    if not rows:
        log.error("Магазины не найдены!")
        sys.exit(1)

    save_to_excel(rows, OUTPUT)

    elapsed = time.time() - t0
    valid = len(deduplicate([r for r in rows if is_valid_address(r.get("address", ""))]))
    print(f"\n✓ Готово за {elapsed:.1f} сек  |  Файл: {OUTPUT}")
    print(f"  Уникальных ТТ: {valid}")
    print(f"  Кэш магазинов: {CACHE_FILE} ({len(cache)} городов)")
    print(f"  Кэш списков городов: {CITIES_CACHE_FILE} ({len(city_cache)} регионов)")


def main():
    fresh = "--fresh" in sys.argv
    asyncio.run(main_async(fresh))


if __name__ == "__main__":
    main()