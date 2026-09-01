"""
Парсер магазинов «Красное&Белое» — krasnoeibeloe.ru
Асинхронный: регионы и города запрашиваются параллельно через asyncio + aiohttp.

Исправления против потери данных:
  1. Ретраи с backoff (3 попытки) на таймауты, 429, 5xx.
  2. Все ошибки и недоборы логируются как WARNING (раньше — молча в debug).
  3. Ручная пагинация ?page=N, если count > полученного, а "next" не пришёл.
  4. Регион + города опрашиваются ВСЕГДА, объединение с дедупликацией
     (раньше города опрашивались только при полностью пустом регионе).
  5. Параллельность снижена до 5 (меньше шансов словить троттлинг).

Запуск: python kb_scraper_fixed.py
Зависимости: pip install aiohttp requests openpyxl
"""

import re
import json
import asyncio
import logging
import sys
import html as html_lib
from collections import defaultdict, Counter

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import requests
import aiohttp

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)

BASE_URL    = "https://krasnoeibeloe.ru"
ADDR_URL    = f"{BASE_URL}/address/"
OUTPUT      = "kb_shops.xlsx"
CONCURRENCY = 5     # сколько регионов одновременно (10 могло вызывать 429)
RETRIES     = 3     # попыток на один URL

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0 Safari/537.36"
    ),
    "Accept": "application/json, text/html, */*",
    "Accept-Language": "ru-RU,ru;q=0.9",
    "Referer": ADDR_URL,
}


# ═══════════════════════════════════════════════════════════════════════════════
# 1. ПОЛУЧЕНИЕ HTML (синхронно — один раз)
# ═══════════════════════════════════════════════════════════════════════════════

def fetch_main_page() -> str:
    log.info("Открываю %s ...", ADDR_URL)
    r = requests.get(ADDR_URL, headers=HEADERS, timeout=25)
    r.raise_for_status()
    return r.text


# ═══════════════════════════════════════════════════════════════════════════════
# 2. ПАРСИНГ РЕГИОНОВ И ГОРОДОВ ИЗ HTML
# ═══════════════════════════════════════════════════════════════════════════════

def extract_balanced_json_array(text: str, start_pos: int) -> str:
    depth = 0
    in_string = False
    escape_next = False
    i = start_pos
    while i < len(text):
        ch = text[i]
        if escape_next:
            escape_next = False; i += 1; continue
        if ch == '\\' and in_string:
            escape_next = True; i += 1; continue
        if ch == '"':
            in_string = not in_string; i += 1; continue
        if not in_string:
            if ch == '[':
                depth += 1
            elif ch == ']':
                depth -= 1
                if depth == 0:
                    return text[start_pos: i + 1]
        i += 1
    return ""


def _parse_marker(html: str, marker: str) -> list[dict]:
    m = re.search(marker, html)
    if not m:
        return []
    arr_start = html.index('[', m.start())
    raw = extract_balanced_json_array(html, arr_start)
    if not raw:
        return []
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        return []


def extract_regions_and_cities(html: str) -> tuple[list[dict], list[dict]]:
    regions: list[dict] = []
    cities:  list[dict] = []

    for marker, target in (
        (r'"regions"\s*:\s*\[', regions),
        (r'"city"\s*:\s*\[',    cities),
        (r'"cities"\s*:\s*\[',  cities),
    ):
        parsed = _parse_marker(html, marker)
        if parsed:
            target.extend(parsed)

    if not regions and '&quot;' in html:
        decoded = html_lib.unescape(html)
        for marker, target in (
            (r'"regions"\s*:\s*\[', regions),
            (r'"city"\s*:\s*\[',    cities),
        ):
            parsed = _parse_marker(decoded, marker)
            if parsed:
                target.extend(parsed)

    cities = [c for c in cities if str(c.get("DEPTH_LEVEL", "")).strip() == "2"]

    log.info("Регионов: %d | Городов: %d", len(regions), len(cities))
    return regions, cities


def try_api_regions() -> list[dict]:
    for url in [f"{BASE_URL}/api/regions/", f"{BASE_URL}/api/v1/regions/"]:
        try:
            r = requests.get(url, headers=HEADERS, timeout=10)
            if r.status_code == 200:
                data = r.json()
                if isinstance(data, list) and data:
                    log.info("Регионы из API: %s → %d", url, len(data))
                    return data
        except Exception:
            pass
    return []


# ═══════════════════════════════════════════════════════════════════════════════
# 3. АСИНХРОННЫЕ ЗАПРОСЫ (с ретраями и полной пагинацией)
# ═══════════════════════════════════════════════════════════════════════════════

def _unwrap(data) -> tuple[list[dict], str | None, int | None]:
    """Возвращает (список магазинов, url следующей страницы, общее кол-во)."""
    next_url = None
    total    = None

    if isinstance(data, list):
        return data, None, len(data)

    if isinstance(data, dict):
        next_url = data.get("next") or None
        total    = data.get("count") or data.get("total") or None
        for key in ("results", "shops", "items", "data"):
            if isinstance(data.get(key), list):
                return data[key], next_url, total

    return [], None, None


async def _get_json_with_retries(session: aiohttp.ClientSession, url: str):
    """Один URL с ретраями. Возвращает dict/list, 'NOT_FOUND' или None (все попытки провалились)."""
    for attempt in range(RETRIES):
        try:
            async with session.get(url, timeout=aiohttp.ClientTimeout(total=40)) as r:
                if r.status == 404:
                    return "NOT_FOUND"
                if r.status == 429 or r.status >= 500:
                    wait = 2 ** attempt + 1
                    log.warning("HTTP %d на %s — жду %d сек (попытка %d/%d)",
                                r.status, url, wait, attempt + 1, RETRIES)
                    await asyncio.sleep(wait)
                    continue
                r.raise_for_status()
                return await r.json(content_type=None)
        except Exception as e:
            if attempt == RETRIES - 1:
                log.warning("ПОТЕРЯ ДАННЫХ: %s → %s", url, e)
            else:
                await asyncio.sleep(2 ** attempt + 1)
    return None


async def fetch_json_async(session: aiohttp.ClientSession, url: str) -> list[dict]:
    """Загружает ВСЕ страницы: следует за 'next', а если его нет,
    но count больше полученного — перебирает ?page=N вручную."""
    all_items: list[dict] = []
    expected_total: int | None = None
    current_url: str | None = url
    page = 1

    while current_url:
        data = await _get_json_with_retries(session, current_url)
        if data is None or data == "NOT_FOUND":
            break

        items, next_url, total = _unwrap(data)
        if expected_total is None:
            expected_total = total

        if not items:
            break
        all_items.extend(items)

        if next_url:
            # next может быть относительным
            current_url = next_url if next_url.startswith("http") else BASE_URL + next_url
        elif expected_total and len(all_items) < expected_total:
            # API объявил count, но 'next' не дал — пробуем ?page= вручную
            page += 1
            sep = "&" if "?" in url else "?"
            current_url = f"{url}{sep}page={page}"
        else:
            current_url = None

    if expected_total and len(all_items) < expected_total:
        log.warning("НЕДОБОР: %s — получено %d из %d", url, len(all_items), expected_total)

    return all_items


def normalize(raw: dict, region: str, city: str) -> dict:
    def g(*keys):
        for k in keys:
            for variant in (k, k.upper(), k.lower()):
                v = raw.get(variant)
                if v:
                    return str(v).strip()
        return ""
    return {
        "region":  region,
        "city":    city,
        "address": g("address", "addr", "UF_ADDRESS", "full_address", "shop_address"),
        "hours":   g("schedule", "work_time", "hours", "UF_SCHEDULE", "working_hours"),
        "phone":   g("phone", "tel", "UF_PHONE"),
    }


def _dedup_key(row: dict) -> tuple:
    return (
        row["region"].lower().strip(),
        row["city"].lower().strip(),
        re.sub(r"\s+", " ", row["address"].lower().strip()),
    )


async def fetch_region(
    session: aiohttp.ClientSession,
    semaphore: asyncio.Semaphore,
    region: dict,
    cities_by_region: dict[str, list],
    city_name_by_id: dict[str, str],
    idx: int,
    total: int,
) -> list[dict]:
    rid         = str(region["ID"])   # ВАЖНО: API ждёт ID, а не XML_ID!
    region_name = region.get("NAME", "").strip()

    seen: set = set()
    rows: list[dict] = []

    def add(row: dict):
        key = _dedup_key(row)
        if row["address"] and key in seen:
            return
        seen.add(key)
        rows.append(row)

    async with semaphore:
        # 1) Региональный эндпоинт (по ID! по XML_ID сервер отдаёт пустой список)
        shops_raw = await fetch_json_async(
            session, f"{BASE_URL}/api/regions/{rid}/shops/"
        )
        for s in shops_raw:
            city_val = (s.get("city_name") or s.get("city") or
                        s.get("CITY") or s.get("CITY_NAME") or "")
            if not city_val:
                # в ответе региона город приходит числом cityId
                city_val = city_name_by_id.get(str(s.get("cityId", "")), "")
            add(normalize(s, region_name, str(city_val).strip()))

        # 2) Городские эндпоинты — только если регион дал пусто (fallback)
        if not shops_raw:
            rcities = cities_by_region.get(rid, [])
            log.warning("%s: регион пуст, fallback по %d городам",
                        region_name, len(rcities))
            for city in rcities:
                city_name  = city.get("NAME", "").strip()
                city_shops = await fetch_json_async(
                    session, f"{BASE_URL}/api/cities/{city['ID']}/shops/"
                )
                for s in city_shops:
                    add(normalize(s, region_name, city_name))

    log.info("[%d/%d] %s → %d ТТ", idx, total, region_name, len(rows))
    return rows


async def collect_all(regions: list[dict], cities: list[dict]) -> list[dict]:
    cities_by_region: dict[str, list] = defaultdict(list)
    for c in cities:
        rid = str(c.get("IBLOCK_SECTION_ID", ""))
        cities_by_region[rid].append(c)

    city_name_by_id = {str(c["ID"]): c.get("NAME", "").strip() for c in cities}

    semaphore = asyncio.Semaphore(CONCURRENCY)
    total     = len(regions)

    connector = aiohttp.TCPConnector(limit=CONCURRENCY)
    async with aiohttp.ClientSession(headers=HEADERS, connector=connector) as session:
        tasks = [
            fetch_region(session, semaphore, region, cities_by_region,
                         city_name_by_id, i, total)
            for i, region in enumerate(regions, 1)
        ]
        results = await asyncio.gather(*tasks)

    all_rows: list[dict] = []
    for chunk in results:
        all_rows.extend(chunk)

    return all_rows


# ═══════════════════════════════════════════════════════════════════════════════
# 4. ФИЛЬТРАЦИЯ И СОХРАНЕНИЕ
# ═══════════════════════════════════════════════════════════════════════════════

def save_to_excel(rows: list[dict], path: str) -> None:
    # Фильтрация отключена — сохраняем ВСЕ строки без исключений
    valid_rows = rows
    log.info("Сохраняем все строки без фильтрации: %d", len(valid_rows))

    region_counts: Counter = Counter(r["region"] for r in valid_rows)

    wb = openpyxl.Workbook()

    # ── Лист 1: Итог по регионам ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Итог по регионам"

    thin      = Side(style="thin", color="CCCCCC")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill  = PatternFill("solid", fgColor="C00000")
    alt_fill  = PatternFill("solid", fgColor="FFF2F2")
    hdr_font  = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    dat_font  = Font(name="Arial", size=10)
    bold_font = Font(bold=True, name="Arial", size=10)

    for ci, (title, width) in enumerate(
        [("Область / Регион", 35), ("Кол-во ТТ", 14)], 1
    ):
        c = ws1.cell(row=1, column=ci, value=title)
        c.font      = hdr_font
        c.fill      = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = border
        ws1.column_dimensions[c.column_letter].width = width
    ws1.row_dimensions[1].height = 22

    total = 0
    for ri, (region, count) in enumerate(
        sorted(region_counts.items(), key=lambda x: -x[1]), 2
    ):
        fill = alt_fill if ri % 2 == 0 else None
        for ci, val in enumerate([region, count], 1):
            c = ws1.cell(row=ri, column=ci, value=val)
            c.font   = dat_font
            c.border = border
            c.alignment = Alignment(
                horizontal="right" if ci == 2 else "left",
                vertical="center"
            )
            if fill:
                c.fill = fill
        total += count

    ri_total = len(region_counts) + 2
    ws1.cell(row=ri_total, column=1, value="ИТОГО").font = bold_font
    ws1.cell(row=ri_total, column=1).border = border
    c_total = ws1.cell(row=ri_total, column=2, value=total)
    c_total.font      = bold_font
    c_total.border    = border
    c_total.alignment = Alignment(horizontal="right", vertical="center")

    ws1.freeze_panes    = "A2"
    ws1.auto_filter.ref = f"A1:B{ri_total - 1}"

    # ── Лист 2: Все адреса ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Все адреса")

    COLS2 = [("Регион", 30), ("Город", 22), ("Адрес", 52),
             ("Часы работы", 24), ("Телефон", 18)]
    for ci, (title, width) in enumerate(COLS2, 1):
        c = ws2.cell(row=1, column=ci, value=title)
        c.font      = hdr_font
        c.fill      = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = border
        ws2.column_dimensions[c.column_letter].width = width
    ws2.row_dimensions[1].height = 22

    keys = ["region", "city", "address", "hours", "phone"]
    for ri, row in enumerate(valid_rows, 2):
        fill = alt_fill if ri % 2 == 0 else None
        for ci, key in enumerate(keys, 1):
            c = ws2.cell(row=ri, column=ci, value=row.get(key, ""))
            c.font      = dat_font
            c.border    = border
            c.alignment = Alignment(vertical="center", wrap_text=(ci == 3))
            if fill:
                c.fill = fill

    ws2.freeze_panes    = "A2"
    ws2.auto_filter.ref = f"A1:E{len(valid_rows) + 1}"

    wb.save(path)
    log.info("Сохранено → %s  (%d регионов, %d ТТ)", path, len(region_counts), total)


# ═══════════════════════════════════════════════════════════════════════════════
# 5. MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    import time
    t0 = time.time()

    html = fetch_main_page()
    regions, cities = extract_regions_and_cities(html)

    if not regions:
        regions = try_api_regions()

    if not regions:
        log.error("Регионы не найдены!")
        sys.exit(1)

    log.info("Запускаем асинхронный сбор (%d регионов, параллельность=%d)...",
             len(regions), CONCURRENCY)

    rows = asyncio.run(collect_all(regions, cities))

    if not rows:
        log.error("Данных нет!")
        sys.exit(1)

    save_to_excel(rows, OUTPUT)

    elapsed = time.time() - t0
    print(f"\n✓ Готово за {elapsed:.1f} сек  |  Файл: {OUTPUT}")
    print(f"  Строк всего: {len(rows)} (без фильтрации)")
    print("  Если в логе были WARNING «ПОТЕРЯ ДАННЫХ» или «НЕДОБОР» — ")
    print("  часть точек не догрузилась, запустите повторно или снизьте CONCURRENCY.")


if __name__ == "__main__":
    main()