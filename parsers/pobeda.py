"""
Парсер магазинов «Победа» — победадискаунтер.рф
Скачивает страницу напрямую с сайта, извлекает данные из JS-переменной groups.
Сохраняет в Excel: «Итог по регионам» + «Все адреса».

Установка:
    pip install requests openpyxl

Запуск:
    python pobeda_parser.py
"""

import re
import sys
from collections import Counter
from itertools import groupby

import requests
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

URL    = "https://xn--80aaadiigoj9aqmm.xn--p1ai/%D0%BC%D0%B0%D0%B3%D0%B0%D0%B7%D0%B8%D0%BD%D1%8B/"
OUTPUT = "pobeda_shops.xlsx"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "ru-RU,ru;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}


# ═══════════════════════════════════════════════════════════════════════════════
# 1. ЗАГРУЗКА И ПАРСИНГ
# ═══════════════════════════════════════════════════════════════════════════════

def fetch_html() -> str:
    """Скачивает страницу магазинов с сайта Победа."""
    print(f"Загружаем {URL} ...")
    r = requests.get(URL, headers=HEADERS, timeout=30)
    r.raise_for_status()
    r.encoding = "utf-8"
    print(f"  → получено {len(r.text):,} символов, статус {r.status_code}")
    return r.text


def parse_html(html: str) -> list[dict]:
    """Извлекает все магазины из JS-переменной groups."""

    region_pattern = re.compile(
        r'name: "([^"]+)",\s*listid:.*?items: \[(.*?)\]\s*\}(?=\s*,?\s*\{|\s*\])',
        re.DOTALL
    )
    item_pattern = re.compile(
        r"balloonContent: '(.+?)'(?=,\s*\n\s*center:).*?center: \[([^\]]+)\].*?name: \"([^\"]+)\"",
        re.DOTALL
    )

    results = []

    for region_m in region_pattern.finditer(html):
        region_name = region_m.group(1)
        items_block = region_m.group(2)

        for item_m in item_pattern.finditer(items_block):
            balloon_html = item_m.group(1)
            coords       = item_m.group(2)
            address      = item_m.group(3).strip()

            # Часы работы
            hours_m = re.search(r'time\.svg[^/]*/>\s*([\d:]+\s*[\d:]+)', balloon_html)
            hours = hours_m.group(1).strip() if hours_m else ""

            # Телефон
            phone_m = re.search(r'phone\.svg[^/]*/>([\d\-\s]+)<', balloon_html)
            phone = phone_m.group(1).strip() if phone_m else "8-800-333-16-31"

            # Координаты
            lat, lon = [x.strip() for x in coords.split(',')]

            # Город = первый элемент до запятой
            city = address.split(',')[0].strip() if ',' in address else ""

            results.append({
                'region':  region_name,
                'city':    city,
                'address': address,
                'hours':   hours,
                'phone':   phone,
                'lat':     lat,
                'lon':     lon,
            })

    return results


# ═══════════════════════════════════════════════════════════════════════════════
# 2. EXCEL
# ═══════════════════════════════════════════════════════════════════════════════

def save_excel(rows: list[dict], path: str) -> None:
    region_counts = Counter(r['region'] for r in rows)

    wb = openpyxl.Workbook()

    thin       = Side(style="thin", color="CCCCCC")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill   = PatternFill("solid", fgColor="1F4E79")
    alt_fill   = PatternFill("solid", fgColor="EBF3FB")
    tot_fill   = PatternFill("solid", fgColor="D6E4F0")
    grand_fill = PatternFill("solid", fgColor="BDD7EE")
    hdr_font   = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    dat_font   = Font(name="Arial", size=10)
    tot_font   = Font(bold=True, italic=True, name="Arial", size=10)
    grand_font = Font(bold=True, name="Arial", size=11)

    def hcell(ws, row, col, val, width=None):
        c = ws.cell(row=row, column=col, value=val)
        c.font = hdr_font; c.fill = hdr_fill; c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center")
        if width:
            ws.column_dimensions[c.column_letter].width = width
        return c

    def dcell(ws, row, col, val, align="left", fill=None, font=None, wrap=False):
        c = ws.cell(row=row, column=col, value=val)
        c.font = font or dat_font; c.border = border
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        if fill:
            c.fill = fill
        return c

    # ── Лист 1: Итог по регионам ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Итог по регионам"
    ws1.row_dimensions[1].height = 22
    hcell(ws1, 1, 1, "Область / Регион", width=38)
    hcell(ws1, 1, 2, "Кол-во ТТ",        width=14)
    ws1.freeze_panes = "A2"

    grand = 0
    for ri, (region, count) in enumerate(
        sorted(region_counts.items(), key=lambda x: -x[1]), 2
    ):
        fill = alt_fill if ri % 2 == 0 else None
        dcell(ws1, ri, 1, region, fill=fill)
        dcell(ws1, ri, 2, count,  align="right", fill=fill)
        grand += count

    ri_g = len(region_counts) + 2
    dcell(ws1, ri_g, 1, "ИТОГО", font=grand_font, fill=grand_fill)
    dcell(ws1, ri_g, 2, grand,   align="right", font=grand_font, fill=grand_fill)
    ws1.auto_filter.ref = f"A1:B{ri_g - 1}"

    # ── Лист 2: Все адреса ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Все адреса")
    ws2.row_dimensions[1].height = 22
    COLS = [
        ("Регион",      30), ("Город",       22), ("Адрес",       52),
        ("Часы работы", 18), ("Телефон",      18), ("Широта",      14), ("Долгота", 14),
    ]
    for ci, (title, width) in enumerate(COLS, 1):
        hcell(ws2, 1, ci, title, width=width)
    ws2.freeze_panes = "A2"

    sorted_rows = sorted(rows, key=lambda r: r['region'])
    row_num = 2
    sheet_total = 0

    for region_name, group in groupby(sorted_rows, key=lambda r: r['region']):
        group_rows = list(group)
        cnt = len(group_rows)

        for r in group_rows:
            fill = alt_fill if row_num % 2 == 0 else None
            dcell(ws2, row_num, 1, r['region'],  fill=fill)
            dcell(ws2, row_num, 2, r['city'],    fill=fill)
            dcell(ws2, row_num, 3, r['address'], fill=fill, wrap=True)
            dcell(ws2, row_num, 4, r['hours'],   fill=fill, align="center")
            dcell(ws2, row_num, 5, r['phone'],   fill=fill, align="center")
            dcell(ws2, row_num, 6, r['lat'],     fill=fill, align="center")
            dcell(ws2, row_num, 7, r['lon'],     fill=fill, align="center")
            row_num += 1

        # Итог по региону
        for ci in range(1, 8):
            c = ws2.cell(row=row_num, column=ci)
            c.fill = tot_fill; c.border = border; c.font = tot_font
            c.alignment = Alignment(vertical="center")
        ws2.cell(row=row_num, column=1).value = f"Итого по «{region_name}»"
        ws2.cell(row=row_num, column=2).value = f"{cnt} магазинов"
        ws2.cell(row=row_num, column=2).alignment = Alignment(horizontal="center", vertical="center")
        row_num += 1
        sheet_total += cnt

    # Общий итог
    for ci in range(1, 8):
        c = ws2.cell(row=row_num, column=ci)
        c.fill = grand_fill; c.border = border; c.font = grand_font
        c.alignment = Alignment(vertical="center")
    ws2.cell(row=row_num, column=1).value = "ИТОГО ПО ВСЕМ РЕГИОНАМ"
    ws2.cell(row=row_num, column=2).value = f"{sheet_total} магазинов"
    ws2.cell(row=row_num, column=2).alignment = Alignment(horizontal="center", vertical="center")
    ws2.auto_filter.ref = f"A1:G{row_num - 1}"

    wb.save(path)


# ═══════════════════════════════════════════════════════════════════════════════
# 3. MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    # Можно передать HTML-файл как аргумент (для оффлайн-режима)
    if len(sys.argv) > 1:
        print(f"Читаем из файла: {sys.argv[1]}")
        with open(sys.argv[1], encoding="utf-8") as f:
            html = f.read()
    else:
        html = fetch_html()

    if "var groups" not in html:
        print("ОШИБКА: переменная groups не найдена в HTML. Возможно, сайт вернул ошибку.")
        sys.exit(1)

    rows = parse_html(html)

    if not rows:
        print("ОШИБКА: магазины не найдены.")
        sys.exit(1)

    region_counts = Counter(r['region'] for r in rows)
    print(f"\n{'─'*50}")
    for region, count in sorted(region_counts.items(), key=lambda x: -x[1]):
        print(f"  ✓ {region}: {count} магазинов")
    print(f"{'─'*50}")
    print(f"  ИТОГО: {len(region_counts)} регионов, {len(rows)} магазинов")
    print(f"{'─'*50}\n")

    save_excel(rows, OUTPUT)
    print(f"✓ Готово → {OUTPUT}")


if __name__ == "__main__":
    main()