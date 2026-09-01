"""
Парсер Бристоль v3 — CDP + перехват живых заголовков.

Идея: не угадываем токен/подпись, а ПЕРЕХВАТЫВАЕМ заголовки, которые сайт
bristol.ru сам отправляет в api.mobile.bristol.ru из вашего браузера.
Работает даже если Бристоль сменил токен, appversion или секрет подписи.

Как работает:
  1. Подключаемся к вашему Chrome по CDP (порт 9222).
  2. Открываем bristol.ru/shops и слушаем сетевой трафик.
  3. Как только сайт сам дёргает API — копируем его заголовки
     (токен, подпись и всё остальное). Ответ coords тоже забираем
     прямо из трафика, если повезёт.
  4. Открываем вкладку на домене API и качаем детали магазинов
     same-origin fetch'ами с перехваченными заголовками (CORS не мешает).
  5. Если подпись протухает — перезагружаем bristol.ru/shops
     и перехватываем свежие заголовки автоматически.

Шаги:
1. Закройте Chrome полностью:  taskkill /F /IM chrome.exe /T
2. Запустите Chrome с флагом отладки:
   "C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe" --remote-debugging-port=9222 --user-data-dir="C:\\chrome-debug" --no-first-run
3. Запустите скрипт: python bristol.py

Установка:
    pip install playwright openpyxl
    playwright install chromium
"""

import sys
import json
import time
import asyncio
import hashlib
from pathlib import Path
from collections import Counter
from itertools import groupby

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from playwright.async_api import async_playwright

OUTPUT      = "bristol_shops.xlsx"
CHECKPOINT  = Path("bristol_checkpoint.json")
DEBUG_PORT  = 9222
API_HOST    = "api.mobile.bristol.ru"
BASE_API    = f"https://{API_HOST}/api/v2"
SHOPS_URL   = "https://bristol.ru/shops"

# Старая схема подписи — используется, только если перехваченные
# заголовки её подтверждают (тогда можем сами обновлять timestamp/hash).
SECRET      = "K2DLLQnw51KtLRg1WmjhzmJBUluaTQcf"

BATCH       = 150   # магазинов на один вызов page.evaluate
JS_PARALLEL = 1     # строго по одному запросу за раз
PAUSE_MS    = 200   # пауза между запросами, мс
MAX_PASSES  = 3     # проходов по недокачанным id

# Автоподстройка: если скорость падает (Qrator начал тормозить),
# скрипт сам снижает параллельность и увеличивает паузу.

def ts() -> str:
    return time.strftime("%H:%M:%S")


# Заголовки, которые НЕ копируем из перехваченного запроса
SKIP_HEADERS = {
    "host", "content-length", "connection", "cookie", "origin", "referer",
    "user-agent", "accept-encoding", "accept-language", "sec-fetch-dest",
    "sec-fetch-mode", "sec-fetch-site", "sec-ch-ua", "sec-ch-ua-mobile",
    "sec-ch-ua-platform", "pragma", "cache-control", "priority",
}


# ═══════════════════════════════════════════════════════════════════════════════
# Перехват заголовков и подпись
# ═══════════════════════════════════════════════════════════════════════════════

class ApiAuth:
    """Хранит перехваченные заголовки и умеет их освежать."""

    def __init__(self):
        self.headers: dict = {}
        self.can_sign = False      # знаем ли схему подписи
        self.appversion = ""

    def adopt(self, raw_headers: dict) -> None:
        self.headers = {
            k: v for k, v in raw_headers.items()
            if not k.startswith(":") and k.lower() not in SKIP_HEADERS
        }
        self.appversion = raw_headers.get("appversion", "")
        ts, h = raw_headers.get("timestamp"), raw_headers.get("hash")
        self.can_sign = False
        if ts and h and self.appversion:
            d = hashlib.md5((SECRET + self.appversion).encode()).hexdigest()
            if hashlib.md5((ts + d).encode()).hexdigest() == h:
                self.can_sign = True
        print(f"  Перехвачены заголовки API: {sorted(self.headers)}")
        print(f"  Схема подписи {'ПОДТВЕРЖДЕНА — освежаем сами' if self.can_sign else 'неизвестна — используем как есть'}")

    def fresh(self) -> dict:
        hdrs = dict(self.headers)
        if self.can_sign:
            ts = str(int(time.time() * 1000))
            d  = hashlib.md5((SECRET + self.appversion).encode()).hexdigest()
            hdrs["timestamp"] = ts
            hdrs["hash"] = hashlib.md5((ts + d).encode()).hexdigest()
        return hdrs


async def capture_auth(ctx, auth: ApiAuth, timeout_s: int = 45) -> list | None:
    """Открывает/перезагружает bristol.ru/shops и ждёт первый запрос к API.
    Возвращает тело coords, если удалось выхватить его из трафика."""
    fut_headers: asyncio.Future = asyncio.get_event_loop().create_future()
    coords_body: list | None = None
    coords_event = asyncio.Event()

    def on_request(req):
        if API_HOST in req.url and not fut_headers.done():
            fut_headers.set_result(dict(req.headers))

    async def on_response(resp):
        nonlocal coords_body
        if API_HOST in resp.url and "/shops/coords" in resp.url and resp.ok:
            try:
                body = await resp.json()
                if isinstance(body, list):
                    coords_body = body
                    coords_event.set()
            except Exception:
                pass

    # Ищем существующую вкладку bristol.ru или создаём новую
    page = None
    for p in ctx.pages:
        if "bristol.ru" in p.url and API_HOST not in p.url:
            page = p
            break
    if page is None:
        page = await ctx.new_page()

    page.on("request", on_request)
    page.on("response", lambda r: asyncio.ensure_future(on_response(r)))

    print(f"  {ts()} Открываю {SHOPS_URL} и слушаю трафик...")
    try:
        await page.goto(SHOPS_URL, wait_until="domcontentloaded", timeout=60_000)
    except Exception as e:
        print(f"  (переход: {e})")

    try:
        raw = await asyncio.wait_for(fut_headers, timeout=timeout_s)
        auth.adopt(raw)
    except asyncio.TimeoutError:
        print("  ⚠ За 45 сек сайт не сделал ни одного запроса к API.")
        print("    Покликайте по карте магазинов в открывшейся вкладке — ")
        print("    скрипт ждёт ещё 60 сек...")
        try:
            raw = await asyncio.wait_for(fut_headers, timeout=60)
            auth.adopt(raw)
        except asyncio.TimeoutError:
            return None

    # Даём шанс выхватить coords из трафика страницы (сайт грузит карту сам)
    try:
        await asyncio.wait_for(coords_event.wait(), timeout=10)
    except asyncio.TimeoutError:
        pass

    return coords_body


# ═══════════════════════════════════════════════════════════════════════════════
# JS, выполняемый во вкладке на домене API (same-origin — CORS не мешает)
# ═══════════════════════════════════════════════════════════════════════════════

JS_FETCH_ONE = """
async ({ url, headers }) => {
    try {
        const r = await fetch(url, { headers, credentials: 'include' });
        const text = await r.text();
        if (!r.ok) return { __error: 'HTTP ' + r.status + ': ' + text.slice(0, 120) };
        return JSON.parse(text);
    } catch (e) {
        return { __error: String(e) };
    }
}
"""

JS_FETCH_MANY = """
async ({ ids, headers, parallel, pauseMs }) => {
    const results = {};
    const failed  = [];
    let   authFail = 0;
    let   slow     = 0;   // ответов дольше 3 сек — признак троттлинга Qrator
    const queue   = [...ids];

    const report = async (info) => {
        try { if (window.__bristol_progress) await window.__bristol_progress(info); }
        catch (e) {}
    };

    async function worker() {
        while (queue.length) {
            const id = queue.shift();
            const t0 = Date.now();
            try {
                const r = await fetch(
                    `/api/v2/shops/${id}?consumer=website`,
                    { headers, credentials: 'include' }
                );
                const ms = Date.now() - t0;
                if (ms > 3000) slow++;
                if (r.ok) {
                    const data = await r.json();
                    results[id] = data;
                    const name = (data && data.shop && data.shop.name) || '';
                    await report({ id, ok: true, status: r.status, ms, name });
                } else {
                    failed.push(id);
                    if (r.status === 401 || r.status === 403) authFail++;
                    await report({ id, ok: false, status: r.status, ms, name: '' });
                    if (r.status === 429) await new Promise(res => setTimeout(res, 3000));
                }
            } catch (e) {
                failed.push(id);
                await report({ id, ok: false, status: 0, ms: Date.now() - t0,
                               name: String(e).slice(0, 60) });
            }
            // Пауза с джиттером — ровный человеческий темп вместо шквала
            await new Promise(res => setTimeout(res, pauseMs + Math.random() * pauseMs));
        }
    }

    await Promise.all(Array.from({ length: parallel }, worker));
    return { results, failed, authFail, slow };
}
"""


# ═══════════════════════════════════════════════════════════════════════════════
# Сбор данных
# ═══════════════════════════════════════════════════════════════════════════════

def load_checkpoint() -> dict:
    if CHECKPOINT.exists():
        try:
            data = json.loads(CHECKPOINT.read_text(encoding="utf-8"))
            print(f"  Чекпоинт: уже скачано {len(data)} магазинов, продолжаем")
            return data
        except Exception:
            pass
    return {}


def save_checkpoint(details: dict) -> None:
    CHECKPOINT.write_text(json.dumps(details, ensure_ascii=False), encoding="utf-8")


async def get_api_page(ctx):
    """Вкладка на домене API для same-origin запросов."""
    for p in ctx.pages:
        if API_HOST in p.url:
            return p
    page = await ctx.new_page()
    try:
        await page.goto(f"{BASE_API}/shops/coords?consumer=website",
                        wait_until="domcontentloaded", timeout=60_000)
    except Exception as e:
        print(f"  (переход на API-домен: {e})")
    await page.wait_for_timeout(3000)   # время на Qrator-челлендж
    return page


async def fetch_coords(api_page, auth: ApiAuth) -> list:
    for attempt in range(1, 7):
        data = await api_page.evaluate(
            JS_FETCH_ONE,
            {"url": "/api/v2/shops/coords?consumer=website", "headers": auth.fresh()}
        )
        if isinstance(data, list):
            return data
        err = data.get("__error") if isinstance(data, dict) else str(data)[:150]
        print(f"  {ts()} coords: попытка {attempt}/6 → {err}")
        if attempt == 3:
            try:
                await api_page.reload(wait_until="domcontentloaded", timeout=60_000)
            except Exception:
                pass
        await api_page.wait_for_timeout(4000)
    return []


async def fetch_all_details(api_page, ctx, auth: ApiAuth,
                            ids: list, details: dict) -> None:
    remaining = [i for i in ids if str(i) not in details]
    total     = len(ids)
    t0        = time.time()

    parallel = JS_PARALLEL
    pause_ms = PAUSE_MS

    # Живой прогресс из браузера: печатаем каждый запрос
    counter = {"n": len(details)}

    def on_progress(info):
        counter["n"] += 1 if info.get("ok") else 0
        n    = counter["n"]
        sid  = info.get("id", "?")
        ms   = info.get("ms", 0)
        name = (info.get("name") or "")[:55]
        if info.get("ok"):
            mark = "✓" if ms < 3000 else "✓ МЕДЛЕННО"
            print(f"  {ts()} [{n:5}/{total}] id={sid:<9} {mark} {ms:5} мс  {name}")
        else:
            st = info.get("status", 0)
            what = f"HTTP {st}" if st else f"сбой: {name}"
            print(f"  {ts()} [{n:5}/{total}] id={sid:<9} ✗ {what} ({ms} мс) → добор позже")

    try:
        await api_page.expose_function("__bristol_progress", on_progress)
        print(f"  {ts()} Живой прогресс подключён")
    except Exception:
        pass   # уже зарегистрирована (повторный запуск в той же вкладке)

    for pass_num in range(1, MAX_PASSES + 1):
        if not remaining:
            break
        if pass_num > 1:
            print(f"\n  Проход {pass_num}: добираем {len(remaining)} шт...")

        next_remaining = []
        for start in range(0, len(remaining), BATCH):
            chunk = remaining[start:start + BATCH]
            out = await api_page.evaluate(JS_FETCH_MANY, {
                "ids": chunk,
                "headers": auth.fresh(),
                "parallel": parallel,
                "pauseMs": pause_ms,
            })
            for sid, data in out["results"].items():
                details[str(sid)] = data
            next_remaining.extend(out["failed"])
            save_checkpoint(details)

            done  = len(details)
            speed = (done - (total - len(remaining) - len(next_remaining))) and \
                    done / max(time.time() - t0, 0.1)
            eta_m = (total - done) / max(speed, 0.01) / 60 if speed else 0
            print(f"  {ts()} ── чекпоинт сохранён: {done}/{total} │ "
                  f"скорость {speed:.1f} ТТ/сек │ осталось ~{eta_m:.0f} мин │ "
                  f"ошибок в батче: {len(out['failed'])} ──")

            # Автоподстройка темпа: Qrator начал тормозить → сбавляем
            slow = out.get("slow", 0)
            if slow > len(chunk) * 0.3 or len(out["failed"]) > len(chunk) * 0.3:
                parallel = max(2, parallel - 1)
                pause_ms = min(1500, int(pause_ms * 1.5))
                print(f"  {ts()} ↓ ТРОТТЛИНГ QRATOR: медленных ответов {slow}, "
                      f"ошибок {len(out['failed'])} из {len(chunk)}")
                print(f"  {ts()}   снижаю темп: пауза={pause_ms} мс, отдыхаю 20 сек...")
                await api_page.wait_for_timeout(20_000)
            elif slow == 0 and not out["failed"] and pause_ms > PAUSE_MS:
                # Всё гладко — потихоньку возвращаем темп
                pause_ms = max(PAUSE_MS, int(pause_ms / 1.3))
                if pause_ms > PAUSE_MS:
                    print(f"  {ts()} ↑ всё гладко, ускоряюсь: пауза={pause_ms} мс")

            # Подпись протухла — перехватываем свежие заголовки и продолжаем
            if out.get("authFail", 0) > len(chunk) * 0.5:
                print("  Подпись протухла → перехватываю свежие заголовки...")
                await capture_auth(ctx, auth, timeout_s=45)

        remaining = next_remaining

    if remaining:
        print(f"\n  ⚠ НЕ СКАЧАЛОСЬ {len(remaining)} магазинов после "
              f"{MAX_PASSES} проходов: {remaining[:20]}{'...' if len(remaining) > 20 else ''}")


def details_to_rows(details: dict) -> list:
    rows = []
    for data in details.values():
        if not isinstance(data, dict) or "shop" not in data:
            continue
        s      = data["shop"]
        region = data.get("region") or {}
        city   = data.get("city") or {}
        sched  = s.get("schedule") or []
        hours  = "; ".join(
            f"{x['days']} {x['time']}" for x in sched if x.get("days")
        )
        rows.append({
            "region":  region.get("name", "Неизвестный регион"),
            "city":    city.get("name", "") if isinstance(city, dict) else "",
            "address": s.get("name", ""),
            "hours":   hours,
            "phone":   s.get("phone", "") or "",
            "lat":     s.get("lat", ""),
            "lon":     s.get("lon", ""),
        })
    return rows


async def collect_all() -> list:
    async with async_playwright() as pw:
        print(f"{ts()} Подключаемся к Chrome на порту {DEBUG_PORT}...")
        try:
            browser = await pw.chromium.connect_over_cdp(
                f"http://127.0.0.1:{DEBUG_PORT}"
            )
        except Exception as e:
            print(f"\nОШИБКА подключения: {e}")
            print("\nУбедитесь что Chrome запущен с флагом "
                  "--remote-debugging-port=9222")
            return []

        ctx = browser.contexts[0] if browser.contexts else await browser.new_context()

        # 1. Перехват заголовков (и, если повезёт, самого списка coords)
        auth = ApiAuth()
        coords = await capture_auth(ctx, auth)
        if not auth.headers:
            print("  ОШИБКА: не удалось перехватить заголовки API.")
            return []

        # 2. Вкладка на домене API для same-origin запросов
        api_page = await get_api_page(ctx)

        # 3. Список магазинов
        if coords:
            print(f"  ✓ coords выхвачен из трафика страницы: {len(coords)} магазинов")
        else:
            print("  Получаем список магазинов через API...")
            coords = await fetch_coords(api_page, auth)
        if not coords:
            print("  ОШИБКА: не удалось получить список магазинов.")
            return []

        opened = [s for s in coords if s.get("store_status") == "opened"]
        ids    = [s["id"] for s in opened]
        print(f"  ✓ Магазинов всего: {len(coords)}, открытых: {len(ids)}\n")

        # 4. Детали
        details = load_checkpoint()
        await fetch_all_details(api_page, ctx, auth, ids, details)

        # НЕ закрываем browser: это живой Chrome пользователя.
        return details_to_rows(details)


# ═══════════════════════════════════════════════════════════════════════════════
# Excel
# ═══════════════════════════════════════════════════════════════════════════════

def save_excel(rows, path):
    wb = openpyxl.Workbook()
    thin       = Side(style="thin", color="CCCCCC")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill   = PatternFill("solid", fgColor="1F4E79")
    alt_fill   = PatternFill("solid", fgColor="EBF3FB")
    tot_fill   = PatternFill("solid", fgColor="D6E4F0")
    grand_fill = PatternFill("solid", fgColor="BDD7EE")
    hdr_font   = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    dat_font   = Font(name="Arial", size=10)
    tot_font   = Font(bold=True, italic=True, name="Arial", size=10)
    grand_font = Font(bold=True, name="Arial", size=11)

    def hcell(ws, r, c, v, w=None):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if w: ws.column_dimensions[cell.column_letter].width = w

    def dcell(ws, r, c, v, align="left", fill=None, font=None, wrap=False):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = font or dat_font; cell.border = border
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        if fill: cell.fill = fill

    # ── Лист 1: Итог по регионам ─────────────────────────────────────────────
    ws1 = wb.active; ws1.title = "Итог по регионам"
    ws1.row_dimensions[1].height = 22
    hcell(ws1, 1, 1, "Область / Регион", w=40)
    hcell(ws1, 1, 2, "Кол-во ТТ", w=12)
    ws1.freeze_panes = "A2"

    reg_counts = Counter(r["region"] for r in rows)
    gs = 0
    for ri, (region, cnt) in enumerate(
        sorted(reg_counts.items(), key=lambda x: -x[1]), 2
    ):
        fill = alt_fill if ri % 2 == 0 else None
        dcell(ws1, ri, 1, region, fill=fill)
        dcell(ws1, ri, 2, cnt, fill=fill, align="right")
        gs += cnt
    ri_g = len(reg_counts) + 2
    dcell(ws1, ri_g, 1, "ИТОГО", font=grand_font, fill=grand_fill)
    dcell(ws1, ri_g, 2, gs, font=grand_font, fill=grand_fill, align="right")
    ws1.auto_filter.ref = f"A1:B{ri_g - 1}"

    # ── Лист 2: Все адреса ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("Все адреса")
    ws2.row_dimensions[1].height = 22
    for ci, (t, w) in enumerate([
        ("Регион", 34), ("Город", 20), ("Адрес", 60), ("Часы работы", 26),
        ("Телефон", 18), ("Широта", 12), ("Долгота", 12)
    ], 1):
        hcell(ws2, 1, ci, t, w=w)
    ws2.freeze_panes = "A2"
    NCOLS = 7

    sorted_rows = sorted(rows, key=lambda r: (r["region"], r["city"], r["address"]))
    row_num = 2; sheet_total = 0

    for region, grp in groupby(sorted_rows, key=lambda r: r["region"]):
        grp = list(grp)
        for r in grp:
            fill = alt_fill if row_num % 2 == 0 else None
            dcell(ws2, row_num, 1, r["region"], fill=fill)
            dcell(ws2, row_num, 2, r["city"], fill=fill)
            dcell(ws2, row_num, 3, r["address"], fill=fill, wrap=True)
            dcell(ws2, row_num, 4, r["hours"], fill=fill, align="center")
            dcell(ws2, row_num, 5, r["phone"], fill=fill, align="center")
            dcell(ws2, row_num, 6, r["lat"], fill=fill, align="center")
            dcell(ws2, row_num, 7, r["lon"], fill=fill, align="center")
            row_num += 1
        for ci in range(1, NCOLS + 1):
            c = ws2.cell(row=row_num, column=ci)
            c.fill = tot_fill; c.border = border; c.font = tot_font
            c.alignment = Alignment(vertical="center")
        ws2.cell(row=row_num, column=1).value = f"Итого по «{region}»"
        ws2.cell(row=row_num, column=2).value = f"{len(grp)} ТТ"
        ws2.cell(row=row_num, column=2).alignment = Alignment(
            horizontal="center", vertical="center")
        row_num += 1; sheet_total += len(grp)

    for ci in range(1, NCOLS + 1):
        c = ws2.cell(row=row_num, column=ci)
        c.fill = grand_fill; c.border = border; c.font = grand_font
        c.alignment = Alignment(vertical="center")
    ws2.cell(row=row_num, column=1).value = "ИТОГО ПО ВСЕМ РЕГИОНАМ"
    ws2.cell(row=row_num, column=2).value = f"{sheet_total} ТТ"
    ws2.cell(row=row_num, column=2).alignment = Alignment(
        horizontal="center", vertical="center")
    ws2.auto_filter.ref = f"A1:G{row_num - 1}"
    wb.save(path)


def main():
    print("=" * 55)
    print("  Парсер Бристоль v3 — перехват живых заголовков")
    print("=" * 55 + "\n")

    t0 = time.time()
    rows = asyncio.run(collect_all())
    if not rows:
        sys.exit(1)

    reg_counts = Counter(r["region"] for r in rows)
    print(f"\n{'─' * 55}")
    for region, count in sorted(reg_counts.items(), key=lambda x: -x[1]):
        print(f"  ✓ {region}: {count} ТТ")
    print(f"{'─' * 55}")
    print(f"  ИТОГО: {len(reg_counts)} регионов, {sum(reg_counts.values())} ТТ")
    print(f"{'─' * 55}\n")

    save_excel(rows, OUTPUT)
    print(f"✓ Готово за {time.time() - t0:.0f} сек → {OUTPUT}")
    print(f"  Чекпоинт {CHECKPOINT} можно удалить (или оставить для докачки).")


if __name__ == "__main__":
    main()