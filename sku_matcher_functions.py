"""
sku_matcher_functions.py

Модель матчинга SKU обучается на индикативных парах справочника:
  col[2] raw_name  →  col[3] corrected_name

Новый CSV SKU приходит в формате col[2] (длинное сырое название).
Модель находит ближайшую пару и возвращает col[3] как нормализованный результат.

Ключевые правила:
  1. Матчим только по индикативным строкам — они «правильные примеры»
  2. Новый SKU не должен быть 1-в-1 копией col[2] — значит он уже есть в базе
  3. Используем RapidFuzz (уже в requirements) — умнее и быстрее ручного Левенштейна
"""

import re
from collections import defaultdict
from pathlib import Path

import pandas as pd


# ─── Стоп-слова ───────────────────────────────────────────────────────────────

_STOP = frozenset({
    "г", "кг", "мл", "л", "шт", "уп", "пак", "бут", "бан", "кор",
    "пл", "жб", "пэт", "мас", "кал",
    "и", "в", "с", "от", "из", "для", "по", "на", "не", "за",
    "со", "во", "при", "без", "об", "же", "до", "то",
    "что", "как", "все", "это", "или",
    "ml", "kg", "gr", "lt",
})


# ─── Утилиты ──────────────────────────────────────────────────────────────────

def _norm(s: str) -> str:
    """Нижний регистр + схлопывание пробелов — для сравнения строк."""
    return re.sub(r"\s+", " ", str(s).lower().strip())


def extract_weight(s: str):
    matches = re.findall(
        r"(\d+(?:[.,]\d+)?)\s*(мл|мг|кг|л(?=[^а-яёa-z]|$)|г(?=[^а-яёa-z]|$)|kg|ml|mg|lb)",
        s, re.IGNORECASE,
    )
    if not matches:
        return None
    vals = []
    for num_str, unit in matches:
        num = float(num_str.replace(",", "."))
        u = unit.lower()
        if u == "л":       num *= 1000
        if u in ("кг","kg"): num *= 1000
        vals.append(num)
    return max(vals)


def weight_penalty(a: str, b: str) -> float:
    """Штраф за несовпадение веса/объёма. 390мл vs 250мл → сильный штраф."""
    wa, wb = extract_weight(a), extract_weight(b)
    if wa is None or wb is None:
        return 1.0
    return (min(wa, wb) / max(wa, wb)) ** 2


def _is_near_duplicate(query: str, ref_raw: str, threshold: float = 0.97) -> bool:
    """
    True если новый SKU практически идентичен уже существующему col[2].
    Такой SKU уже фактически есть в справочнике — не нужно его показывать как новый.
    threshold=0.97 — можно 1-2 символа отличаться (опечатка), но не более.
    """
    from rapidfuzz import fuzz
    return fuzz.ratio(_norm(query), _norm(ref_raw)) / 100.0 >= threshold


# ─── Построение справочных структур ───────────────────────────────────────────

def build_from_reference(ref_rows: list) -> tuple[dict, set]:
    """
    Строит из ВСЕХ строк справочника:
      brand_map   : {ключ_нижний → канонический_бренд}
      descriptors : слова, реально различающие варианты одного бренда
                    (обучается только на индикативных парах)
    """
    # Brand map — из всех строк
    brand_map: dict[str, str] = {}
    for r in ref_rows:
        canonical = str(r[1] or "").strip()
        if not canonical:
            continue
        brand_map[canonical.lower()] = canonical
        for token in re.findall(r"[а-яёa-z\w]{3,}", canonical.lower()):
            if token not in _STOP and token not in brand_map:
                brand_map[token] = canonical

    # Дескрипторы — только из индикативных пар (col[2] raw_name)
    # Слово = дескриптор, если оно есть в части SKU бренда, но не во всех
    brand_sku_words: dict[str, list] = defaultdict(list)
    for r in ref_rows:
        if str(r[4] or "") != "индикативное":
            continue
        brand    = str(r[1] or "").strip()
        raw_name = str(r[2] or "").strip()
        if not brand or not raw_name:
            continue
        tokens = frozenset(
            w for w in re.findall(r"[а-яёa-z]{3,}", raw_name.lower())
            if w not in _STOP
        )
        if tokens:
            brand_sku_words[brand.lower()].append(tokens)

    descriptors: set[str] = set()
    for brand, word_sets in brand_sku_words.items():
        if len(word_sets) < 2:
            continue
        all_words = set().union(*word_sets)
        n = len(word_sets)
        for word in all_words:
            count = sum(1 for ws in word_sets if word in ws)
            if 0 < count < n:
                descriptors.add(word)

    return brand_map, descriptors


# ─── Модель матчинга ──────────────────────────────────────────────────────────

class SKUMatcher:
    """
    TF-IDF модель, обученная на индикативных парах справочника.

    Что реально происходит при обучении (fit):
      TfidfVectorizer разбивает каждое raw-название (col[2]) на символьные
      2-4-граммы и считает IDF — насколько редок каждый фрагмент в корпусе.
      Слова-константы ("россия", "мл", "г") получают низкий вес,
      уникальные фрагменты ("лёгкий", "провансаль", "390") — высокий.

    Что происходит при поиске (transform + cosine):
      Новый SKU превращается в тот же вектор n-грамм → косинусное сходство
      со всеми обучающими векторами → топ-10 кандидатов → штрафы за вес/объём
      и дескрипторы → лучший результат.

    Почему лучше чистого Левенштейна:
      - Не зависит от порядка слов ("майонез лёгкий 390" = "390 лёгкий майонез")
      - IDF автоматически игнорирует шумовые слова без хардкода стоп-слов
      - Векторная операция — быстро даже на тысячах кандидатов
    """

    def __init__(self, ref_rows: list, descriptors: set):
        from sklearn.feature_extraction.text import TfidfVectorizer
        import numpy as np

        self.descriptors = descriptors

        # Берём только индикативные строки с непустым raw-именем (col[2])
        self._rows = [
            r for r in ref_rows
            if str(r[4] or "") == "индикативное"
            and str(r[2] or "").strip()   # ← фикс пустых строк
        ]
        self._raw_names = [_norm(str(r[2])) for r in self._rows]

        if not self._raw_names:
            self._vectorizer = None
            self._matrix     = None
            self._brand_idx  = {}
            return

        # Символьные 2-4-граммы: ловят опечатки, сокращения, частичные слова
        # char_wb — учитывает границы слов, лучше чем просто char для текста
        # sublinear_tf — log(tf+1) вместо tf: снижает влияние очень частых n-грамм
        self._vectorizer = TfidfVectorizer(
            analyzer    = "char_wb",
            ngram_range = (2, 4),
            min_df      = 1,
            sublinear_tf= True,
        )
        self._matrix = self._vectorizer.fit_transform(self._raw_names)

        # Индекс бренд → позиции строк для быстрой фильтрации
        self._brand_idx: dict[str, list[int]] = defaultdict(list)
        for i, r in enumerate(self._rows):
            brand = str(r[1] or "").strip().lower()
            if brand:
                self._brand_idx[brand].append(i)

    def find_match(
        self, query: str, csv_brand: str, threshold: float
    ) -> tuple[list | None, float]:
        """
        Находит лучшую обучающую пару для нового CSV SKU.
        Возвращает (row, score) или (None, 0.0).
        """
        # Фикс пустых строк: пустой запрос → нет результата
        if not query or not query.strip():
            return None, 0.0
        if self._vectorizer is None or not self._raw_names:
            return None, 0.0

        import numpy as np

        q_norm = _norm(query)
        if not q_norm:
            return None, 0.0

        # ── TF-IDF: новый SKU → вектор → косинусное сходство со всем корпусом
        try:
            q_vec = self._vectorizer.transform([q_norm])
        except Exception:
            return None, 0.0

        cosine = (self._matrix @ q_vec.T).toarray().flatten()

        # ── Фильтр по бренду: обнуляем кандидатов другого бренда
        brand_lower = _norm(csv_brand)
        if brand_lower and brand_lower not in ("nan", "none", ""):
            mask = np.zeros(len(self._rows), dtype=bool)
            for bkey, idxs in self._brand_idx.items():
                if bkey in brand_lower or brand_lower in bkey:
                    for i in idxs:
                        mask[i] = True
            if mask.any():
                cosine = cosine * mask
            # Если никакого бренда не нашли — не обнуляем, ищем по всем

        # ── Топ-10 кандидатов по косинусу для финального ранжирования
        top_k   = min(10, len(cosine))
        top_idx = np.argpartition(cosine, -top_k)[-top_k:]

        best_score = 0.0
        best_row   = None

        for i in top_idx:
            if cosine[i] < 0.05:   # явный мусор — пропускаем
                continue

            row = self._rows[i]
            raw = str(row[2] or "").strip()

            # Фикс пустых строк в справочнике
            if not raw:
                continue

            # Блокируем 1-в-1 дубли: SKU уже фактически есть в справочнике
            if _is_near_duplicate(query, raw):
                continue

            # Штраф за несовпадение объёма/веса (390мл vs 250мл → сильный штраф)
            wp = weight_penalty(query, raw)

            # Штраф за расхождение дескрипторов (лёгкий vs оливковый → конфликт)
            dp = self._descriptor_penalty(query, raw)

            combined = cosine[i] * wp * dp

            if combined > best_score:
                best_score = combined
                best_row   = row

        if best_score >= threshold and best_row is not None:
            return best_row, round(best_score, 2)

        return None, 0.0

    def _descriptor_penalty(self, a: str, b: str) -> float:
        if not self.descriptors:
            return 1.0
        al, bl = a.lower(), b.lower()
        da = {d for d in self.descriptors if d in al}
        db = {d for d in self.descriptors if d in bl}
        if not da and not db:
            return 1.0
        conflicts = len(da - db) + len(db - da)
        return max(0.3, 1.0 - conflicts * 0.15)


# ─── Классический матчер (Левенштейн) ────────────────────────────────────────

class ClassicMatcher:
    """
    Оригинальный алгоритм: символьный Левенштейн + штрафы за вес и дескрипторы.
    Без ML, без внешних зависимостей. Медленнее на больших справочниках,
    но предсказуем и не требует scikit-learn.
    """

    def __init__(self, ref_rows: list, descriptors: set):
        self.descriptors = descriptors
        self._rows = [
            r for r in ref_rows
            if str(r[4] or "") == "индикативное"
            and str(r[2] or "").strip()
        ]

    @staticmethod
    def _levenshtein(a: str, b: str) -> float:
        a, b = a.lower(), b.lower()
        if a == b: return 1.0
        la, lb = len(a), len(b)
        if la == 0 or lb == 0: return 0.0
        prev = list(range(lb + 1))
        for i in range(1, la + 1):
            curr = [i] + [0] * lb
            for j in range(1, lb + 1):
                curr[j] = (prev[j-1] if a[i-1] == b[j-1]
                           else 1 + min(prev[j], curr[j-1], prev[j-1]))
            prev = curr
        return 1 - prev[lb] / max(la, lb)

    def _descriptor_penalty(self, a: str, b: str) -> float:
        if not self.descriptors: return 1.0
        al, bl = a.lower(), b.lower()
        da = {d for d in self.descriptors if d in al}
        db = {d for d in self.descriptors if d in bl}
        if not da and not db: return 1.0
        conflicts = len(da - db) + len(db - da)
        return max(0.3, 1.0 - conflicts * 0.15)

    def find_match(self, query: str, csv_brand: str, threshold: float):
        if not query.strip() or not self._rows:
            return None, 0.0
        csv_brand_lower = csv_brand.strip().lower()
        best_score, best_row = 0.0, None
        for r in self._rows:
            raw = str(r[2] or "").strip()
            if not raw: continue
            if _is_near_duplicate(query, raw): continue
            if csv_brand_lower and csv_brand_lower not in ("nan", "none"):
                ind_brand = str(r[1] or "").lower()
                if ind_brand and ind_brand not in csv_brand_lower and csv_brand_lower not in ind_brand:
                    continue
            score = self._levenshtein(query, raw) * weight_penalty(query, raw) * self._descriptor_penalty(query, raw)
            if score > best_score:
                best_score, best_row = score, r
        if best_score >= threshold and best_row is not None:
            return best_row, round(best_score, 2)
        return None, 0.0


# ─── Основной пайплайн ────────────────────────────────────────────────────────

def run_matching(
    ref_path: str, csv_folder: str, threshold: float, on_progress, on_done,
    mode: str = "ml",
):
    """
    Запускается в отдельном потоке.
    on_progress(msg: str)
    on_done(results, error=None, all_new_skus=None)
    """
    try:
        import polars as pl

        # Чтение справочника через calamine — быстрее openpyxl
        try:
            df_ref = pl.read_excel(ref_path, sheet_name="SKU", engine="calamine")
        except Exception:
            df_ref = pl.read_excel(ref_path, engine="calamine")

        df_ref   = df_ref.fill_null("")
        ref_rows = df_ref.to_numpy().tolist()

        indicative_count = sum(1 for r in ref_rows if str(r[4] or "") == "индикативное")
        existing = {str(r[2] or "").strip() for r in ref_rows if str(r[2] or "").strip()}

        on_progress(
            f"Справочник: {len(ref_rows)} строк, индикативных (обучающих пар): {indicative_count}"
        )

        # Строим brand_map и дескрипторы из данных справочника
        brand_map, descriptors = build_from_reference(ref_rows)
        on_progress(f"Брендов: {len(set(brand_map.values()))} | Дескрипторов: {len(descriptors)}")

        # Строим модель в зависимости от выбранного режима
        if mode == "ml":
            try:
                matcher = SKUMatcher(ref_rows, descriptors)
                on_progress(f"Режим: ML (TF-IDF) | Обучающих пар: {len(matcher._rows)}")
            except ImportError:
                on_progress("⚠ scikit-learn не установлен → fallback на классический режим")
                matcher = ClassicMatcher(ref_rows, descriptors)
        else:
            matcher = ClassicMatcher(ref_rows, descriptors)
            on_progress(f"Режим: Классический (Левенштейн) | Пар: {len(matcher._rows)}")

        # Загрузка CSV
        csv_files = list(Path(csv_folder).glob("*.csv"))
        if not csv_files:
            on_done([], error="В папке нет CSV файлов")
            return

        frames = []
        for f in csv_files:
            try:
                df = pd.read_csv(f)
                if "pd_sku" not in df.columns:
                    on_progress(f"⚠ Пропущен {f.name} — нет pd_sku")
                    continue
                cols = [c for c in ["pd_sku", "brand", "category"] if c in df.columns]
                frames.append(df[cols].drop_duplicates("pd_sku"))
                on_progress(f"→ {f.name}: {len(df)} строк")
            except Exception as e:
                on_progress(f"⚠ Ошибка {f.name}: {e}")

        if not frames:
            on_done([], error="Ни один CSV не подошёл")
            return

        all_skus = pd.concat(frames).drop_duplicates("pd_sku")

        # Отбираем новые: не совпадают с existing col[2] ни точно, ни как дубль
        truly_new = []
        for _, row in all_skus.iterrows():
            sku = str(row["pd_sku"])
            if sku in existing:
                continue  # точное совпадение — уже в справочнике
            truly_new.append(row)

        new_skus = pd.DataFrame(truly_new) if truly_new else pd.DataFrame(columns=all_skus.columns)
        on_progress(f"Новых SKU для матчинга: {len(new_skus)}")

        # Полный список всех новых (для вкладки «Все новые SKU»)
        all_new_skus = (
            new_skus.rename(columns={"pd_sku": "SKU", "brand": "Бренд", "category": "Категория"})
            .fillna("")
            .to_dict("records")
        )

        results   = []
        total     = len(new_skus)

        for idx, (_, row) in enumerate(new_skus.iterrows()):
            csv_sku   = str(row["pd_sku"]).strip()
            csv_brand = str(row.get("brand", "")).strip()

            if not csv_sku or csv_sku.lower() in ("nan", "none", ""):
                continue

            # Матчинг через выбранную модель
            best_row, score = matcher.find_match(csv_sku, csv_brand, threshold)

            if best_row is not None:
                results.append({
                    "Категория":        str(best_row[0] or ""),
                    "Бренд":            str(best_row[1] or ""),
                    "Наименование SKU": csv_sku,
                    "Совпало с":        str(best_row[2] or ""),
                    "SKU скорр":        str(best_row[3] or ""),
                    "Статус SKU":       "индикативное",
                    "Уверенность":      score,
                })

            if idx % 50 == 0:
                on_progress(f"Обработано: {idx + 1}/{total}...")

        on_progress(f"Найдено совпадений: {len(results)}")
        results.sort(key=lambda x: -x["Уверенность"])
        on_done(results, all_new_skus=all_new_skus)

    except Exception as e:
        import traceback
        on_done([], error=f"{e}\n{traceback.format_exc()}")


# ─── Сохранение ───────────────────────────────────────────────────────────────

def save_to_reference(selected_results: list, ref_path: str) -> int:
    """Дописывает выбранные строки в лист SKU справочника. openpyxl только для записи."""
    import openpyxl
    wb = openpyxl.load_workbook(ref_path)
    ws = wb["SKU"] if "SKU" in wb.sheetnames else wb.active

    if ws.cell(row=1, column=6).value != "Уверенность совпадения":
        ws.cell(row=1, column=6).value = "Уверенность совпадения"

    next_row = ws.max_row + 1
    for r in selected_results:
        ws.cell(row=next_row, column=1).value = r.get("Категория", "")
        ws.cell(row=next_row, column=2).value = r.get("Бренд", "")
        ws.cell(row=next_row, column=3).value = r.get("Наименование SKU", "")
        ws.cell(row=next_row, column=4).value = r.get("SKU скорр", "")
        ws.cell(row=next_row, column=5).value = r.get("Статус SKU", "индикативное")
        ws.cell(row=next_row, column=6).value = r.get("Уверенность", "")
        next_row += 1

    wb.save(ref_path)
    return len(selected_results)