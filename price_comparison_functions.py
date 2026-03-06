"""
price_comparison_functions.py
Логика сравнения цен Купер vs PromoData — модуль для EFKO FlowManager.
Код идентичен standalone comparison_gui.py, адаптирован для вызова из app.py.
"""

import re
import logging
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


NETWORK_MAP = {
    "lentagp":        "Лента Гипер",
    "auchan":         "Ашан",
    "auchansm":       "Ашан",
    "metro":          "Метро",
    "okey":           "О'кей",
    "magnit_express": "Магнит (у дома)",
    "magnit":         "Магнит Семейный",
    "magnit_gmexp":   "Магнит Экстра",
    "5ka":            "Пятёрочка",
    "perekrestok":    "Перекрёсток*",
    "globusgiper":    "Глобус",
    "dixy":           "Дикси",
    "fdlenta":        "Лента Супер",
}


def extract_network(url):
    if pd.isna(url):
        return None
    m = re.search(r"kuper\.ru/([^/]+)/", str(url))
    return m.group(1) if m else None


def ml_to_g(ml):
    """Конвертирует мл → г (×0.94), округляет до 10 г."""
    return round(ml * 0.94 / 10) * 10


def normalize(s):
    s = str(s).lower().strip()
    def _ml(m):
        return f"{ml_to_g(float(m.group(1)))} г"
    s = re.sub(r"(\d+(?:\.\d+)?)\s*мл", _ml, s)
    s = re.sub(r"[^\w\s]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def extract_weight(s):
    """Возвращает вес в граммах (мл→г ×0.94, кг→г ×1000, л→мл ×1000×0.94)."""
    m = re.search(r"(\d[\d\.,]*)\s*(г|мл|кг|л)\b", str(s), re.IGNORECASE)
    if not m:
        return None
    try:
        val  = float(m.group(1).replace(",", "."))
        unit = m.group(2).lower()
        if unit == "кг": val *= 1000
        if unit == "мл": val  = ml_to_g(val)
        if unit == "л":  val  = ml_to_g(val * 1000)
        return val
    except Exception:
        return None


def weights_ok(sku_a, sku_b, tol=0.10):
    """False если оба веса известны и отличаются более чем на tol (10%)."""
    wa, wb = extract_weight(sku_a), extract_weight(sku_b)
    if wa is None or wb is None:
        return True
    return abs(wa - wb) / max(wa, wb) <= tol


def jaccard(a, b):
    if not weights_ok(a, b):
        return 0.0
    wa, wb = set(normalize(a).split()), set(normalize(b).split())
    if not wa or not wb:
        return 0.0
    return len(wa & wb) / len(wa | wb)


def run_comparison(
    path_kuper: str,
    path_promodata: str,
    path_sprav: str,
    output_file: str,
    threshold: float,
    log,
    stop_event=None,
):
    """
    Сравнивает цены Купера с PromoData и сохраняет Excel.
    log        — callable(str)
    stop_event — threading.Event или None
    Возвращает количество итоговых строк.
    """

    def stopped():
        return stop_event is not None and stop_event.is_set()

    log("▶ Загрузка файла Купера...")
    df3 = pd.read_excel(path_kuper)
    df3["Сеть"]        = df3["Url"].apply(extract_network).map(NETWORK_MAP)
    df3["sku_clean"]   = df3["Наименование SKU"].str.strip()
    df3["promo_kuper"] = pd.to_numeric(df3["Неперечеркнутая цена"], errors="coerce")
    df3["reg_kuper"]   = pd.to_numeric(
        df3["Перечеркнутая цена"].replace("Нет данных", None), errors="coerce"
    )
    df3["тип"] = df3["Тип продукта"].str.strip()

    df3 = df3[
        (df3["тип"] == "Майонез") |
        (
            (df3["тип"] == "Кетчуп") &
            df3["sku_clean"].str.lower().str.contains("томат") &
            ~df3["sku_clean"].str.lower().str.contains("шашлык|гриль|bbq|барбекю")
        )
    ]
    log(f"   Строк после фильтра (Майонез + Кетчуп томатный): {len(df3):,}")

    kuper = (
        df3[df3["Сеть"].notna()]
        .groupby(["Сеть", "sku_clean", "тип"])
        .agg(
            Регулярная_Kuper  =("reg_kuper",   lambda x: round(x.dropna().mean(), 2) if x.dropna().any() else None),
            Промо_Kuper       =("promo_kuper", lambda x: round(x.dropna().mean(), 2) if x.dropna().any() else None),
            N_рег_Kuper       =("reg_kuper",   lambda x: int(x.notna().sum())),
            N_промо_Kuper     =("promo_kuper", lambda x: int(x.notna().sum())),
        )
        .reset_index()
    )
    log(f"   Уникальных SKU×Сеть в Купере: {len(kuper):,}")

    if stopped():
        return 0

    log("▶ Загрузка PromoData (Лист3)...")
    raw = pd.read_excel(path_promodata, sheet_name="Лист3", header=None)
    date_cols = list(range(3, raw.shape[1]))
    data = raw.iloc[13:].copy()
    data.columns = range(raw.shape[1])
    data[0] = data[0].replace("(пусто)", None).ffill()
    data[1] = data[1].ffill()

    def last_value(row):
        vals = row[date_cols].dropna()
        return vals.iloc[-1] if len(vals) > 0 else None

    data["last_val"] = data.apply(last_value, axis=1)
    data["metric"]   = data[2].apply(lambda x: "regular" if "регулярная" in str(x) else "promo")

    # Кол-во периодов где цена была заполнена (не null)
    data["price_cnt"] = data[date_cols].notna().sum(axis=1)

    promo = (
        data[[0, 1, "metric", "last_val"]]
        .rename(columns={0: "SKU", 1: "Сеть"})
        .dropna(subset=["SKU", "Сеть"])
        .pivot_table(index=["SKU", "Сеть"], columns="metric", values="last_val", aggfunc="first")
        .reset_index()
    )
    promo.columns.name = None
    promo.rename(columns={"regular": "Регулярная_PromoData", "promo": "Промо_PromoData"}, inplace=True)

    # Кол-во вхождений цены по периодам
    promo_cnt = (
        data[[0, 1, "metric", "price_cnt"]]
        .rename(columns={0: "SKU", 1: "Сеть"})
        .dropna(subset=["SKU", "Сеть"])
        .pivot_table(index=["SKU", "Сеть"], columns="metric", values="price_cnt", aggfunc="first")
        .reset_index()
    )
    promo_cnt.columns.name = None
    promo_cnt.rename(columns={"regular": "N_рег_PD", "promo": "N_промо_PD"}, inplace=True)
    promo = promo.merge(promo_cnt, on=["SKU", "Сеть"], how="left")
    log(f"   Строк PromoData: {len(promo):,}")

    if stopped():
        return 0

    log("▶ Загрузка справочника SKU...")
    df_sku = pd.read_excel(path_sprav, sheet_name="SKU")
    df_sku["наим"]  = df_sku["Наименование SKU"].str.strip()
    df_sku["скорр"] = df_sku["SKU скорр"].str.strip()
    ref_map = df_sku[df_sku["скорр"].notna()].set_index("наим")["скорр"].to_dict()
    log(f"   Записей со SKU скорр: {len(ref_map):,}")

    log(f"▶ Нечёткий матч SKU (порог Жаккара ≥ {threshold})...")
    ref_names  = list(ref_map.keys())
    kuper_skus = kuper["sku_clean"].unique().tolist()
    fuzzy_map  = {}
    total = len(kuper_skus)

    for i, kp in enumerate(kuper_skus):
        if stopped():
            return 0
        best_score, best_ref = 0.0, None
        for ref in ref_names:
            s = jaccard(kp, ref)
            if s > best_score:
                best_score, best_ref = s, ref
        if best_score >= threshold:
            fuzzy_map[kp] = ref_map[best_ref]
        if (i + 1) % 100 == 0 or (i + 1) == total:
            log(f"   Обработано {i+1}/{total} SKU...")

    kuper["SKU_скорр"] = kuper["sku_clean"].map(fuzzy_map)
    kuper["SKU_скорр"] = kuper["SKU_скорр"].fillna(kuper["sku_clean"].map(ref_map))
    matched_count = kuper["SKU_скорр"].notna().sum()
    log(f"   Найдено совпадений: {matched_count:,}")

    if stopped():
        return 0

    log("▶ Сборка итоговой таблицы...")
    matched = (
        kuper[kuper["SKU_скорр"].notna()]
        .merge(promo, left_on=["Сеть", "SKU_скорр"], right_on=["Сеть", "SKU"], how="inner")
        .drop_duplicates(subset=["Сеть", "SKU_скорр"])
        .sort_values(["тип", "Сеть", "SKU_скорр"])
        .reset_index(drop=True)
    )
    matched["Δ Регулярная"] = (matched["Регулярная_Kuper"] - matched["Регулярная_PromoData"]).round(2)
    matched["Δ Промо"]      = (matched["Промо_Kuper"]      - matched["Промо_PromoData"]).round(2)
    log(f"   Итоговых строк: {len(matched):,}  (Кетчуп: {len(matched[matched['тип']=='Кетчуп'])}, Майонез: {len(matched[matched['тип']=='Майонез'])})")

    if stopped():
        return 0

    log("▶ Запись Excel...")
    _write_excel(matched, output_file)
    log(f"✅ Готово! Файл сохранён: {output_file}")
    return len(matched)


def _write_excel(matched, output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Купер vs PromoData"
    ws.freeze_panes = "A4"

    H_FILL  = PatternFill("solid", start_color="1F4E79")
    PD_FILL = PatternFill("solid", start_color="375623")
    KP_FILL = PatternFill("solid", start_color="833C00")
    DF_FILL = PatternFill("solid", start_color="44546A")
    CN_FILL = PatternFill("solid", start_color="2E4057")
    EVEN    = PatternFill("solid", start_color="EBF3FB")
    ODD     = PatternFill("solid", start_color="FFFFFF")
    RED     = PatternFill("solid", start_color="FFB3B3")
    GREEN   = PatternFill("solid", start_color="CCFFCC")
    THIN    = Border(
        left=Side(style="thin", color="BBBBBB"), right=Side(style="thin", color="BBBBBB"),
        top=Side(style="thin", color="BBBBBB"),  bottom=Side(style="thin", color="BBBBBB"),
    )
    H_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=9)
    C = Alignment(horizontal="center", vertical="center", wrap_text=True)
    L = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    R = Alignment(horizontal="right",  vertical="center")

    ws.merge_cells("A1:N1")
    ws["A1"] = (f"Купер vs PromoData  |  Майонез + Кетчуп томатный  |  "
                f"Строк: {len(matched)}  |  26.02.2026")
    ws["A1"].font      = Font(name="Arial", bold=True, size=12, color="1F4E79")
    ws["A1"].alignment = C
    ws.row_dimensions[1].height = 22

    for start, end, label, fill in [
        (1,  4,  "",                       H_FILL),
        (5,  6,  "PromoData",              PD_FILL),
        (7,  8,  "Купер",                  KP_FILL),
        (9,  10, "Δ Отклонение",           DF_FILL),
        (11, 14, "Кол-во встреч цены",     CN_FILL),
    ]:
        if start != end:
            ws.merge_cells(start_row=2, start_column=start, end_row=2, end_column=end)
        c = ws.cell(row=2, column=start, value=label)
        c.alignment = C; c.font = H_FONT; c.fill = fill
    ws.row_dimensions[2].height = 18

    for i, (h, f) in enumerate([
        ("Категория",    H_FILL), ("Сеть",           H_FILL),
        ("SKU (скорр.)", H_FILL), ("SKU Купер",       KP_FILL),
        ("Рег. PD",      PD_FILL),("Промо PD",        PD_FILL),
        ("Рег. Купер",   KP_FILL),("Промо Купер",     KP_FILL),
        ("Δ Рег.",       DF_FILL),("Δ Промо",         DF_FILL),
        ("N рег. PD",    CN_FILL),("N промо PD",      CN_FILL),
        ("N рег. Купер", CN_FILL),("N промо Купер",   CN_FILL),
    ], 1):
        c = ws.cell(row=3, column=i, value=h)
        c.font=H_FONT; c.fill=f; c.alignment=C; c.border=THIN
    ws.row_dimensions[3].height = 30

    DATA_COLS = [
        "тип", "Сеть", "SKU_скорр", "sku_clean",
        "Регулярная_PromoData", "Промо_PromoData",
        "Регулярная_Kuper",     "Промо_Kuper",
        "Δ Регулярная",         "Δ Промо",
        "N_рег_PD",             "N_промо_PD",
        "N_рег_Kuper",          "N_промо_Kuper",
    ]
    PRICE_COLS = {5, 6, 7, 8, 9, 10}
    COUNT_COLS = {11, 12, 13, 14}

    for r_idx, row in matched.iterrows():
        er   = r_idx + 4
        fill = EVEN if r_idx % 2 == 0 else ODD
        for ci, col in enumerate(DATA_COLS, 1):
            val = row.get(col, None)
            if pd.isna(val): val = None
            cell = ws.cell(row=er, column=ci, value=val)
            cell.font = Font(name="Arial", size=9); cell.border = THIN
            if ci in PRICE_COLS:
                cell.alignment = R
                if val is not None: cell.number_format = "#,##0.00"
                if ci in (9, 10) and val is not None:
                    cell.fill = RED if float(val) > 0 else (GREEN if float(val) < 0 else fill)
                else:
                    cell.fill = fill
            elif ci in COUNT_COLS:
                cell.alignment = C
                if val is not None:
                    cell.value = int(val)
                    cell.number_format = "0"
                cell.fill = fill
            else:
                cell.alignment = L; cell.fill = fill

    for i, w in enumerate([14, 17, 28, 55, 12, 12, 13, 13, 11, 11, 10, 11, 11, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    lr = len(matched) + 5
    for i, (txt, fnt) in enumerate([
        ("Легенда:", Font(name="Arial", bold=True, size=9)),
        ("  🔴 Красный Δ = Купер дороже PromoData",  Font(name="Arial", size=9, color="CC0000")),
        ("  🟢 Зелёный Δ = Купер дешевле PromoData", Font(name="Arial", size=9, color="006600")),
        ("  N рег./промо PD — сколько периодов в PromoData цена встречалась именно такой.",
         Font(name="Arial", size=9, italic=True, color="555555")),
        ("  N рег./промо Купер — сколько магазинов сети показали эту цену в выгрузке.",
         Font(name="Arial", size=9, italic=True, color="555555")),
        ("  Кетчуп: только томатный (без гриль/шашлык).",
         Font(name="Arial", size=9, italic=True, color="555555")),
    ]):
        ws.cell(row=lr + i, column=1, value=txt).font = fnt

    wb.save(output_file)