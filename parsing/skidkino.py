"""
Парсер магазинов «Скидкино» — скидкино.рф (xn--d1ahbfcthr.xn--p1ai)

Сайт на Tilda, список магазинов лежит как «товары» в Tilda Store.
Данные забираются напрямую из Tilda API без браузера:
    https://store.tildaapi.com/api/getproductslist/?storepartuid=604847375591&recid=373449275

Регион определяется:
  1) из самого адреса, если он начинается с области/республики;
  2) иначе — по городу через справочник city_and_regions_Russia.xlsx
     (файл должен лежать рядом со скриптом).

Запуск:  python skidkino_parser.py
Зависимости:  pip install requests openpyxl
"""

import re
import sys
import time
from collections import Counter
from itertools import groupby

import requests
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

OUTPUT       = "skidkino_shops.xlsx"
REGIONS_XLSX = "city_and_regions_Russia.xlsx"

API = ("https://store.tildaapi.com/api/getproductslist/"
       "?storepartuid=604847375591&recid=373449275"
       "&getparts=true&getoptions=true&size=100")

HEADERS = {
    "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                   "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"),
    "Accept": "application/json",
    "Referer": "https://xn--d1ahbfcthr.xn--p1ai/",
}

# Нормализация названий регионов из адресов (включая опечатки на сайте)
REGION_WORDS = re.compile(
    r"(область|обл[юь]?\.?|край|республика|респ\.?)", re.IGNORECASE
)

CITY_PREFIX = re.compile(r"^(р\.?\s*п|пгт|рп|г|п|с|д|ст)\s*\.?\s*", re.IGNORECASE)


def fetch_shops() -> list[dict]:
    all_prods, slice_n = [], 1
    while True:
        url = API + f"&slice={slice_n}"
        for attempt in range(1, 4):
            try:
                r = requests.get(url, headers=HEADERS, timeout=30)
                r.raise_for_status()
                d = r.json()
                break
            except Exception as e:
                print(f"  попытка {attempt}: {e}")
                time.sleep(2)
        else:
            return all_prods
        prods = d.get("products", [])
        all_prods.extend(prods)
        print(f"  страница {slice_n}: +{len(prods)} (всего {len(all_prods)}/{d.get('total')})")
        if not d.get("nextslice") or not prods:
            break
        slice_n = d["nextslice"]
    return all_prods


def load_city2region(path: str) -> dict:
    mapping = {}
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
        for row in wb.active.iter_rows(min_row=2, values_only=True):
            city, region = row[1], row[2]
            if city and region:
                key = str(city).strip().lower().replace("ё", "е")
                mapping[key] = str(region).strip()
        print(f"  Справочник регионов: {len(mapping)} городов")
    except FileNotFoundError:
        print(f"  ⚠ Файл {path} не найден — регионы только из адресов")
    return mapping


def normalize_region(raw: str) -> str:
    """'Ульяновская обл.' → 'Ульяновская область', 'Мордовия' → 'Республика Мордовия'."""
    s = raw.strip().rstrip(".").strip()
    s = re.sub(r"\bобл[юь]?\.?$", "область", s, flags=re.IGNORECASE)
    s = re.sub(r"\bресп\.?\b", "Республика", s, flags=re.IGNORECASE)
    if re.fullmatch(r"мордовия", s, re.IGNORECASE):
        s = "Республика Мордовия"
    return s


def parse_title(title: str, city2region: dict) -> dict:
    """Разбирает 'Область, г. Город, ул. Улица, 1' на регион/город/адрес."""
    # Опечатка на сайте: 'Мордовия. г. Ковылкино' (точка вместо запятой)
    t = re.sub(r"^(Мордовия)\.\s*", r"\1, ", title.strip())
    parts = [p.strip() for p in t.split(",") if p.strip()]

    region, city = "", ""
    if parts and (REGION_WORDS.search(parts[0])
                  or re.fullmatch(r"мордовия", parts[0], re.IGNORECASE)):
        region = normalize_region(parts.pop(0))
    if parts and (CITY_PREFIX.match(parts[0]) or not region):
        city = CITY_PREFIX.sub("", parts.pop(0)).strip()

    if not region and city:
        region = city2region.get(city.lower().replace("ё", "е"), "")

    return {
        "region":  region,
        "city":    city,
        "address": ", ".join(parts),
        "full":    title.strip(),
    }


def normalize(raw: list[dict], city2region: dict) -> list[dict]:
    rows = []
    for p in raw:
        row = parse_title(p.get("title", ""), city2region)
        descr = (p.get("descr") or "").strip()
        row["hours"] = re.sub(r"^Режим работы:\s*", "", descr)
        rows.append(row)
    return rows


def save_excel(rows: list[dict], path: str) -> None:
    wb = openpyxl.Workbook()
    thin       = Side(style="thin", color="CCCCCC")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill   = PatternFill("solid", fgColor="E97B00")
    alt_fill   = PatternFill("solid", fgColor="FDF3E7")
    tot_fill   = PatternFill("solid", fgColor="FAE3C8")
    grand_fill = PatternFill("solid", fgColor="F5CE9E")
    hdr_font   = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    dat_font   = Font(name="Arial", size=10)
    tot_font   = Font(bold=True, italic=True, name="Arial", size=10)
    grand_font = Font(bold=True, name="Arial", size=11)

    def hcell(ws, r, c, v, w=None):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if w: ws.column_dimensions[cell.column_letter].width = w

    def dcell(ws, r, c, v, align="left", fill=None, font=None, wrap=False):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = font or dat_font; cell.border = border
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        if fill: cell.fill = fill

    # ── Лист 1: Итог по регионам ─────────────────────────────────────────────
    ws1 = wb.active; ws1.title = "Итог по регионам"
    ws1.row_dimensions[1].height = 22
    hcell(ws1, 1, 1, "Область / Регион", w=34)
    hcell(ws1, 1, 2, "Кол-во ТТ", w=12)
    ws1.freeze_panes = "A2"

    counts = Counter(r["region"] or "Регион не определён" for r in rows)
    total = 0
    for ri, (region, cnt) in enumerate(sorted(counts.items(), key=lambda x: -x[1]), 2):
        fill = alt_fill if ri % 2 == 0 else None
        dcell(ws1, ri, 1, region, fill=fill)
        dcell(ws1, ri, 2, cnt, fill=fill, align="right")
        total += cnt
    ri_g = len(counts) + 2
    dcell(ws1, ri_g, 1, "ИТОГО", font=grand_font, fill=grand_fill)
    dcell(ws1, ri_g, 2, total, font=grand_font, fill=grand_fill, align="right")
    ws1.auto_filter.ref = f"A1:B{ri_g - 1}"

    # ── Лист 2: Все адреса ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("Все адреса")
    ws2.row_dimensions[1].height = 22
    for ci, (t, w) in enumerate([
        ("Регион", 28), ("Город / нас. пункт", 24), ("Адрес", 40),
        ("Часы работы", 22), ("Полный адрес (как на сайте)", 52)
    ], 1):
        hcell(ws2, 1, ci, t, w=w)
    ws2.freeze_panes = "A2"
    NCOLS = 5

    sorted_rows = sorted(rows, key=lambda r: (r["region"], r["city"], r["address"]))
    row_num = 2; sheet_total = 0

    for region, grp in groupby(sorted_rows, key=lambda r: r["region"]):
        grp = list(grp)
        for r in grp:
            fill = alt_fill if row_num % 2 == 0 else None
            dcell(ws2, row_num, 1, r["region"], fill=fill)
            dcell(ws2, row_num, 2, r["city"], fill=fill)
            dcell(ws2, row_num, 3, r["address"], fill=fill, wrap=True)
            dcell(ws2, row_num, 4, r["hours"], fill=fill, align="center")
            dcell(ws2, row_num, 5, r["full"], fill=fill, wrap=True)
            row_num += 1
        for ci in range(1, NCOLS + 1):
            c = ws2.cell(row=row_num, column=ci)
            c.fill = tot_fill; c.border = border; c.font = tot_font
            c.alignment = Alignment(vertical="center")
        ws2.cell(row=row_num, column=1).value = f"Итого по «{region or 'без региона'}»"
        ws2.cell(row=row_num, column=2).value = f"{len(grp)} ТТ"
        ws2.cell(row=row_num, column=2).alignment = Alignment(
            horizontal="center", vertical="center")
        row_num += 1; sheet_total += len(grp)

    for ci in range(1, NCOLS + 1):
        c = ws2.cell(row=row_num, column=ci)
        c.fill = grand_fill; c.border = border; c.font = grand_font
        c.alignment = Alignment(vertical="center")
    ws2.cell(row=row_num, column=1).value = "ИТОГО ПО ВСЕМ РЕГИОНАМ"
    ws2.cell(row=row_num, column=2).value = f"{sheet_total} ТТ"
    ws2.cell(row=row_num, column=2).alignment = Alignment(
        horizontal="center", vertical="center")
    ws2.auto_filter.ref = f"A1:E{row_num - 1}"
    wb.save(path)


def main():
    print("=" * 55)
    print("  Парсер «Скидкино» (скидкино.рф)")
    print("=" * 55 + "\n")

    t0 = time.time()
    print("Запрашиваем список магазинов из Tilda API...")
    raw = fetch_shops()
    if not raw:
        print("ОШИБКА: данные не получены.")
        sys.exit(1)

    city2region = load_city2region(REGIONS_XLSX)
    rows = normalize(raw, city2region)

    unmatched = sorted({r["full"] for r in rows if not r["region"]})
    if unmatched:
        print(f"  ⚠ Не определён регион ({len(unmatched)}):")
        for u in unmatched:
            print(f"     {u}")

    counts = Counter(r["region"] or "Регион не определён" for r in rows)
    print(f"\n{'─' * 55}")
    for region, count in sorted(counts.items(), key=lambda x: -x[1]):
        print(f"  ✓ {region}: {count} ТТ")
    print(f"{'─' * 55}")
    print(f"  ИТОГО: {len(counts)} регионов, {len(rows)} ТТ")
    print(f"{'─' * 55}\n")

    save_excel(rows, OUTPUT)
    print(f"✓ Готово за {time.time() - t0:.1f} сек → {OUTPUT}")


if __name__ == "__main__":
    main()