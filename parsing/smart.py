"""
Парсер магазинов Smart (Сладкая жизнь НН) — smart.swnn.ru

API — открытый 1С HTTP-сервис, авторизация не нужна.
Все магазины отдаются одним запросом:
    GET https://smart.swnn.ru/WS/hs/exchange/getStores/1/1/1?noauth

Регион определяется по справочнику city_and_regions_Russia.xlsx
(файл должен лежать рядом со скриптом).

Запуск:  python smart_parser.py
Зависимости:  pip install requests openpyxl
"""

import sys
import time
from collections import Counter
from itertools import groupby

import requests
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

OUTPUT   = "smart_shops.xlsx"
REGIONS_XLSX = "city_and_regions_Russia.xlsx"

# Города, которых нет в справочнике или которые пишутся иначе
CITY_FIXES = {
    "голубое":         "Московская область",   # посёлок, г.о. Солнечногорск
    "сходня":          "Московская область",   # мкр. Химок
    "починки":         "Нижегородская область",
    "семёнов":         "Нижегородская область",
    "ростов великий":  "Ярославская область",
    # Тёзки: в справочнике Радужный = ХМАО, но у Smart это ЗАТО под Владимиром
    # (координаты магазина 56.0, 40.3 — Владимирская область)
    "радужный":        "Владимирская область",
}
BASE     = "https://smart.swnn.ru/WS/hs/exchange/"
HEADERS  = {
    "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                   "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"),
    "Accept": "application/json",
    "Referer": "https://smart.swnn.ru/stores",
}


def fetch_shops() -> list[dict]:
    url = BASE + "getStores/1/1/1?noauth"
    for attempt in range(1, 4):
        try:
            r = requests.get(url, headers=HEADERS, timeout=30)
            r.raise_for_status()
            data = r.json()
            if data.get("Success") and data.get("Data", {}).get("Дискаунтеры"):
                return data["Data"]["Дискаунтеры"]
            print(f"  попытка {attempt}: неожиданный ответ: {str(data)[:150]}")
        except Exception as e:
            print(f"  попытка {attempt}: {e}")
        time.sleep(2)
    return []


def load_city2region(path: str) -> dict:
    """Справочник город → регион. Ключи в нижнем регистре, ё → е."""
    mapping = {}
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            city, region = row[1], row[2]
            if city and region:
                key = str(city).strip().lower().replace("ё", "е")
                mapping[key] = str(region).strip()
        print(f"  Справочник регионов: {len(mapping)} городов")
    except FileNotFoundError:
        print(f"  ⚠ Файл {path} не найден — колонка «Регион» будет пустой")
    return mapping


def region_for(city: str, city2region: dict) -> str:
    key = city.strip().lower()
    if key in CITY_FIXES:
        return CITY_FIXES[key]
    return city2region.get(key.replace("ё", "е"), "")


def normalize(raw: list[dict], city2region: dict) -> list[dict]:
    rows = []
    for s in raw:
        city = (s.get("НаселенныйПункт") or "").strip()
        rows.append({
            "region":  region_for(city, city2region),
            "city":    city,
            "name":    (s.get("Наименование") or "").strip(),   # полный адрес с городом
            "address": (s.get("Адрес") or "").strip(),
            "hours":   (s.get("ВремяРаботы") or "").strip(),
            "lat":     s.get("Широта", ""),
            "lon":     s.get("Долгота", ""),
        })
    return rows


def save_excel(rows: list[dict], path: str) -> None:
    wb = openpyxl.Workbook()
    thin       = Side(style="thin", color="CCCCCC")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill   = PatternFill("solid", fgColor="5032C8")
    alt_fill   = PatternFill("solid", fgColor="F0EDFB")
    tot_fill   = PatternFill("solid", fgColor="DCD6F5")
    grand_fill = PatternFill("solid", fgColor="C3B8EE")
    hdr_font   = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    dat_font   = Font(name="Arial", size=10)
    tot_font   = Font(bold=True, italic=True, name="Arial", size=10)
    grand_font = Font(bold=True, name="Arial", size=11)

    def hcell(ws, r, c, v, w=None):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if w: ws.column_dimensions[cell.column_letter].width = w

    def dcell(ws, r, c, v, align="left", fill=None, font=None, wrap=False):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = font or dat_font; cell.border = border
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        if fill: cell.fill = fill

    # ── Лист 1: Итог по регионам ─────────────────────────────────────────────
    ws1 = wb.active; ws1.title = "Итог по регионам"
    ws1.row_dimensions[1].height = 22
    hcell(ws1, 1, 1, "Область / Регион", w=34)
    hcell(ws1, 1, 2, "Кол-во ТТ", w=12)
    ws1.freeze_panes = "A2"

    counts = Counter(r["region"] or "Регион не определён" for r in rows)
    total = 0
    for ri, (city, cnt) in enumerate(sorted(counts.items(), key=lambda x: -x[1]), 2):
        fill = alt_fill if ri % 2 == 0 else None
        dcell(ws1, ri, 1, city, fill=fill)
        dcell(ws1, ri, 2, cnt, fill=fill, align="right")
        total += cnt
    ri_g = len(counts) + 2
    dcell(ws1, ri_g, 1, "ИТОГО", font=grand_font, fill=grand_fill)
    dcell(ws1, ri_g, 2, total, font=grand_font, fill=grand_fill, align="right")
    ws1.auto_filter.ref = f"A1:B{ri_g - 1}"

    # ── Лист 2: Все адреса ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("Все адреса")
    ws2.row_dimensions[1].height = 22
    for ci, (t, w) in enumerate([
        ("Регион", 28), ("Город", 22), ("Полный адрес", 50),
        ("Часы работы", 28), ("Широта", 12), ("Долгота", 12)
    ], 1):
        hcell(ws2, 1, ci, t, w=w)
    ws2.freeze_panes = "A2"
    NCOLS = 6

    sorted_rows = sorted(rows, key=lambda r: (r["region"], r["city"], r["address"]))
    row_num = 2; sheet_total = 0

    for region, grp in groupby(sorted_rows, key=lambda r: r["region"]):
        grp = list(grp)
        for r in grp:
            fill = alt_fill if row_num % 2 == 0 else None
            dcell(ws2, row_num, 1, r["region"], fill=fill)
            dcell(ws2, row_num, 2, r["city"], fill=fill)
            dcell(ws2, row_num, 3, r["name"], fill=fill, wrap=True)
            dcell(ws2, row_num, 4, r["hours"], fill=fill, align="center")
            dcell(ws2, row_num, 5, r["lat"], fill=fill, align="center")
            dcell(ws2, row_num, 6, r["lon"], fill=fill, align="center")
            row_num += 1
        for ci in range(1, NCOLS + 1):
            c = ws2.cell(row=row_num, column=ci)
            c.fill = tot_fill; c.border = border; c.font = tot_font
            c.alignment = Alignment(vertical="center")
        ws2.cell(row=row_num, column=1).value = f"Итого по «{region or 'без региона'}»"
        ws2.cell(row=row_num, column=2).value = f"{len(grp)} ТТ"
        ws2.cell(row=row_num, column=2).alignment = Alignment(
            horizontal="center", vertical="center")
        row_num += 1; sheet_total += len(grp)

    for ci in range(1, NCOLS + 1):
        c = ws2.cell(row=row_num, column=ci)
        c.fill = grand_fill; c.border = border; c.font = grand_font
        c.alignment = Alignment(vertical="center")
    ws2.cell(row=row_num, column=1).value = "ИТОГО ПО ВСЕМ РЕГИОНАМ"
    ws2.cell(row=row_num, column=2).value = f"{sheet_total} ТТ"
    ws2.cell(row=row_num, column=2).alignment = Alignment(
        horizontal="center", vertical="center")
    ws2.auto_filter.ref = f"A1:F{row_num - 1}"
    wb.save(path)


def main():
    print("=" * 55)
    print("  Парсер Smart (smart.swnn.ru)")
    print("=" * 55 + "\n")

    t0 = time.time()
    print("Запрашиваем список магазинов...")
    raw = fetch_shops()
    if not raw:
        print("ОШИБКА: данные не получены.")
        sys.exit(1)

    city2region = load_city2region(REGIONS_XLSX)
    rows = normalize(raw, city2region)

    unmatched = sorted({r["city"] for r in rows if not r["region"]})
    if unmatched:
        print(f"  ⚠ Не определён регион для городов: {unmatched}")

    counts = Counter(r["region"] or "Регион не определён" for r in rows)

    print(f"\n{'─' * 55}")
    for region, count in sorted(counts.items(), key=lambda x: -x[1]):
        print(f"  ✓ {region}: {count} ТТ")
    print(f"{'─' * 55}")
    print(f"  ИТОГО: {len(counts)} регионов, {len(rows)} ТТ")
    print(f"{'─' * 55}\n")

    save_excel(rows, OUTPUT)
    print(f"✓ Готово за {time.time() - t0:.1f} сек → {OUTPUT}")


if __name__ == "__main__":
    main()