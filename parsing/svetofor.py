"""
Парсер магазинов «Светофор» — официальный сайт сети svetofor-nsk.ru

Почему не svetoformagazin.com: тот сайт закрылся антиботом (KillBot) и
отдаёт страницу проверки вместо данных, причём с кодом 200 — старый парсер
из-за этого молча получал «нет данных» по всем регионам.

Здесь используется официальный сайт сети. Структура: округ → регион →
карточки магазинов. Список регионов берётся из меню самого сайта, поэтому
локальные html-файлы больше не нужны и новый регион подхватится сам.

Установка:
    pip install requests beautifulsoup4 openpyxl

Результат:
    svetofor_shops.xlsx    — итог по регионам + все адреса
    svetofor_by_region.json
"""

import re
import json
import time
import random
from collections import Counter, defaultdict

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

BASE = "https://svetofor-nsk.ru"
INDEX_URL = f"{BASE}/shops.html"
OUTPUT_XLSX = "svetofor_shops.xlsx"
OUTPUT_JSON = "svetofor_by_region.json"

DELAY_MIN = 0.8
DELAY_MAX = 1.8

SESSION = requests.Session()
SESSION.headers.update({
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) "
                  "Chrome/124.0.0.0 Safari/537.36",
    "Accept-Language": "ru-RU,ru;q=0.9",
})

# Ссылка на страницу региона:      /shops/szfo/pskov.html
RE_REGION = re.compile(r"^/shops/([a-z]+)/([a-z0-9\-]+)\.html$")
# Ссылка на карточку магазина:     /shops/szfo/pskov/pskov-3.html
RE_SHOP = re.compile(r"^/shops/([a-z]+)/([a-z0-9\-]+)/([a-z0-9\-]+)\.html$")

OKRUG_NAMES = {
    "cfo": "Центральный ФО",
    "szfo": "Северо-Западный ФО",
    "ufo": "Южный ФО",
    "skfo": "Северо-Кавказский ФО",
    "pfo": "Приволжский ФО",
    "urfo": "Уральский ФО",
    "sibfo": "Сибирский ФО",
    "dvfo": "Дальневосточный ФО",
}

# Признаки закрытой точки — такие карточки на сайте остаются висеть
CLOSED_KW = ["магазин закрыт", "закрыт с", "не работает", "временно закрыт"]


def pause():
    time.sleep(random.uniform(DELAY_MIN, DELAY_MAX))


def get(url, retries=3):
    for attempt in range(1, retries + 1):
        try:
            r = SESSION.get(url, timeout=25)
            if r.status_code == 200:
                r.encoding = "utf-8"
                return r.text
            if r.status_code in (403, 429, 503):
                time.sleep(6 * attempt)
            else:
                return None
        except Exception:
            time.sleep(3 * attempt)
    return None


def _path(href: str) -> str:
    """Приводит ссылку к виду /shops/... независимо от абсолютности."""
    if not href:
        return ""
    href = href.split("?")[0].split("#")[0]
    if href.startswith("http"):
        href = re.sub(r"^https?://[^/]+", "", href)
    return href


# ═══════════════════════════════════════════════════════════════════════════
# 1. СПИСОК РЕГИОНОВ ИЗ МЕНЮ САЙТА
# ═══════════════════════════════════════════════════════════════════════════

def load_regions():
    html = get(INDEX_URL)
    if not html:
        return []

    soup = BeautifulSoup(html, "html.parser")
    seen, regions = set(), []

    for a in soup.find_all("a", href=True):
        path = _path(a["href"])
        m = RE_REGION.match(path)
        if not m:
            continue
        okrug_code, _slug = m.group(1), m.group(2)
        if okrug_code not in OKRUG_NAMES:
            continue
        # Страница самого округа: /shops/szfo.html — сюда не попадает,
        # т.к. RE_REGION требует два уровня вложенности.
        name = a.get_text(strip=True)
        if not name or name.lower() == "back":
            continue
        if path in seen:
            continue
        seen.add(path)
        regions.append({
            "name": name,
            "okrug": OKRUG_NAMES[okrug_code],
            "url": BASE + path,
        })

    return regions


# ═══════════════════════════════════════════════════════════════════════════
# 2. КАРТОЧКИ МАГАЗИНОВ НА СТРАНИЦЕ РЕГИОНА
# ═══════════════════════════════════════════════════════════════════════════

def parse_region_page(html, region_slug=None):
    """
    Ищем ссылки на карточки магазинов — так подсчёт не зависит от вёрстки.
    Город и адрес достаём из блока вокруг ссылки; если разметка изменится,
    количество всё равно посчитается верно, потеряются только детали.
    """
    soup = BeautifulSoup(html, "html.parser")
    shops, seen = [], set()

    for a in soup.find_all("a", href=True):
        path = _path(a["href"])
        m = RE_SHOP.match(path)
        if not m:
            continue
        if region_slug and m.group(2) != region_slug:
            continue
        if path in seen:
            continue
        seen.add(path)

        city, address, closed = "", "", False

        card = a
        for _ in range(4):
            card = card.parent
            if card is None:
                break
            head = card.find(["h2", "h3", "h4", "h5"])
            if head:
                city = head.get_text(strip=True)
                break

        if card is not None:
            text = card.get_text("\n", strip=True)
            low = text.lower()
            closed = any(kw in low for kw in CLOSED_KW)
            for line in text.split("\n"):
                line = line.strip()
                if not line or line == city:
                    continue
                if line.lower().startswith(("подробнее", "вакансии")):
                    continue
                if re.match(r"^\d{2}[.:]\d{2}", line):   # часы работы
                    continue
                address = line
                break

        city = re.sub(r"^(г\.|пгт\.?|с\.|п\.)\s*", "", city).strip()
        shops.append({"city": city, "address": address,
                      "closed": closed, "url": BASE + path})

    return shops


# ═══════════════════════════════════════════════════════════════════════════
# 3. ОБХОД
# ═══════════════════════════════════════════════════════════════════════════

def collect_all(regions, limit=None):
    rows = []
    total = len(regions) if limit is None else min(limit, len(regions))

    for i, reg in enumerate(regions[:total], 1):
        print(f"[{i:2}/{total}] {reg['name']:<40}", end=" ", flush=True)

        html = get(reg["url"])
        if not html:
            print("ОШИБКА загрузки")
            pause()
            continue

        slug = RE_REGION.match(_path(reg["url"].replace(BASE, ""))).group(2)
        shops = parse_region_page(html, slug)

        if not shops:
            print("– нет данных")
        else:
            closed = sum(1 for s in shops if s["closed"])
            active = len(shops) - closed
            note = f" (+{closed} закрытых)" if closed else ""
            print(f"{active} ТТ{note}")

        for s in shops:
            rows.append({
                "okrug": reg["okrug"],
                "region": reg["name"],
                "city": s["city"],
                "address": s["address"],
                "closed": s["closed"],
            })

        pause()

    return rows


# ═══════════════════════════════════════════════════════════════════════════
# 4. EXCEL
# ═══════════════════════════════════════════════════════════════════════════

def save_excel(rows, path):
    wb = openpyxl.Workbook()
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    alt_fill = PatternFill("solid", fgColor="EBF3FB")
    grand_fill = PatternFill("solid", fgColor="BDD7EE")
    hdr_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    dat_font = Font(name="Arial", size=10)
    grand_font = Font(bold=True, name="Arial", size=11)

    def hcell(ws, row, col, val, width=None):
        c = ws.cell(row=row, column=col, value=val)
        c.font = hdr_font
        c.fill = hdr_fill
        c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center")
        if width:
            ws.column_dimensions[c.column_letter].width = width

    def dcell(ws, row, col, val, align="left", fill=None, font=None, wrap=False):
        c = ws.cell(row=row, column=col, value=val)
        c.font = font or dat_font
        c.border = border
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        if fill:
            c.fill = fill

    active = [r for r in rows if not r["closed"]]

    # ── Лист 1: итог по регионам ──
    ws1 = wb.active
    ws1.title = "Итог по регионам"
    ws1.row_dimensions[1].height = 22
    for ci, (t, w) in enumerate(
            [("Округ", 24), ("Регион", 40), ("Кол-во ТТ", 12), ("Закрытых", 11)], 1):
        hcell(ws1, 1, ci, t, width=w)
    ws1.freeze_panes = "A2"

    by_region = Counter((r["okrug"], r["region"]) for r in active)
    closed_by_region = Counter(
        (r["okrug"], r["region"]) for r in rows if r["closed"])

    total = 0
    ri = 2
    for (okrug, region), cnt in sorted(by_region.items(), key=lambda x: -x[1]):
        fill = alt_fill if ri % 2 == 0 else None
        dcell(ws1, ri, 1, okrug, fill=fill)
        dcell(ws1, ri, 2, region, fill=fill)
        dcell(ws1, ri, 3, cnt, fill=fill, align="right")
        dcell(ws1, ri, 4, closed_by_region.get((okrug, region), 0),
              fill=fill, align="right")
        total += cnt
        ri += 1

    dcell(ws1, ri, 1, "ИТОГО", font=grand_font, fill=grand_fill)
    dcell(ws1, ri, 2, f"{len(by_region)} регионов", font=grand_font, fill=grand_fill)
    dcell(ws1, ri, 3, total, font=grand_font, fill=grand_fill, align="right")
    dcell(ws1, ri, 4, sum(closed_by_region.values()),
          font=grand_font, fill=grand_fill, align="right")
    ws1.auto_filter.ref = f"A1:D{ri - 1}"

    # ── Лист 2: итог по округам ──
    ws3 = wb.create_sheet("Итог по округам")
    for ci, (t, w) in enumerate([("Округ", 28), ("Регионов", 12), ("Кол-во ТТ", 12)], 1):
        hcell(ws3, 1, ci, t, width=w)
    by_okrug = Counter(r["okrug"] for r in active)
    regions_in_okrug = defaultdict(set)
    for r in active:
        regions_in_okrug[r["okrug"]].add(r["region"])
    ri = 2
    for okrug, cnt in sorted(by_okrug.items(), key=lambda x: -x[1]):
        fill = alt_fill if ri % 2 == 0 else None
        dcell(ws3, ri, 1, okrug, fill=fill)
        dcell(ws3, ri, 2, len(regions_in_okrug[okrug]), fill=fill, align="right")
        dcell(ws3, ri, 3, cnt, fill=fill, align="right")
        ri += 1
    dcell(ws3, ri, 1, "ИТОГО", font=grand_font, fill=grand_fill)
    dcell(ws3, ri, 2, len(by_region), font=grand_font, fill=grand_fill, align="right")
    dcell(ws3, ri, 3, total, font=grand_font, fill=grand_fill, align="right")

    # ── Лист 3: все адреса ──
    ws2 = wb.create_sheet("Все адреса")
    ws2.row_dimensions[1].height = 22
    for ci, (t, w) in enumerate(
            [("Округ", 24), ("Регион", 36), ("Город / НП", 26),
             ("Адрес", 60), ("Статус", 12)], 1):
        hcell(ws2, 1, ci, t, width=w)
    ws2.freeze_panes = "A2"

    row_num = 2
    for r in sorted(rows, key=lambda x: (x["okrug"], x["region"], x["city"])):
        fill = alt_fill if row_num % 2 == 0 else None
        dcell(ws2, row_num, 1, r["okrug"], fill=fill)
        dcell(ws2, row_num, 2, r["region"], fill=fill)
        dcell(ws2, row_num, 3, r["city"], fill=fill)
        dcell(ws2, row_num, 4, r["address"], fill=fill, wrap=True)
        dcell(ws2, row_num, 5, "закрыт" if r["closed"] else "работает",
              fill=fill, align="center")
        row_num += 1
    ws2.auto_filter.ref = f"A1:E{row_num - 1}"

    wb.save(path)


# ═══════════════════════════════════════════════════════════════════════════
# 5. MAIN
# ═══════════════════════════════════════════════════════════════════════════

def main(limit=None):
    print("=" * 55)
    print("  Парсер Светофор — svetofor-nsk.ru (сайт сети)")
    print("=" * 55 + "\n")

    print("Читаем список регионов из меню сайта...")
    regions = load_regions()
    if not regions:
        print("ОШИБКА: регионы не найдены. Проверьте доступность сайта.")
        raise SystemExit(1)
    print(f"  → Найдено регионов: {len(regions)}\n")

    rows = collect_all(regions, limit=limit)

    active = [r for r in rows if not r["closed"]]
    closed = len(rows) - len(active)
    by_region = Counter(r["region"] for r in active)

    print(f"\n{'─' * 55}")
    for region, count in sorted(by_region.items(), key=lambda x: -x[1]):
        print(f"  {region}: {count} ТТ")
    print(f"{'─' * 55}")
    print(f"  ИТОГО: {len(by_region)} регионов, {len(active)} действующих ТТ")
    if closed:
        print(f"  Закрытых точек на сайте: {closed} (в подсчёт не входят)")
    print(f"{'─' * 55}\n")

    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump({r: c for r, c in by_region.items()}, f,
                  ensure_ascii=False, indent=2)

    save_excel(rows, OUTPUT_XLSX)
    print(f"✓ Готово → {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()