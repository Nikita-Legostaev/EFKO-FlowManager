"""
Парсер магазинов «Семишагофф» — semishagoff.org/shops/
Скачивает страницу напрямую, парсит по районам.
Регион для всех точек: Ленинградская область (включая СПб).

Установка:
    pip install requests beautifulsoup4 openpyxl

Запуск:
    python semishagoff_parser.py
"""

import sys
from collections import Counter
from itertools import groupby

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

URL    = "https://semishagoff.org/shops/"
OUTPUT = "semishagoff_shops.xlsx"
REGION = "Ленинградская область"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "ru-RU,ru;q=0.9",
}


# ═══════════════════════════════════════════════════════════════════════════════
# 1. ПАРСИНГ
# ═══════════════════════════════════════════════════════════════════════════════

def fetch_and_parse() -> list[dict]:
    print(f"Загружаем {URL} ...")
    r = requests.get(URL, headers=HEADERS, timeout=30)
    r.raise_for_status()
    r.encoding = "utf-8"
    print(f"  → {len(r.text):,} символов, статус {r.status_code}")
    return parse(r.text)


def parse(html: str) -> list[dict]:
    soup = BeautifulSoup(html, "html.parser")
    results = []
    current_district = "Неизвестный район"

    # Ищем секцию со списком магазинов (оба блока: обычные + 24ч)
    # Проходим по всем элементам последовательно
    for el in soup.find_all(["h2", "h3", "a"]):

        # Заголовок района
        if el.name in ("h2", "h3"):
            text = el.get_text(strip=True)
            if any(kw in text for kw in [
                "район", "Адмиралтейский", "Василеостровский", "Выборгский",
                "Калининский", "Кировский", "Красногвардейский", "Красносельский",
                "Колпинский", "Московский", "Невский", "Петроградский",
                "Петродворцовый", "Приморский", "Пушкинский", "Тосненский",
                "Фрунзенский", "Центральный", "Кронштадтский", "Курортный",
                "Киришский", "Волховский", "Гатчинский", "Всеволожский",
                "Волосовский", "Ломоносовский"
            ]):
                current_district = text
            continue

        # Карточка магазина — ссылка вида /shops/slug/
        if el.name == "a":
            href = el.get("href", "")
            if not (href.startswith("/shops/") and href != "/shops/"):
                continue

            lines = [ln.strip() for ln in el.get_text("\n").split("\n") if ln.strip()]
            if len(lines) < 2:
                continue

            # lines[0]  = короткое название (улица)
            # lines[-2] = полный адрес
            # lines[-1] = часы работы
            short_name   = lines[0]
            full_address = lines[-2] if len(lines) >= 3 else lines[0]
            hours        = lines[-1]

            # Пропускаем дублирующиеся записи (страница дублирует 24ч)
            key = (current_district, full_address)
            if key in {(r["district"], r["full_address"]) for r in results}:
                continue

            results.append({
                "region":       REGION,
                "district":     current_district,
                "name":         short_name,
                "full_address": full_address,
                "hours":        hours,
                "url":          "https://semishagoff.org" + href,
            })

    return results


# ═══════════════════════════════════════════════════════════════════════════════
# 2. EXCEL
# ═══════════════════════════════════════════════════════════════════════════════

def save_excel(rows: list[dict], path: str) -> None:
    wb = openpyxl.Workbook()

    thin       = Side(style="thin", color="CCCCCC")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill   = PatternFill("solid", fgColor="1F4E79")
    alt_fill   = PatternFill("solid", fgColor="EBF3FB")
    tot_fill   = PatternFill("solid", fgColor="D6E4F0")
    grand_fill = PatternFill("solid", fgColor="BDD7EE")
    hdr_font   = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    dat_font   = Font(name="Arial", size=10)
    tot_font   = Font(bold=True, italic=True, name="Arial", size=10)
    grand_font = Font(bold=True, name="Arial", size=11)

    def hcell(ws, row, col, val, width=None):
        c = ws.cell(row=row, column=col, value=val)
        c.font = hdr_font; c.fill = hdr_fill; c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center")
        if width:
            ws.column_dimensions[c.column_letter].width = width
        return c

    def dcell(ws, row, col, val, align="left", fill=None, font=None, wrap=False):
        c = ws.cell(row=row, column=col, value=val)
        c.font = font or dat_font; c.border = border
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        if fill: c.fill = fill
        return c

    # ── Лист 1: Итог по регионам (один регион — Ленинградская область) ────────
    ws1 = wb.active
    ws1.title = "Итог по регионам"
    ws1.row_dimensions[1].height = 22
    hcell(ws1, 1, 1, "Область / Регион", width=38)
    hcell(ws1, 1, 2, "Кол-во ТТ",        width=14)
    ws1.freeze_panes = "A2"

    total = len(rows)
    dcell(ws1, 2, 1, REGION, fill=alt_fill)
    dcell(ws1, 2, 2, total,  align="right", fill=alt_fill)
    dcell(ws1, 3, 1, "ИТОГО", font=grand_font, fill=grand_fill)
    dcell(ws1, 3, 2, total,  align="right", font=grand_font, fill=grand_fill)

    # ── Лист 2: Все адреса, сгруппированные по районам ───────────────────────
    ws2 = wb.create_sheet("Все адреса")
    ws2.row_dimensions[1].height = 22
    COLS = [
        ("Регион",        22),
        ("Район",         28),
        ("Название",      30),
        ("Полный адрес",  52),
        ("Часы работы",   16),
    ]
    for ci, (title, width) in enumerate(COLS, 1):
        hcell(ws2, 1, ci, title, width=width)
    ws2.freeze_panes = "A2"

    sorted_rows = sorted(rows, key=lambda r: r["district"])
    row_num = 2
    sheet_total = 0

    for district, group in groupby(sorted_rows, key=lambda r: r["district"]):
        group_rows = list(group)
        cnt = len(group_rows)

        for r in group_rows:
            fill = alt_fill if row_num % 2 == 0 else None
            dcell(ws2, row_num, 1, r["region"],       fill=fill)
            dcell(ws2, row_num, 2, r["district"],     fill=fill)
            dcell(ws2, row_num, 3, r["name"],         fill=fill)
            dcell(ws2, row_num, 4, r["full_address"], fill=fill, wrap=True)
            dcell(ws2, row_num, 5, r["hours"],        fill=fill, align="center")
            row_num += 1

        # Итог по району
        for ci in range(1, 6):
            c = ws2.cell(row=row_num, column=ci)
            c.fill = tot_fill; c.border = border; c.font = tot_font
            c.alignment = Alignment(vertical="center")
        ws2.cell(row=row_num, column=1).value = REGION
        ws2.cell(row=row_num, column=2).value = f"Итого по «{district}»"
        ws2.cell(row=row_num, column=3).value = f"{cnt} магазинов"
        ws2.cell(row=row_num, column=3).alignment = Alignment(horizontal="center", vertical="center")
        row_num += 1
        sheet_total += cnt

    # Общий итог
    for ci in range(1, 6):
        c = ws2.cell(row=row_num, column=ci)
        c.fill = grand_fill; c.border = border; c.font = grand_font
        c.alignment = Alignment(vertical="center")
    ws2.cell(row=row_num, column=1).value = REGION
    ws2.cell(row=row_num, column=2).value = "ИТОГО ПО ВСЕМ РАЙОНАМ"
    ws2.cell(row=row_num, column=3).value = f"{sheet_total} магазинов"
    ws2.cell(row=row_num, column=3).alignment = Alignment(horizontal="center", vertical="center")
    ws2.auto_filter.ref = f"A1:E{row_num - 1}"

    wb.save(path)


# ═══════════════════════════════════════════════════════════════════════════════
# 3. MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    rows = fetch_and_parse()

    if not rows:
        print("ОШИБКА: магазины не найдены.")
        sys.exit(1)

    district_counts = Counter(r["district"] for r in rows)

    print(f"\n{'─'*50}")
    for district, count in sorted(district_counts.items(), key=lambda x: -x[1]):
        print(f"  ✓ {district}: {count} магазинов")
    print(f"{'─'*50}")
    print(f"  {REGION}: {len(rows)} магазинов ({len(district_counts)} районов)")
    print(f"{'─'*50}\n")

    save_excel(rows, OUTPUT)
    print(f"✓ Готово → {OUTPUT}")


if __name__ == "__main__":
    main()