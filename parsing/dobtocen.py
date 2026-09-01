"""
Парсер магазинов «Доброцен» — magazinnoff.ru
Antibot Cloud обходится через playwright-stealth.

Установка:
    pip install playwright openpyxl pandas beautifulsoup4 playwright-stealth
    playwright install chromium

Подготовка:
    1. cities.html  — сохранённый HTML страницы /magazin/dobrocen/cities
    2. city_and_regions_Russia.xlsx — справочник рядом со скриптом

Запуск:
    python dobrocen_parser.py
"""

import re, sys, asyncio
from collections import Counter, defaultdict
from itertools import groupby

from bs4 import BeautifulSoup
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from playwright.async_api import async_playwright

BASE_URL    = "https://www.magazinnoff.ru"
CITIES_HTML = "cities.html"
DICT_FILE   = "city_and_regions_Russia.xlsx"
OUTPUT      = "dobrocen_shops.xlsx"


# ═══════════════════════════════════════════════════════════════════════════════
# 1. СПРАВОЧНИК
# ═══════════════════════════════════════════════════════════════════════════════

def load_city_dict(path):
    try:
        df = pd.read_excel(path, usecols=["city", "region_name"]).dropna()
        d = dict(zip(df["city"].str.strip(), df["region_name"].str.strip()))
        print(f"  Справочник: {len(d)} городов из {path}")
        return d
    except FileNotFoundError:
        print(f"  ⚠  {path} не найден")
        return {}

def resolve_region(city_name, city_dict):
    r = city_dict.get(city_name.strip())
    if r: return r
    r = city_dict.get(city_name.strip().replace("ё", "е"))
    if r: return r
    return f"[не найден] {city_name}"


# ═══════════════════════════════════════════════════════════════════════════════
# 2. СПИСОК ГОРОДОВ
# ═══════════════════════════════════════════════════════════════════════════════

def load_cities(html_path):
    with open(html_path, encoding="utf-8") as f:
        html = f.read()
    soup = BeautifulSoup(html, "html.parser")
    seen, cities = set(), []
    for a in soup.find_all("a", href=re.compile(r"^/magazin/dobrocen/c/[^/]+$")):
        href = a["href"]
        slug = href.rstrip("/").split("/")[-1]
        name = a.get_text(strip=True)
        if slug and slug not in seen and name:
            seen.add(slug)
            cities.append({"name": name, "slug": slug, "url": BASE_URL + href})
    return cities


# ═══════════════════════════════════════════════════════════════════════════════
# 3. ПАРСИНГ HTML
# ═══════════════════════════════════════════════════════════════════════════════

def parse_html(html):
    soup = BeautifulSoup(html, "html.parser")
    results = []
    for small in soup.find_all("small", itemprop="streetAddress"):
        address = small.get_text(strip=True)
        if not address:
            continue
        parent = small.find_parent("a") or small.find_parent("li") or small.parent
        phone, hours = "", ""
        if parent:
            tel = parent.find(itemprop="telephone")
            oh  = parent.find(itemprop="openingHours")
            if tel:
                phone = re.sub(r"^Телефон:\s*", "", tel.get_text(strip=True))
            if oh:
                hours = re.sub(r"^Часы работы:\s*", "", oh.get_text(strip=True))
        results.append({"address": address, "phone": phone, "hours": hours})
    return results


# ═══════════════════════════════════════════════════════════════════════════════
# 4. PLAYWRIGHT + STEALTH
# ═══════════════════════════════════════════════════════════════════════════════

async def scrape_cities(cities, city_dict):
    rows = []
    total = len(cities)

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(
            headless=False,
        )
        ctx = await browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            locale="ru-RU",
            viewport={"width": 1280, "height": 800},
        )
        page = await ctx.new_page()


        # Открываем первый город — пусть пользователь решит капчу вручную
        print("\n  Открываем браузер...")
        await page.goto(
            "https://www.magazinnoff.ru/magazin/dobrocen/c/abdylino",
            wait_until="domcontentloaded", timeout=30000
        )
        print("\n" + "="*55)
        print("  РЕШИТЕ КАПЧУ В БРАУЗЕРЕ, затем нажмите Enter здесь")
        print("="*55)
        await asyncio.get_event_loop().run_in_executor(None, input)
        print("  Продолжаем...\n")

        for i, city in enumerate(cities, 1):
            print(f"  [{i:3}/{total}] {city['name']:<22}", end=" ", flush=True)
            region = resolve_region(city["name"], city_dict)
            shops = []

            try:
                await page.goto(city["url"], wait_until="domcontentloaded",
                                timeout=30000)

                # Ждём адресный блок
                try:
                    await page.wait_for_selector(
                        "small[itemprop='streetAddress']", timeout=8000
                    )
                except Exception:
                    pass

                html = await page.content()
                shops = parse_html(html)

            except Exception as e:
                print(f"ERR: {e}")

            if shops:
                for s in shops:
                    rows.append({"region": region, "city": city["name"],
                                 "address": s["address"], "phone": s["phone"],
                                 "hours": s["hours"]})
                print(f"✓ {len(shops)} ТТ  | {region}")
            else:
                rows.append({"region": region, "city": city["name"],
                             "address": "", "phone": "", "hours": ""})
                print(f"– нет ТТ     | {region}")

        await browser.close()

    return rows


# ═══════════════════════════════════════════════════════════════════════════════
# 5. EXCEL
# ═══════════════════════════════════════════════════════════════════════════════

def save_excel(rows, path):
    wb = openpyxl.Workbook()
    thin       = Side(style="thin", color="CCCCCC")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill   = PatternFill("solid", fgColor="1F4E79")
    alt_fill   = PatternFill("solid", fgColor="EBF3FB")
    tot_fill   = PatternFill("solid", fgColor="D6E4F0")
    grand_fill = PatternFill("solid", fgColor="BDD7EE")
    hdr_font   = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    dat_font   = Font(name="Arial", size=10)
    tot_font   = Font(bold=True, italic=True, name="Arial", size=10)
    grand_font = Font(bold=True, name="Arial", size=11)

    def hcell(ws, row, col, val, width=None):
        c = ws.cell(row=row, column=col, value=val)
        c.font = hdr_font; c.fill = hdr_fill; c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center")
        if width: ws.column_dimensions[c.column_letter].width = width

    def dcell(ws, row, col, val, align="left", fill=None, font=None, wrap=False):
        c = ws.cell(row=row, column=col, value=val)
        c.font = font or dat_font; c.border = border
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        if fill: c.fill = fill

    ws1 = wb.active
    ws1.title = "Итог по регионам"
    ws1.row_dimensions[1].height = 22
    hcell(ws1, 1, 1, "Область / Регион", width=40)
    hcell(ws1, 1, 2, "Городов",          width=12)
    hcell(ws1, 1, 3, "Кол-во ТТ",        width=12)
    ws1.freeze_panes = "A2"

    reg = defaultdict(lambda: {"cities": set(), "shops": 0})
    for r in rows:
        reg[r["region"]]["cities"].add(r["city"])
        if r["address"]: reg[r["region"]]["shops"] += 1

    gc = gs = 0
    for ri, (region, d) in enumerate(
            sorted(reg.items(), key=lambda x: -x[1]["shops"]), 2):
        fill = alt_fill if ri % 2 == 0 else None
        dcell(ws1, ri, 1, region,           fill=fill)
        dcell(ws1, ri, 2, len(d["cities"]), fill=fill, align="right")
        dcell(ws1, ri, 3, d["shops"],       fill=fill, align="right")
        gc += len(d["cities"]); gs += d["shops"]

    ri_g = len(reg) + 2
    dcell(ws1, ri_g, 1, "ИТОГО", font=grand_font, fill=grand_fill)
    dcell(ws1, ri_g, 2, gc,      font=grand_font, fill=grand_fill, align="right")
    dcell(ws1, ri_g, 3, gs,      font=grand_font, fill=grand_fill, align="right")
    ws1.auto_filter.ref = f"A1:C{ri_g - 1}"

    ws2 = wb.create_sheet("Все адреса")
    ws2.row_dimensions[1].height = 22
    for ci, (t, w) in enumerate([("Регион",38),("Город",20),("Адрес",50),
                                   ("Телефон",20),("Часы работы",25)], 1):
        hcell(ws2, 1, ci, t, width=w)
    ws2.freeze_panes = "A2"

    sorted_rows = sorted(rows, key=lambda r: (r["region"], r["city"]))
    row_num = 2; sheet_total = 0

    for region, grp in groupby(sorted_rows, key=lambda r: r["region"]):
        grp = list(grp)
        reg_shops = sum(1 for r in grp if r["address"])
        for r in grp:
            fill = alt_fill if row_num % 2 == 0 else None
            dcell(ws2, row_num, 1, r["region"],  fill=fill)
            dcell(ws2, row_num, 2, r["city"],    fill=fill)
            dcell(ws2, row_num, 3, r["address"], fill=fill, wrap=True)
            dcell(ws2, row_num, 4, r["phone"],   fill=fill, align="center")
            dcell(ws2, row_num, 5, r["hours"],   fill=fill, align="center")
            row_num += 1

        for ci in range(1, 6):
            c = ws2.cell(row=row_num, column=ci)
            c.fill = tot_fill; c.border = border; c.font = tot_font
            c.alignment = Alignment(vertical="center")
        ws2.cell(row=row_num, column=1).value = f"Итого по «{region}»"
        ws2.cell(row=row_num, column=2).value = f"{len({r['city'] for r in grp})} городов"
        ws2.cell(row=row_num, column=3).value = f"{reg_shops} ТТ"
        for ci in (2, 3):
            ws2.cell(row=row_num, column=ci).alignment = Alignment(
                horizontal="center", vertical="center")
        row_num += 1; sheet_total += reg_shops

    for ci in range(1, 6):
        c = ws2.cell(row=row_num, column=ci)
        c.fill = grand_fill; c.border = border; c.font = grand_font
        c.alignment = Alignment(vertical="center")
    ws2.cell(row=row_num, column=1).value = "ИТОГО ПО ВСЕМ РЕГИОНАМ"
    ws2.cell(row=row_num, column=2).value = f"{len({r['city'] for r in rows})} городов"
    ws2.cell(row=row_num, column=3).value = f"{sheet_total} ТТ"
    for ci in (2, 3):
        ws2.cell(row=row_num, column=ci).alignment = Alignment(
            horizontal="center", vertical="center")
    ws2.auto_filter.ref = f"A1:E{row_num - 1}"

    wb.save(path)


# ═══════════════════════════════════════════════════════════════════════════════
# 6. MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    print("=" * 55)
    print("  Парсер Доброцен — Playwright + Stealth")
    print("=" * 55 + "\n")

    city_dict = load_city_dict(DICT_FILE)

    print(f"\nЧитаем список городов из: {CITIES_HTML}")
    cities = load_cities(CITIES_HTML)
    if not cities:
        print("ОШИБКА: список городов пуст.")
        sys.exit(1)
    print(f"  → Найдено: {len(cities)} городов\n")

    rows = asyncio.run(scrape_cities(cities, city_dict))

    region_counts = Counter(r["region"] for r in rows if r["address"])
    print(f"\n{'─'*55}")
    for region, count in sorted(region_counts.items(), key=lambda x: -x[1]):
        print(f"  ✓ {region}: {count} ТТ")
    print(f"{'─'*55}")
    print(f"  ИТОГО: {len(region_counts)} регионов, "
          f"{sum(region_counts.values())} ТТ")
    print(f"{'─'*55}\n")

    save_excel(rows, OUTPUT)
    print(f"✓ Готово → {OUTPUT}")

if __name__ == "__main__":
    main()