# services/production.py
import os

# Тяжёлые библиотеки — ленивый импорт

MONTHS_RU = {
    1: "Январь",
    2: "Февраль",
    3: "Март",
    4: "Апрель",
    5: "Май",
    6: "Июнь",
    7: "Июль",
    8: "Август",
    9: "Сентябрь",
    10: "Октябрь",
    11: "Ноябрь",
    12: "Декабрь",
}

MONTH_LABELS = [
    "1 - Январь",
    "2 - Февраль",
    "3 - Март",
    "4 - Апрель",
    "5 - Май",
    "6 - Июнь",
    "7 - Июль",
    "8 - Август",
    "9 - Сентябрь",
    "10 - Октябрь",
    "11 - Ноябрь",
    "12 - Декабрь",
]

# ─────────────────────────────────────────────────────────────
# 1. СВОД
# ─────────────────────────────────────────────────────────────


def extract_month_data(filepath, sheet_index=0):
    import pandas as pd  # ленивый импорт

    wb_sheets = pd.read_excel(filepath, header=None, sheet_name=None)
    sheet_names = list(wb_sheets.keys())
    sheet_name = sheet_names[sheet_index]
    df = wb_sheets[sheet_name]

    companies = df.iloc[1, 3:].tolist()
    groups = df.iloc[2, 3:].tolist()
    results = []

    ROW_MASLO_TOTAL = 112
    ROW_MASLO_SOY = 114
    ROW_MASLO_RAPE = 115
    ROW_MARG = 119
    ROW_KETCH = 124

    for col_offset, (company, group) in enumerate(zip(companies, groups)):
        col = 3 + col_offset
        if pd.isna(company):
            continue
        company = str(company).strip()
        group = str(group).strip() if not pd.isna(group) else ""

        def val(row, _col=col):
            v = df.iloc[row, _col]
            return 0.0 if pd.isna(v) else float(v)

        maslo_total = val(ROW_MASLO_TOTAL)
        maslo_soy = val(ROW_MASLO_SOY)
        maslo_rape = val(ROW_MASLO_RAPE)
        maslo_other = maslo_total - maslo_soy - maslo_rape

        if maslo_total != 0 or maslo_soy != 0 or maslo_rape != 0:
            if maslo_other != 0:
                results.append(
                    [
                        group,
                        company,
                        "Производство фасованного масла",
                        "Масло",
                        maslo_other,
                    ]
                )
            if maslo_soy != 0:
                results.append(
                    [
                        group,
                        company,
                        "Производство фасованного масла",
                        "Соевое",
                        maslo_soy,
                    ]
                )
            if maslo_rape != 0:
                results.append(
                    [
                        group,
                        company,
                        "Производство фасованного масла",
                        "Рапсовое",
                        maslo_rape,
                    ]
                )

        marg_total = val(ROW_MARG)
        marg_mono = val(120)
        marg_fas = val(121)
        marg_spread = val(122)

        if marg_total != 0:
            results.append(
                [
                    group,
                    company,
                    "Производство маргариновой продукции",
                    "Всего",
                    marg_total,
                ]
            )
        if marg_mono != 0:
            results.append(
                [
                    group,
                    company,
                    "Производство маргариновой продукции",
                    "Монолит",
                    marg_mono,
                ]
            )
        if marg_fas != 0:
            results.append(
                [
                    group,
                    company,
                    "Производство маргариновой продукции",
                    "Фасовка без спредов",
                    marg_fas,
                ]
            )
        if marg_spread != 0:
            results.append(
                [
                    group,
                    company,
                    "Производство маргариновой продукции",
                    "Спреды/крем",
                    marg_spread,
                ]
            )

        ketch = val(ROW_KETCH)
        if ketch != 0:
            results.append([group, company, "Производство кетчупов", "Кетчуп", ketch])

        mayo_total = val(131)
        sauce = val(134)
        mayo = mayo_total - sauce
        if mayo != 0:
            results.append(
                [
                    group,
                    company,
                    "Производство майонезов и майонезных соусов",
                    "Майонез",
                    mayo,
                ]
            )
        if sauce != 0:
            results.append(
                [
                    group,
                    company,
                    "Производство майонезов и майонезных соусов",
                    "Соус",
                    sauce,
                ]
            )

    return results, sheet_name


# ─────────────────────────────────────────────────────────────
# 2. НПК
# ─────────────────────────────────────────────────────────────


def get_npk_value(npk_filepath, month_num, year):
    import pandas as pd

    df = pd.read_excel(npk_filepath, header=None, sheet_name="свод")
    month_name = MONTHS_RU[int(month_num)]

    year_row = df.iloc[8, :]
    month_row = df.iloc[9, :]
    val_row = df.iloc[12, :]

    year_start_col = None
    for ci, v in enumerate(year_row):
        try:
            if int(float(str(v))) == int(year):
                year_start_col = ci
                break
        except (ValueError, TypeError):
            continue

    if year_start_col is None:
        raise ValueError(f"Год {year} не найден в файле НПК")

    for ci in range(year_start_col, len(month_row)):
        if str(month_row[ci]).strip() == month_name:
            return float(val_row[ci]) / 1000

    raise ValueError(f"Месяц {month_name} {year} не найден в файле НПК")


# ─────────────────────────────────────────────────────────────
# 3. Тольяти
# ─────────────────────────────────────────────────────────────


def get_tolyatti_value(folder):
    import pandas as pd

    xlsx_files = [
        os.path.join(folder, f)
        for f in os.listdir(folder)
        if f.endswith(".xlsx") and not f.startswith("~")
    ]
    if not xlsx_files:
        raise ValueError(f"В папке Тольяти нет xlsx файлов: {folder}")

    latest = max(xlsx_files, key=os.path.getmtime)
    df = pd.read_excel(latest, header=None, sheet_name="Для ежедневного отчёта")
    value = df.iloc[9, 2]
    return (float(value) if not pd.isna(value) else 0.0), os.path.basename(latest)


# ─────────────────────────────────────────────────────────────
# 4. Маппинг и запись в целевой файл
# ─────────────────────────────────────────────────────────────


def load_mapping(mapping_path):
    import pandas as pd

    df = pd.read_excel(mapping_path, sheet_name="Маппинг")
    df.columns = ["завод", "подкат", "лист", "строка", "родитель", "кол"]
    result = []
    for _, row in df.iterrows():
        if (
            pd.notna(row["завод"])
            and pd.notna(row["строка"])
            and str(row["строка"]).strip()
        ):
            result.append(
                {
                    "завод": str(row["завод"]).strip(),
                    "подкат": (
                        str(row["подкат"]).strip() if pd.notna(row["подкат"]) else ""
                    ),
                    "лист": str(row["лист"]).strip(),
                    "строка": str(row["строка"]).strip(),
                    "родитель": (
                        str(row["родитель"]).strip()
                        if pd.notna(row["родитель"])
                        else ""
                    ),
                    "кол": (
                        str(row["кол"]).strip().upper() if pd.notna(row["кол"]) else "A"
                    ),
                }
            )
    return result


def write_to_target(target_path, results, mapping, month_num, year, log):
    from openpyxl import load_workbook
    import datetime as dt

    wb = load_workbook(target_path)
    month_num = int(month_num)
    year = int(year)

    # Строим lookup с проверкой дублей
    val_lookup = {}
    for row in results:
        key = (str(row[1]).strip(), str(row[3]).strip())
        if key in val_lookup:
            log(
                f"⚠ Дубль ключа: завод='{key[0]}' подкат='{key[1]}' "
                f"старое={val_lookup[key]:.3f} → новое={row[4]:.3f}"
            )
        val_lookup[key] = row[4]

    log(
        f"Уникальных ключей в данных: {len(val_lookup)}, строк маппинга: {len(mapping)}"
    )

    written = []
    not_found = []
    # year/month фиксированы на весь вызов — колонка для листа одна и та
    # же для всех строк маппинга этого листа, нет смысла искать её заново
    # на каждой строке.
    target_col_cache = {}

    for m in mapping:
        key = (m["завод"], m["подкат"])

        if key not in val_lookup:
            log(f"  — Нет данных: завод={repr(key[0])} подкат={repr(key[1])}")
            not_found.append(f"Нет данных: завод='{key[0]}' подкат='{key[1]}'")
            continue

        value = val_lookup[key]
        sheet_name = m["лист"]
        row_label = m["строка"]
        parent_label = m["родитель"]
        search_col = 1 if m["кол"] == "A" else 2

        if sheet_name not in wb.sheetnames:
            available = ", ".join(f"'{s}'" for s in wb.sheetnames)
            not_found.append(f"Лист '{sheet_name}' не найден. Доступные: {available}")
            continue

        ws = wb[sheet_name]

        if sheet_name in target_col_cache:
            target_col = target_col_cache[sheet_name]
        else:
            target_col = None
            for col in ws.iter_cols(min_col=2, max_row=2):
                for cell in col:
                    v = cell.value
                    if (
                        isinstance(v, dt.datetime)
                        and v.year == year
                        and v.month == month_num
                    ):
                        target_col = cell.column
                        break
                if target_col:
                    break
            target_col_cache[sheet_name] = target_col

        if target_col is None:
            not_found.append(f"{sheet_name}: колонка {year}-{month_num:02d} не найдена")
            continue

        search_start = 1
        if parent_label:
            parent_row = None
            for r in range(1, ws.max_row + 1):
                for c in [1, 2]:
                    v = ws.cell(row=r, column=c).value
                    if v and str(v).strip() == parent_label:
                        parent_row = r
                        break
                if parent_row:
                    break
            if parent_row is None:
                not_found.append(f"{sheet_name}: родитель '{parent_label}' не найден")
                continue
            search_start = parent_row

        target_row = None
        for r in range(search_start, ws.max_row + 1):
            v = ws.cell(row=r, column=search_col).value
            if v and str(v).strip() == row_label:
                target_row = r
                break

        if target_row is None:
            not_found.append(
                f"{sheet_name}: строка '{row_label}' не найдена (родитель='{parent_label}')"
            )
            continue

        ws.cell(row=target_row, column=target_col).value = value
        written.append(f"{sheet_name} | {row_label} | {value:.3f}")
        log(f"  ✓ {sheet_name} → {row_label} = {value:.3f}")

    try:
        wb.save(target_path)
    except PermissionError:
        raise PermissionError(
            f"Закройте файл в Excel и попробуйте снова:\n{target_path}"
        )

    return written, not_found


# ─────────────────────────────────────────────────────────────
# 5. Основная функция для вызова из GUI
# ─────────────────────────────────────────────────────────────


def run_production(
    svod_file,  # путь к файлу СВОД (раньше была папка)
    npk_file,
    tolyatti_folder,
    target_file,
    mapping_file,
    month_str,
    year,
    log,
    messagebox,
    stop_event,
):
    if not svod_file or not os.path.isfile(svod_file):
        messagebox.showerror("Ошибка", "Укажите файл СВОД (.xlsx)")
        return
    if not target_file or not os.path.isfile(target_file):
        messagebox.showerror("Ошибка", "Укажите файл тестовые_данные.xlsx")
        return
    if not mapping_file or not os.path.isfile(mapping_file):
        messagebox.showerror("Ошибка", "Укажите файл маппинга")
        return

    log(f"Читаем СВОД: {os.path.basename(svod_file)}")
    sheet_index = max(0, int(month_str) - 1)
    results, sheet_name = extract_month_data(svod_file, sheet_index=sheet_index)
    log(f"Извлечено {len(results)} строк из СВОД (лист: {sheet_name})")

    if stop_event.is_set():
        return

    if npk_file and os.path.isfile(npk_file):
        try:
            npk_val = get_npk_value(npk_file, month_str, year)
            results.append(["НПК", "НПК (из первичн продаж)", "НПК", "НПК", npk_val])
            log(f"НПК: {npk_val:.3f} тн")
        except Exception as e:
            log(f"⚠ НПК пропущен: {e}")

    if stop_event.is_set():
        return

    if tolyatti_folder and os.path.isdir(tolyatti_folder):
        try:
            tol_val, tol_fname = get_tolyatti_value(tolyatti_folder)
            results.append(["Тольяти", "Тольяти", "Тольяти", "Тольяти", tol_val])
            log(f"Тольяти ({tol_fname}): {tol_val:.3f}")
        except Exception as e:
            log(f"⚠ Тольяти пропущен: {e}")

    if stop_event.is_set():
        return

    log("Загружаем маппинг и пишем в целевой файл...")
    try:
        mapping = load_mapping(mapping_file)
        written, not_found = write_to_target(
            target_file, results, mapping, month_str, year, log
        )
    except PermissionError as e:
        messagebox.showerror("Файл занят", str(e))
        return
    except Exception as e:
        messagebox.showerror("Ошибка записи", str(e))
        return

    month_name = MONTHS_RU[int(month_str)]
    msg = f"Записано: {len(written)} ячеек\nФайл: {os.path.basename(target_file)}\nПериод: {month_name} {year}"
    if not_found:
        msg += f"\n\nНе найдено ({len(not_found)}):\n" + "\n".join(not_found[:15])
        log(f"⚠ Не найдено {len(not_found)} позиций:")
        for nf in not_found:
            log(f"  — {nf}")

    log(f"✅ Готово! Записано {len(written)} ячеек — {month_name} {year}")
    messagebox.showinfo("Готово", msg)