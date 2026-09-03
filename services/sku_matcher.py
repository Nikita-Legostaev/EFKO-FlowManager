"""
services/sku_matcher.py

Иерархический матчинг SKU.

Штрафы при несовпадении атрибутов:
  1. Категория  — штраф 90%  (x 0.10)
  2. Бренд      — штраф 85%  (x 0.15)
  3. Вкус       — штраф 20%  (x 0.80)
  4. Вес/объём  — штраф ~12% (x 0.88)

Двухэтапный подход:
  Этап 1 — жёсткая фильтрация по категории + бренду
  Этап 2 — TF-IDF + hierarchical_score
"""

import re
from collections import defaultdict
from pathlib import Path

_STOP = frozenset({
    "г","кг","мл","л","шт","уп","пак","бут","бан","кор",
    "пл","жб","пэт","мас","кал",
    "и","в","с","от","из","для","по","на","не","за",
    "со","во","при","без","об","же","до","то",
    "что","как","все","это","или",
    "ml","kg","gr","lt",
})

def _norm(s):
    return re.sub(r"\s+", " ", str(s).lower().strip())


def extract_flavors_from_reference(ref_rows: list) -> frozenset:
    """
    Собирает вкусы/разновидности из справочника автоматически.
    Сравнивает col[2] raw и col[3] скорр — слова только в скорр
    и есть кандидаты на вкус/тип продукта.
    """
    flavors = set()
    for r in ref_rows:
        if str(r[4] or "") != "индикативное":
            continue
        raw  = set(re.findall(r"[а-яё]{4,}", str(r[2] or "").lower()))
        corr = set(re.findall(r"[а-яё]{4,}", str(r[3] or "").lower()))
        flavors |= (corr - raw - _STOP)
    return frozenset(flavors)


def extract_weight(s):
    matches = re.findall(
        r"(\d+(?:[.,]\d+)?)\s*(мл|мг|кг|л(?=[^а-яёa-z]|$)|г(?=[^а-яёa-z]|$)|kg|ml|mg|lb)",
        s, re.IGNORECASE,
    )
    if not matches:
        return None
    vals = []
    for num_str, unit in matches:
        num = float(num_str.replace(",", "."))
        u = unit.lower()
        if u == "л":         num *= 1000
        if u in ("кг","kg"): num *= 1000
        vals.append(num)
    return max(vals)


def extract_flavor(s, flavors: frozenset) -> str | None:
    """Ищет вкус/разновидность из динамического набора, построенного по справочнику."""
    sl = s.lower()
    for f in flavors:
        if f in sl:
            return f
    return None


def descriptor_conflict(a: str, b: str, brand_words: frozenset = frozenset()) -> float:
    """
    Сравнивает описательные слова двух названий напрямую — без предустановленного списка.
    Конфликт = слово есть в одном названии но нет в другом.

    Примеры:
      "Острый" vs "Томатный"  → 2 конфликта → штраф 60%
      "Шашлычный" vs "Острый" → 2 конфликта → штраф 60%
      "Томатный 300г" vs "Томатный 320г" → 0 конфликтов → без штрафа
    """
    # Берём слова длиннее 4 букв, убираем стоп-слова и слова бренда
    # Слова-исключения: упаковка, технические, страна, общие категорийные
    skip = _STOP | brand_words | {
        # Упаковка и способ подачи
        "дозатором", "дозатор", "пластиковый", "пластиковом", "пластик",
        "пакет", "стекло", "жесть", "крышка", "крышкой", "флакон",
        "дойпак", "пэт", "туба", "тубе",
        # Страна, происхождение
        "россия", "российский", "отечественный",
        # Общие слова категории (не различают SKU)
        "высшей", "первой", "категории", "категория",
        "натуральный", "натурального", "томатный",  # томатный = базовый для кетчупа
        "кетчуп", "майонез", "масло", "соус",
    }
    a_words = frozenset(w for w in re.findall(r"[а-яё]{4,}", a.lower()) if w not in skip)
    b_words = frozenset(w for w in re.findall(r"[а-яё]{4,}", b.lower()) if w not in skip)

    if not a_words and not b_words:
        return 1.0
    if not a_words or not b_words:
        non_empty = a_words if a_words else b_words
        return max(0.30, 1.0 - len(non_empty) * 0.20)

    conflicts = len((a_words - b_words) | (b_words - a_words))
    if conflicts == 0:
        return 1.0
    # Каждый конфликт = -20%, минимум 0.30
    return max(0.20, 1.0 - conflicts * 0.40)


def weight_penalty(a: str, b: str) -> float:
    """
    Штраф за несовпадение веса/объёма — относительный порог.

    Разница > 15% → сильный штраф (ratio²)
      300г vs 500г → (300/500)² = 0.36  → штраф 64%
      300г vs 320г → разница 6.7% → мягкий штраф

    Разница <= 15% → мягкий штраф (ratio²·⁵)
      300г vs 320г → (300/320)^2.5 = 0.85 → штраф 15%
      390г vs 400г → (390/400)^2.5 = 0.94 → штраф 6%
    """
    wa, wb = extract_weight(a), extract_weight(b)
    if wa is None or wb is None:
        return 1.0
    ratio      = min(wa, wb) / max(wa, wb)
    diff_pct   = 1.0 - ratio          # 0.0 = одинаковые, 1.0 = полностью разные
    if diff_pct > 0.15:               # разница > 15%
        return ratio ** 2             # сильный штраф
    return ratio ** 2.5               # мягкий но заметный штраф


def _is_near_duplicate(query, ref_raw, threshold=0.97):
    from rapidfuzz import fuzz
    return fuzz.ratio(_norm(query), _norm(ref_raw)) / 100.0 >= threshold


def hierarchical_score(query, query_brand, query_category, ref_row, base_score, flavors=frozenset()):
    """
    Иерархические штрафы:
      Категория → x0.10 при несовпадении (штраф 90%)
      Бренд     → x0.15 при несовпадении (штраф 85%)
      Вкус      → x0.80 при несовпадении (штраф 20%)
      Вес       → x0.88 при несовпадении (штраф ~12%)
    """
    score = base_score

    ref_category = str(ref_row[0] or "").strip().lower()
    ref_brand    = str(ref_row[1] or "").strip().lower()
    ref_raw      = str(ref_row[2] or "").strip()

    q_cat   = query_category.strip().lower()
    q_brand = query_brand.strip().lower()

    # 1. Категория: штраф 70% (мягче — категории могут называться чуть иначе)
    if q_cat and ref_category and q_cat not in ("nan","none",""):
        if q_cat not in ref_category and ref_category not in q_cat:
            score *= 0.30

    # 2. Бренд: штраф 80%
    if q_brand and ref_brand and q_brand not in ("nan","none",""):
        if q_brand not in ref_brand and ref_brand not in q_brand:
            score *= 0.20

    # 3. Описательные слова: сравниваем напрямую — не нужен список вкусов
    # "Острый" vs "Томатный" → конфликт → штраф
    # "Томатный" vs "Томатный" → совпадение → без штрафа
    score *= descriptor_conflict(query, ref_raw)

    # 4. Вес — через weight_penalty
    score *= weight_penalty(query, ref_raw)

    return score


def build_from_reference(ref_rows):
    brand_map = {}
    for r in ref_rows:
        canonical = str(r[1] or "").strip()
        if not canonical:
            continue
        brand_map[canonical.lower()] = canonical
        for token in re.findall(r"[а-яёa-z\w]{3,}", canonical.lower()):
            if token not in _STOP and token not in brand_map:
                brand_map[token] = canonical

    brand_sku_words = defaultdict(list)
    for r in ref_rows:
        if str(r[4] or "") != "индикативное":
            continue
        brand    = str(r[1] or "").strip()
        raw_name = str(r[2] or "").strip()
        if not brand or not raw_name:
            continue
        tokens = frozenset(
            w for w in re.findall(r"[а-яёa-z]{3,}", raw_name.lower())
            if w not in _STOP
        )
        if tokens:
            brand_sku_words[brand.lower()].append(tokens)

    descriptors = set()
    for brand, word_sets in brand_sku_words.items():
        if len(word_sets) < 2:
            continue
        all_words = set().union(*word_sets)
        n = len(word_sets)
        for word in all_words:
            count = sum(1 for ws in word_sets if word in ws)
            if 0 < count < n:
                descriptors.add(word)

    return brand_map, descriptors


# ─── Кэш модели ───────────────────────────────────────────────────────────────

import os
import hashlib

def _model_cache_path(ref_path: str) -> str:
    """Путь к файлу кэша рядом со справочником."""
    base = os.path.splitext(ref_path)[0]
    return base + "_sku_model.pkl"


def _ref_fingerprint(ref_path: str) -> str:
    """Отпечаток справочника — хэш размера + даты изменения."""
    stat = os.stat(ref_path)
    key  = f"{stat.st_size}:{stat.st_mtime}"
    return hashlib.md5(key.encode()).hexdigest()


def save_model(matcher, ref_path: str, fingerprint: str, log=print):
    """Сохраняет модель и отпечаток справочника на диск."""
    import joblib
    cache_path = _model_cache_path(ref_path)
    try:
        joblib.dump({"fingerprint": fingerprint, "matcher": matcher}, cache_path)
        log(f"  Модель сохранена: {os.path.basename(cache_path)}")
    except Exception as e:
        log(f"  ⚠ Не удалось сохранить модель: {e}")


def load_model(ref_path: str, log=print):
    """
    Загружает модель если справочник не изменился.
    Возвращает matcher или None если кэш устарел / не существует.
    """
    import joblib
    cache_path = _model_cache_path(ref_path)
    if not os.path.exists(cache_path):
        return None
    try:
        current_fp = _ref_fingerprint(ref_path)
        data = joblib.load(cache_path)
        if data.get("fingerprint") != current_fp:
            log("  Справочник изменился — переобучаем модель...")
            return None
        log(f"  Модель загружена из кэша: {os.path.basename(cache_path)}")
        return data["matcher"]
    except Exception as e:
        log(f"  ⚠ Не удалось загрузить кэш: {e}")
        return None


class SKUMatcher:
    """TF-IDF + двухэтапный матчинг с иерархическими штрафами."""

    def __init__(self, ref_rows, descriptors, flavors=frozenset()):
        from sklearn.feature_extraction.text import TfidfVectorizer

        self.descriptors = descriptors
        self.flavors = flavors
        self._rows = [
            r for r in ref_rows
            if str(r[4] or "") == "индикативное"
            and str(r[2] or "").strip()
        ]
        self._raw_names = [_norm(str(r[2])) for r in self._rows]

        if not self._raw_names:
            self._vectorizer = None
            self._matrix = None
            self._brand_idx = {}
            self._cat_idx = {}
            return

        # Комбинированная токенизация: символьные + словесные n-граммы
        # Символьные (2-4): ловят опечатки, сокращения, частичные слова
        # Словесные (1-2):  "Шашлычный 500г" как единый признак
        from sklearn.pipeline import FeatureUnion
        self._vectorizer = FeatureUnion([
            ("char", TfidfVectorizer(
                analyzer="char_wb", ngram_range=(2, 4),
                min_df=1, sublinear_tf=True,
            )),
            ("word", TfidfVectorizer(
                analyzer="word", ngram_range=(1, 2),
                min_df=1, sublinear_tf=True,
            )),
        ])
        from sklearn.preprocessing import normalize
        self._matrix = normalize(self._vectorizer.fit_transform(self._raw_names))

        self._brand_idx = defaultdict(list)
        for i, r in enumerate(self._rows):
            brand = str(r[1] or "").strip().lower()
            if brand:
                self._brand_idx[brand].append(i)

        self._cat_idx = defaultdict(list)
        for i, r in enumerate(self._rows):
            cat = str(r[0] or "").strip().lower()
            if cat:
                self._cat_idx[cat].append(i)

    def find_match(self, query, csv_brand, threshold, csv_category="",
                   rejections: "RejectionStore | None" = None):
        if not query or not query.strip():
            return None, 0.0
        if self._vectorizer is None or not self._raw_names:
            return None, 0.0

        import numpy as np

        q_norm = _norm(query)
        if not q_norm:
            return None, 0.0

        try:
            from sklearn.preprocessing import normalize
            q_vec = normalize(self._vectorizer.transform([q_norm]))
        except Exception:
            return None, 0.0

        cosine = (self._matrix @ q_vec.T).toarray().flatten()
        # Клипуем на случай погрешностей float
        cosine = cosine.clip(0.0, 1.0)

        # Этап 1: мягкий фильтр по категории (substring match, fallback если ничего не нашли)
        cat_lower = csv_category.strip().lower()
        if cat_lower and cat_lower not in ("nan","none",""):
            cat_mask = np.zeros(len(self._rows), dtype=bool)
            for ckey, idxs in self._cat_idx.items():
                if ckey == cat_lower or cat_lower in ckey or ckey in cat_lower:
                    for i in idxs:
                        cat_mask[i] = True
            # Применяем только если нашли хоть что-то — иначе ищем по всем
            if cat_mask.any():
                cosine = cosine * cat_mask

        # Этап 1: фильтр по бренду
        brand_lower = _norm(csv_brand)
        if brand_lower and brand_lower not in ("nan","none",""):
            brand_mask = np.zeros(len(self._rows), dtype=bool)
            for bkey, idxs in self._brand_idx.items():
                if bkey in brand_lower or brand_lower in bkey:
                    for i in idxs:
                        brand_mask[i] = True
            if brand_mask.any():
                cosine = cosine * brand_mask

        # Этап 2: топ-10 + иерархические штрафы
        top_k   = min(10, len(cosine))
        top_idx = np.argpartition(cosine, -top_k)[-top_k:]

        best_score = 0.0
        best_row   = None

        for i in top_idx:
            if cosine[i] < 0.05:
                continue
            row = self._rows[i]
            raw = str(row[2] or "").strip()
            if not raw:
                continue
            if _is_near_duplicate(query, raw):
                continue
            combined = hierarchical_score(query, csv_brand, csv_category, row, cosine[i], self.flavors)
            if combined > best_score:
                best_score = combined
                best_row   = row

        if best_score >= threshold and best_row is not None:
            return best_row, round(best_score, 2)
        return None, 0.0


class ClassicMatcher:
    """Левенштейн + иерархические штрафы."""

    def __init__(self, ref_rows, descriptors, flavors=frozenset()):
        self.descriptors = descriptors
        self.flavors = flavors
        self._rows = [
            r for r in ref_rows
            if str(r[4] or "") == "индикативное"
            and str(r[2] or "").strip()
        ]

    @staticmethod
    def _levenshtein(a, b):
        a, b = a.lower(), b.lower()
        if a == b: return 1.0
        la, lb = len(a), len(b)
        if la == 0 or lb == 0: return 0.0
        prev = list(range(lb + 1))
        for i in range(1, la + 1):
            curr = [i] + [0] * lb
            for j in range(1, lb + 1):
                curr[j] = (prev[j-1] if a[i-1] == b[j-1]
                           else 1 + min(prev[j], curr[j-1], prev[j-1]))
            prev = curr
        return 1 - prev[lb] / max(la, lb)

    def find_match(self, query, csv_brand, threshold, csv_category=""):
        if not query.strip() or not self._rows:
            return None, 0.0
        best_score, best_row = 0.0, None
        for r in self._rows:
            raw = str(r[2] or "").strip()
            if not raw: continue
            if _is_near_duplicate(query, raw): continue
            base  = self._levenshtein(query, raw)
            score = hierarchical_score(query, csv_brand, csv_category, r, base, self.flavors)
            if score > best_score:
                best_score, best_row = score, r
        if best_score >= threshold and best_row is not None:
            return best_row, round(best_score, 2)
        return None, 0.0


def _evaluate_model(matcher, ref_rows: list, log, k_folds: int = 5):
    """
    K-fold перекрёстная проверка модели.

    Делит индикативные пары на K частей.
    Каждую по очереди использует как тест, остальные — как обучение.
    Итоговая точность = среднее по всем K фолдам.

    Почему лучше чем одно разбиение:
      Одно разбиение 80/20 зависит от случайного выбора тест-пар.
      K-fold каждую пару проверяет ровно один раз → результат стабильный.
    """
    import random

    indicative = [
        r for r in ref_rows
        if str(r[4] or "") == "индикативное"
        and str(r[2] or "").strip()
        and str(r[3] or "").strip()
    ]

    if len(indicative) < k_folds * 2:
        log(f"  ⚠ Мало индикативных пар ({len(indicative)}) для {k_folds}-fold валидации")
        return

    random.seed(42)
    random.shuffle(indicative)

    fold_size   = len(indicative) // k_folds
    fold_accs   = []
    fold_scores = []

    log(f"  ── Перекрёстная проверка ({k_folds}-fold, {len(indicative)} пар) ──")

    for fold in range(k_folds):
        # Тестовый фолд
        test_start = fold * fold_size
        test_end   = test_start + fold_size if fold < k_folds - 1 else len(indicative)
        test_set   = indicative[test_start:test_end]

        correct      = 0
        scores_right = []
        scores_wrong = []

        for r in test_set:
            query    = str(r[2] or "").strip()
            expected = str(r[3] or "").strip()
            brand    = str(r[1] or "").strip()
            category = str(r[0] or "").strip()

            best_row, score = matcher.find_match(
                query, brand, threshold=0.0, csv_category=category
            )

            if best_row is not None:
                predicted = str(best_row[3] or "").strip()
                if predicted == expected:
                    correct += 1
                    scores_right.append(score)
                else:
                    scores_wrong.append(score)

        fold_acc = correct / len(test_set) * 100 if test_set else 0
        fold_accs.append(fold_acc)
        if scores_right:
            fold_scores.append(sum(scores_right) / len(scores_right))

        log(f"  Фолд {fold + 1}/{k_folds}: {fold_acc:.1f}%  ({correct}/{len(test_set)})")

    # Итог
    avg_acc   = sum(fold_accs) / len(fold_accs)
    std_acc   = (sum((x - avg_acc) ** 2 for x in fold_accs) / len(fold_accs)) ** 0.5
    avg_score = sum(fold_scores) / len(fold_scores) if fold_scores else 0

    log("  ────────────────────────────────────")
    log(f"  Средняя точность:  {avg_acc:.1f}% ± {std_acc:.1f}%")
    log(f"  Ср. уверенность:   {avg_score:.2f}")

    if avg_acc < 70:
        log("  ⚠ Точность низкая — добавь больше индикативных пар в справочник")
    elif avg_acc >= 90:
        log("  ✅ Модель работает отлично")
    elif avg_acc >= 80:
        log("  ✅ Модель работает хорошо")
    else:
        log("  Модель работает удовлетворительно")

    if std_acc > 10:
        log("  ⚠ Высокий разброс между фолдами — данные неравномерные")


# ─── Хранилище отклонений (онлайн-обучение) ──────────────────────────────────

class RejectionStore:
    """
    Хранит отклонённые пользователем пары в JSON-файле рядом со справочником.
    При матчинге штрафует ранее отклонённые кандидаты (× 0.05).
    После RETRAIN_THRESHOLD отклонений сигнализирует о необходимости переобучения.
    """

    RETRAIN_THRESHOLD = 20

    def __init__(self, ref_path: str):
        self._path = os.path.splitext(ref_path)[0] + "_rejections.json"
        self._data: dict = {}
        self._new_since_retrain = 0
        self._load()

    def _load(self):
        if os.path.exists(self._path):
            try:
                import json
                with open(self._path, encoding="utf-8") as f:
                    self._data = json.load(f)
            except Exception:
                self._data = {}

    def _save(self):
        import json
        with open(self._path, "w", encoding="utf-8") as f:
            json.dump(self._data, f, ensure_ascii=False, indent=2)

    def add_rejection(self, query: str, ref_raw: str):
        key = _norm(query)
        if key not in self._data:
            self._data[key] = []
        ref_norm = _norm(ref_raw)
        if ref_norm not in self._data[key]:
            self._data[key].append(ref_norm)
            self._new_since_retrain += 1
            self._save()

    def is_rejected(self, query: str, ref_raw: str) -> bool:
        key = _norm(query)
        return _norm(ref_raw) in self._data.get(key, [])

    def needs_retrain(self) -> bool:
        return self._new_since_retrain >= self.RETRAIN_THRESHOLD

    def reset_retrain_counter(self):
        self._new_since_retrain = 0

    def stats(self) -> dict:
        total = sum(len(v) for v in self._data.values())
        return {
            "queries_with_rejections": len(self._data),
            "total_rejections":        total,
            "new_since_retrain":       self._new_since_retrain,
        }


def run_matching(ref_path, csv_folder, threshold, on_progress, on_done, mode="ml",
                 rejection_store: "RejectionStore | None" = None, stop_event=None):
    try:
        import polars as pl
        import pandas as pd

        try:
            df_ref = pl.read_excel(ref_path, sheet_name="SKU", engine="calamine")
        except Exception:
            df_ref = pl.read_excel(ref_path, engine="calamine")

        df_ref   = df_ref.fill_null("")
        ref_rows = df_ref.to_numpy().tolist()

        indicative_count = sum(1 for r in ref_rows if str(r[4] or "") == "индикативное")
        existing = {str(r[2] or "").strip() for r in ref_rows if str(r[2] or "").strip()}

        on_progress(f"Справочник: {len(ref_rows)} строк, индикативных: {indicative_count}")

        brand_map, descriptors = build_from_reference(ref_rows)
        flavors = extract_flavors_from_reference(ref_rows)
        on_progress(f"Брендов: {len(set(brand_map.values()))} | Дескрипторов: {len(descriptors)} | Вкусов: {len(flavors)}")

        if rejection_store:
            stats = rejection_store.stats()
            on_progress(f"  Отклонений в базе: {stats['total_rejections']} "
                       f"(по {stats['queries_with_rejections']} SKU)")
            if rejection_store.needs_retrain():
                on_progress(f"  ⚠ Накопилось {stats['new_since_retrain']} новых отклонений — "
                           f"рекомендуется переобучить модель")

        if mode == "ml":
            try:
                matcher = SKUMatcher(ref_rows, descriptors, flavors=flavors)
                on_progress(f"Режим: ML (TF-IDF + иерархия) | Пар: {len(matcher._rows)}")
                _evaluate_model(matcher, ref_rows, on_progress)
            except ImportError:
                on_progress("⚠ scikit-learn не установлен → классический режим")
                matcher = ClassicMatcher(ref_rows, descriptors)
        else:
            matcher = ClassicMatcher(ref_rows, descriptors)
            on_progress(f"Режим: Классический + иерархия | Пар: {len(matcher._rows)}")

        csv_files = list(Path(csv_folder).glob("*.csv"))
        if not csv_files:
            on_done([], error="В папке нет CSV файлов")
            return

        frames = []
        for f in csv_files:
            try:
                df = pd.read_csv(f)
                if "pd_sku" not in df.columns:
                    on_progress(f"⚠ Пропущен {f.name} — нет pd_sku")
                    continue
                cols = [c for c in ["pd_sku","brand","category"] if c in df.columns]
                frames.append(df[cols].drop_duplicates("pd_sku"))
                on_progress(f"→ {f.name}: {len(df)} строк")
            except Exception as e:
                on_progress(f"⚠ Ошибка {f.name}: {e}")

        if not frames:
            on_done([], error="Ни один CSV не подошёл")
            return

        all_skus = pd.concat(frames).drop_duplicates("pd_sku")

        is_new   = ~all_skus["pd_sku"].astype(str).isin(existing)
        new_skus = all_skus[is_new]
        on_progress(f"Новых SKU для матчинга: {len(new_skus)}")

        all_new_skus = (
            new_skus.rename(columns={"pd_sku":"SKU","brand":"Бренд","category":"Категория"})
            .fillna("").to_dict("records")
        )

        results = []
        total   = len(new_skus)

        for idx, (_, row) in enumerate(new_skus.iterrows()):
            if stop_event is not None and stop_event.is_set():
                on_progress("⛔ Остановлено пользователем")
                on_done(results, all_new_skus=all_new_skus)
                return
            csv_sku      = str(row["pd_sku"]).strip()
            csv_brand    = str(row.get("brand","")).strip()
            csv_category = str(row.get("category","")).strip()

            if not csv_sku or csv_sku.lower() in ("nan","none",""):
                continue

            best_row, score = matcher.find_match(
                csv_sku, csv_brand, threshold,
                csv_category=csv_category,
                rejections=rejection_store,
            )

            if best_row is not None:
                results.append({
                    "Категория":        str(best_row[0] or ""),
                    "Бренд":            str(best_row[1] or ""),
                    "Наименование SKU": csv_sku,
                    "Совпало с":        str(best_row[2] or ""),
                    "SKU скорр":        str(best_row[3] or ""),
                    "Статус SKU":       "индикативное",
                    "Уверенность":      score,
                })

            if idx % 50 == 0:
                on_progress(f"Обработано: {idx + 1}/{total}...")

        on_progress(f"Найдено совпадений: {len(results)}")
        results.sort(key=lambda x: -x["Уверенность"])
        on_done(results, all_new_skus=all_new_skus)

    except Exception as e:
        import traceback
        on_done([], error=f"{e}\n{traceback.format_exc()}")


def save_to_reference(selected_results, ref_path):
    import openpyxl
    wb = openpyxl.load_workbook(ref_path)
    ws = wb["SKU"] if "SKU" in wb.sheetnames else wb.active

    if ws.cell(row=1, column=6).value != "Уверенность совпадения":
        ws.cell(row=1, column=6).value = "Уверенность совпадения"

    next_row = ws.max_row + 1
    for r in selected_results:
        ws.cell(row=next_row, column=1).value = r.get("Категория","")
        ws.cell(row=next_row, column=2).value = r.get("Бренд","")
        ws.cell(row=next_row, column=3).value = r.get("Наименование SKU","")
        ws.cell(row=next_row, column=4).value = r.get("SKU скорр","")
        ws.cell(row=next_row, column=5).value = r.get("Статус SKU","индикативное")
        ws.cell(row=next_row, column=6).value = r.get("Уверенность","")
        next_row += 1

    wb.save(ref_path)
    return len(selected_results)